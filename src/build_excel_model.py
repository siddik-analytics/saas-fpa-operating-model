"""Builds the generated Excel operating model from the committed marts.

    python -m src.build_excel_model

Writes `build/generated/Helio_SaaS_FP&A_Operating_Model_generated.xlsx`.

IT DOES NOT WRITE `excel/Helio_SaaS_FP&A_Operating_Model.xlsx`. That path holds the REVIEWED
portfolio workbook - the same model after a presentation review carried out in native Excel
through its COM object model: a rebuilt Executive Summary, charts moved into the content
columns and re-typed, tables demoted where a chart already carries the message, and the
per-sheet geometry brought inside a one-screen ceiling. None of that presentation work lives
in this builder, so a build must not be allowed to overwrite it. The two are the same numbers
in different clothes; `docs/excel_operating_model.md` sets out which is which.

Phase 9. The workbook is the financial-management interface over the Phase 3-8 analytical
stack: eleven presentation tabs a recruiter, hiring manager, FP&A leader or CFO can read
without opening a SQL file, plus nine hidden supporting data sheets holding the Excel Tables
every presentation formula reads.

Division of labour, enforced by design and asserted by `src/validate_excel_model.py`:

* SQL owns every business calculation. ARR movement classification, retention, capacity and
  the pipeline constraint, the forecast drivers, the bridges, materiality and polarity, the
  commentary text, billings, deferred revenue and ASC 340-40 all stay in `sql/`.
* Excel owns presentation arithmetic only: variance, variance %, favourable / unfavourable
  from the Phase 7 polarity value, subtotals, bridge running balances, the scenario and
  segment lookups, and the control roll-up.

The build fails loudly if a required mart or column is missing. It never writes a blank tab.

`openpyxl` does not calculate formulas, so no formula cell carries a cached result. The
workbook is written with full-calculation-on-load set, and every formula is checked two ways
by the validation module: the formula string is checked structurally, and the value it should
produce is recomputed independently in Python from the mart.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from . import excel_data as ed
from . import excel_style as st
from .config import REPO_ROOT
from .excel_style import P, Reference

WORKBOOK_VERSION = "1.0"
# Where the builder writes. Gitignored: it is a build artefact, not a deliverable.
GENERATED_PATH = (REPO_ROOT / "build" / "generated"
                  / "Helio_SaaS_FP&A_Operating_Model_generated.xlsx")
# The reviewed workbook that ships. Written by hand through Excel COM, never by this module.
PUBLISHED_PATH = REPO_ROOT / "excel" / "Helio_SaaS_FP&A_Operating_Model.xlsx"
OUTPUT_PATH = GENERATED_PATH

COL = st.CONTENT_COL            # every presentation sheet writes from column B

COMPANY = "Helio Systems, Inc."
SUBTITLE = "Q2 FY2026 Reforecast"
REPORTING_LINE = "Reporting date: 30 June 2026"

VISIBLE_SHEETS = [
    "Executive Summary", "ARR & Retention", "GTM", "Forecast", "P&L", "Budget Bridge",
    "Scenarios", "Runway & Hiring", "Accounting", "Assumptions", "Controls",
]
DATA_SHEETS = [
    "Data_ARR", "Data_Retention", "Data_GTM", "Data_PnL", "Data_Bridge", "Data_Scenario",
    "Data_Runway", "Data_Accounting", "Data_Commentary", "Chart_Data",
]

# One sheet holds every chart's source block. Charts read nothing else: not a presentation
# cell, not a formula, not a column of a wider mart extract that happens to sit at the right
# offset. See `build_chart_data` for what each block contains.
CHART_SHEET = "Chart_Data"


# ---------------------------------------------------------------------------
# Table registry -- how a presentation formula finds a column
# ---------------------------------------------------------------------------
@dataclass
class TableRef:
    """A written Excel Table: where it lives and which column sits where."""

    sheet: str
    name: str
    left: int
    first_row: int
    last_row: int
    columns: list[str]

    def col(self, name: str) -> int:
        return self.left + self.columns.index(name)

    def letter(self, name: str) -> str:
        return get_column_letter(self.col(name))

    def ref(self, ws: Worksheet, column: str, *, include_header: bool = False) -> Reference:
        index = self.col(column)
        return Reference(
            ws, min_col=index, max_col=index,
            min_row=self.first_row - (1 if include_header else 0), max_row=self.last_row)


class Context:
    """Everything the sheet builders share: the workbook, the frames, the table registry."""

    def __init__(self, wb: Workbook, marts: dict[str, pd.DataFrame], stamp: str):
        self.wb = wb
        self.marts = marts
        self.stamp = stamp
        self.tables: dict[str, TableRef] = {}
        self.blocks: dict[str, st.ChartBlock] = {}
        self.frames: dict[str, pd.DataFrame] = {}
        self.headline = ed.headline(marts)

    def add(
        self, sheet: str, name: str, frame: pd.DataFrame, *, top: int = 1, left: int = 1,
        formats: dict[str, str] | None = None) -> TableRef:
        ws = self.wb[sheet]
        first, last = st.write_table(
            ws, frame, name=name, top=top, left=left, formats=formats
        )
        ref = TableRef(sheet, name, left, first, last, [str(c) for c in frame.columns])
        self.tables[name] = ref
        self.frames[name] = frame
        return ref

    def table(self, name: str) -> TableRef:
        return self.tables[name]

    def sheet(self, name: str) -> Worksheet:
        return self.wb[name]


# ---------------------------------------------------------------------------
# Small formula helpers -- one place where structured references are spelled
# ---------------------------------------------------------------------------
def xlookup(key: str, table: str, lookup_col: str, return_col: str, *, if_missing: str | None = None) -> str:
    tail = "," + if_missing if if_missing is not None else ""
    return "XLOOKUP(" + key + "," + table + "[" + lookup_col + "]," + table + "[" + return_col + "]" + tail + ")"


def fav_unfav_formula(variance_cell: str, label_cell: str, table: str) -> str:
    """Favourable / unfavourable from the Phase 7 polarity value -- never re-derived ad hoc.

    Written as plain nested `IF` over `XLOOKUP` rather than with `LET`. `LET` would let the
    polarity lookup be named once instead of repeated, but it declares a name, and a declared
    name is a second OOXML namespace (`_xlpm.`) to get right on top of `_xlfn.` -- a real
    interoperability risk for a formula whose only job is to print one of four words. Repeating
    an `XLOOKUP` against a fourteen-row table costs nothing. The behaviour is identical:

        contextual            -> "n/a"      (headcount-like metrics are never labelled)
        variance = 0          -> "-"
        polarity matches sign -> "Favorable"
        otherwise             -> "Unfavorable"
    """
    polarity = xlookup(label_cell, table, "line_item", "polarity")
    return (
        "=IF(" + polarity + '="contextual","n/a",'
        "IF(" + variance_cell + '=0,"-",'
        "IF((" + polarity + '="higher_favorable")=(' + variance_cell + '>0),'
        '"Favorable","Unfavorable")))'
    )


# ---------------------------------------------------------------------------
# Supporting data sheets
# ---------------------------------------------------------------------------
USD = "usd_plain"
USD2 = "usd_cents"


def build_data_sheets(ctx: Context) -> None:
    marts = ctx.marts

    # --- Data_ARR -------------------------------------------------------
    arr = ed.arr_monthly(marts, start=date(2025, 1, 31), end=ed.FY2026_END)
    ctx.add(
        "Data_ARR", "tbl_arr_monthly", arr, formats={
            "month_end_date": "date_month", "beginning_arr": USD, "new_logo_arr": USD,
            "expansion_arr": USD, "reactivation_arr": USD, "contraction_arr": USD,
            "churn_arr": USD, "net_new_arr": USD, "ending_arr": USD,
        })
    ctx.add(
        "Data_ARR", "tbl_segment_arr", ed.segment_arr_view(marts), top=1, left=15, formats={
            "jun_2026_actual_arr": USD, "dec_2026_base_arr": USD, "h2_2026_net_new_arr": USD,
            "share_of_dec_2026_arr": "pct",
        })

    # --- Data_Retention -------------------------------------------------
    ctx.add(
        "Data_Retention", "tbl_retention", ed.retention_ttm(marts), formats={
            "month_end_date": "date_month", "cohort_customers": "count",
            "cohort_beginning_arr": USD, "nrr": "pct", "grr": "pct", "logo_retention": "pct",
        })
    ctx.add(
        "Data_Retention", "tbl_retention_trend", ed.retention_trend(marts), top=1, left=11,
        formats={"month_end_date": "date_month", "NRR": "pct", "GRR": "pct",
                 "Logo retention": "pct"})
    ctx.add(
        "Data_Retention", "tbl_atr", ed.renewal_base_quarterly(marts), top=1, left=17,
        formats={"contracts": "count", "atr_arr": USD})
    atr_pivot = ed.renewal_base_pivot(marts)
    ctx.add(
        "Data_Retention", "tbl_atr_pivot", atr_pivot, top=1, left=22,
        formats={c: USD for c in atr_pivot.columns if c != "renewal_quarter"})

    # --- Data_GTM -------------------------------------------------------
    ctx.add(
        "Data_GTM", "tbl_gtm_capacity", ed.gtm_capacity_snapshot(marts), formats={
            "quota_carrying_reps": "count", "expected_attainment": "pct",
            "new_logo_share_of_bookings": "pct", "theoretical_quota_capacity": USD,
            "blended_productive_capacity": USD, "new_logo_productive_capacity": USD,
            "actual_bookings": USD,
        })
    ctx.add(
        "Data_GTM", "tbl_gtm_constraint", ed.gtm_constraint_by_segment(marts), top=1, left=10,
        formats={
            "h2_capacity_supported_arr": USD, "h2_pipeline_supported_arr": USD,
            "h2_constrained_new_logo_arr": USD, "h2_segment_months": "count",
            "h2_pipeline_bound_months": "count", "h2_capacity_bound_months": "count",
            "pipeline_share_of_h2_months": "pct", "budget_new_logo_arr": USD,
            "base_new_logo_arr": USD, "new_logo_arr_variance": USD,
        })
    ctx.add(
        "Data_GTM", "tbl_gtm_monthly", ed.gtm_constraint_monthly(marts), top=1, left=24,
        formats={
            "month_end_date": "date_month", "theoretical_capacity": USD,
            "blended_capacity": USD, "new_logo_capacity": USD,
            "pipeline_supported_bookings": USD, "constrained_new_logo_arr": USD,
        })
    ctx.add(
        "Data_GTM", "tbl_win_rate", ed.win_rate_and_cycle(marts), top=1, left=34, formats={
            "closed_won": "count", "closed_lost": "count", "win_rate": "pct",
            "required_pipeline_per_dollar": "ratio", "median_sales_cycle_days": "days",
            "mean_sales_cycle_days": "days",
        })
    ctx.add(
        "Data_GTM", "tbl_unit_econ", ed.fy2025_unit_economics(marts), top=1, left=42, formats={
            "new_logos": "count", "new_logo_arr": USD, "acquisition_sm_lagged": USD,
            "gross_margin_pct": "pct", "new_logo_arpa": USD, "cac": USD,
            "cac_per_dollar_new_logo_arr": "ratio", "cac_payback_months": "months",
        })
    ctx.add(
        "Data_GTM", "tbl_sales_efficiency", ed.sales_efficiency(marts), top=1, left=52,
        formats={"net_new_arr": USD, "prior_quarter_sm": USD,
                 "net_arr_sales_efficiency": "rate2", "magic_number": "rate2"})
    ctx.add(
        "Data_GTM", "tbl_pipeline", ed.pipeline_by_quarter(marts), top=1, left=58,
        formats={"opportunities": "count", "unweighted_acv": USD, "weighted_acv": USD})

    # --- Data_PnL -------------------------------------------------------
    ctx.add(
        "Data_PnL", "tbl_pnl_summary", ed.pnl_summary(marts), formats={
            "fy2025_actual": USD2, "h1_2026_actual": USD2, "h2_2026_base": USD2,
            "fy2026_base": USD2, "fy2026_budget": USD2,
        })
    ctx.add(
        "Data_PnL", "tbl_pnl_monthly",
        ed.pnl_monthly(marts, start=date(2025, 1, 31), end=ed.FY2026_END), top=1, left=13,
        formats={
            "month_end_date": "date_month", "subscription_revenue": USD,
            "services_revenue": USD, "total_revenue": USD, "subscription_cogs": USD,
            "services_cogs": USD, "total_cogs": USD, "gross_profit": USD,
            "gross_margin_pct": "pct", "sales_marketing": USD, "research_development": USD,
            "general_administrative": USD, "total_opex": USD, "operating_income": USD,
            "operating_margin_pct": "pct",
        })
    ctx.add(
        "Data_PnL", "tbl_headcount_function",
        ed.headcount_by_function(marts, ed.FY2026_END), top=1, left=31,
        formats={"ending_headcount": "fte"})

    # --- Data_Bridge ----------------------------------------------------
    bridge_fmt = {
        "amount": USD2, "running_balance": USD2, "residual": USD2, "chart_invisible": USD,
        "chart_total": USD, "chart_increase": USD, "chart_decrease": USD,
    }
    ctx.add("Data_Bridge", "tbl_arr_bridge", ed.arr_bridge(marts), formats=bridge_fmt)
    ctx.add(
        "Data_Bridge", "tbl_gp_bridge", ed.gross_profit_bridge(marts), top=1, left=14,
        formats=bridge_fmt)
    ctx.add(
        "Data_Bridge", "tbl_opex_bridge", ed.opex_bridge(marts, "Total OpEx"), top=1, left=27,
        formats=bridge_fmt)
    ctx.add(
        "Data_Bridge", "tbl_oi_bridge", ed.operating_income_bridge(marts), top=1, left=40,
        formats=bridge_fmt)
    ctx.add(
        "Data_Bridge", "tbl_rev_bridge", ed.revenue_bridge(marts), top=1, left=51,
        formats={"amount": USD2, "running_balance": USD2, "residual": USD2})
    ctx.add(
        "Data_Bridge", "tbl_mgmt_variance", ed.management_variance(marts), top=1, left=58,
        formats={
            "budget_amount": USD2, "base_amount": USD2, "variance": USD2,
            "variance_pct": "pct2",
        })
    ctx.add(
        "Data_Bridge", "tbl_arr_bridge_segment", ed.arr_bridge_by_segment(marts), top=1, left=72,
        formats={"amount": USD2, "running_balance": USD2, "residual": USD2})
    ctx.add(
        "Data_Bridge", "tbl_opex_bridge_all", ed.opex_bridge_all(marts), top=1, left=81,
        formats={"amount": USD2, "running_balance": USD2, "residual": USD2})
    ctx.add(
        "Data_Bridge", "tbl_gm_bridge", ed.gross_margin_bridge(marts), top=1, left=88,
        formats={"amount": "rate3"})

    # --- Data_Scenario --------------------------------------------------
    ctx.add(
        "Data_Scenario", "tbl_scenario_summary", ed.scenario_summary(marts), formats={
            "dec_2026_exit_arr": USD, "fy2026_revenue": USD, "fy2026_operating_income": USD,
            "dec_2027_exit_arr": USD, "dec_2026_cash": USD, "dec_2027_cash": USD,
            "policy_avg_monthly_burn": USD, "policy_runway_months": "months",
            "headroom_months": "months_signed", "board_runway_floor_months": "months",
            "win_rate_multiplier": "rate2", "attainment_multiplier": "rate2",
            "pipeline_creation_multiplier": "rate2", "retention_severity_multiplier": "rate2",
            "expansion_multiplier": "rate2",
        })
    traj = ed.scenario_trajectory(marts, start=date(2026, 1, 31))
    ctx.add(
        "Data_Scenario", "tbl_scenario_trajectory", traj, top=1, left=19,
        formats={"month_end_date": "date_month", "Bear": USD, "Base": USD, "Bull": USD})
    ctx.add(
        "Data_Scenario", "tbl_scenario_drivers", ed.scenario_drivers(marts), top=1, left=26,
        formats={"Bear": "rate3", "Base": "rate3", "Bull": "rate3"})

    # --- Data_Runway ----------------------------------------------------
    ctx.add(
        "Data_Runway", "tbl_runway_policy", ed.runway_policy(marts), formats={
            "policy_avg_monthly_burn": USD, "opening_cash": USD,
            "policy_runway_months": "months", "board_runway_floor_months": "months",
            "headroom_months": "months_signed",
        })
    ctx.add(
        "Data_Runway", "tbl_hiring", ed.hiring_decision(marts), top=1, left=12, formats={
            "cumulative_hires": "fte", "h2_2026_new_logo_capacity": USD,
            "policy_avg_monthly_burn": USD, "policy_runway_months": "months",
            "board_runway_floor_months": "months", "headroom_months": "months_signed",
            "dec_2027_incremental_arr": USD, "dec_2027_incremental_operating_income": USD,
            "dec_2027_incremental_cash": USD, "dec_2026_incremental_arr": USD,
            "dec_2026_incremental_cash": USD,
        })

    # --- Data_Accounting ------------------------------------------------
    ctx.add(
        "Data_Accounting", "tbl_subscription_accounting", ed.subscription_accounting(marts),
        formats={
            "bookings_tcv": USD, "bookings_acv": USD, "subscription_billings": USD,
            "exit_arr": USD, "contract_analytical_revenue": USD,
            "gl_subscription_revenue": USD, "ending_deferred_revenue": USD,
            "ending_unbilled_receivable": USD,
        })
    ctx.add(
        "Data_Accounting", "tbl_deferred_revenue", ed.deferred_revenue_quarterly(marts),
        top=1, left=11, formats={
            "beginning_deferred_revenue": USD, "billings": USD, "revenue_recognised": USD,
            "unbilled_receivable_movement": USD, "ending_deferred_revenue": USD,
            "ending_unbilled_receivable": USD, "rollforward_residual": USD2,
        })
    ctx.add(
        "Data_Accounting", "tbl_commission", ed.commission_accounting(marts), top=1, left=21,
        formats={
            "beginning_commission_asset": USD, "commission_earned": USD,
            "immediate_expense": USD, "capitalised_commission": USD,
            "commission_amortisation": USD, "gaap_commission_expense": USD,
            "commission_paid_cash": USD, "ending_commission_asset": USD,
        })
    ctx.add(
        "Data_Accounting", "tbl_accounting_adjustment", ed.accounting_adjustment(marts),
        top=1, left=33, formats={
            "commission_accounting_adjustment": USD2, "phase6_total_revenue": USD,
            "adjustment_pct_of_revenue": "pct2", "largest_monthly_adjustment": USD2,
        })

    # --- Data_Commentary ------------------------------------------------
    ctx.add(
        "Data_Commentary", "tbl_commentary", ed.commentary(marts),
        formats={"exec_rank": "count", "commentary_id": "count",
                 "materiality_score": USD2})

    for name in DATA_SHEETS:
        ws = ctx.sheet(name)
        st.data_sheet(ws)
        ws.sheet_state = "hidden"
        st.cell(ws, 1, 1)


# ---------------------------------------------------------------------------
# The chart-data layer
# ---------------------------------------------------------------------------
def _short_bridge_label(line_item: str) -> str:
    """Shorten a bridge line for a category axis without losing which driver it is."""
    text = re.sub(r"\s*\([^)]*\)", "", str(line_item)).strip()
    for suffix in (" ARR variance", " impact", " variance"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
    replacements = {
        "Budget Exit ARR": "Budget", "Base Reforecast Exit ARR": "Base",
        "Budget Gross Profit": "Budget", "Base Gross Profit": "Base",
        "Budget Total OpEx": "Budget", "Base Total OpEx": "Base",
        "Budget Operating Income / (Loss)": "Budget",
        "Base Operating Income / (Loss)": "Base",
        "Opening ARR": "Opening", "Payroll": "Payroll",
        "Non-payroll run-rate": "Non-payroll", "Sales commissions": "Commissions",
        "Revenue variance - Subscription": "Subscription rev",
        "Revenue variance - Services": "Services rev",
        "Subscription COGS": "Subs COGS", "Services COGS": "Svcs COGS",
        "Sales & Marketing OpEx": "S&M", "Research & Development OpEx": "R&D",
        "General & Administrative OpEx": "G&A",
        "Subscription COGS - payroll": "Subs COGS payroll",
        "Subscription COGS - non-payroll": "Subs COGS non-payroll",
        "Services COGS - payroll": "Svcs COGS payroll",
        "Services COGS - non-payroll": "Svcs COGS non-payroll",
        "Revenue": "Revenue",
    }
    return replacements.get(text, text)


def _waterfall_series(frame) -> tuple[list[str], dict[str, list]]:
    """Turn a bridge frame into the four stacked series a waterfall needs.

    A movement of zero is emitted as `None`, not 0.0, so Excel draws neither a bar nor a data
    label for it -- the brief's "do not label every $0 movement", solved in the data rather
    than by fighting Excel's label engine.
    """
    categories, base, anchor, up, down = [], [], [], [], []
    for _, record in frame.iterrows():
        amount = float(record["amount"])
        kind = str(record["line_kind"])
        categories.append(_short_bridge_label(record["line_item"]))
        if kind in ("anchor", "result"):
            base.append(0.0)
            anchor.append(amount)
            up.append(None)
            down.append(None)
        else:
            base.append(float(record["chart_invisible"]))
            anchor.append(None)
            increase = float(record["chart_increase"])
            decrease = float(record["chart_decrease"])
            up.append(increase if increase > 0 else None)
            down.append(decrease if decrease > 0 else None)
    return categories, {"Base": base, "Anchor": anchor, "Increase": up, "Decrease": down}


def build_chart_data(ctx: Context) -> None:
    """Write one compact, purpose-built block per chart, from approved mart values only.

    This adds no calculation. Every number here already exists in a mart or in a supporting
    table on a Data_* sheet; the block reshapes it into the contiguous category-plus-series
    form Excel charts read reliably.
    """
    ws = ctx.sheet(CHART_SHEET)
    marts = ctx.marts
    row = 1

    def add(title, categories, series, fmt="usd_plain"):
        nonlocal row
        block = st.write_chart_block(ws, row, 1, title, categories, series, number_format=fmt)
        ctx.blocks[title] = block
        row = block.last_row + 3
        return block

    # --- Bridges ---------------------------------------------------------
    for key, frame in (
        ("arr_bridge", ed.arr_bridge(marts)),
        ("gp_bridge", ed.gross_profit_bridge(marts)),
        ("opex_bridge", ed.opex_bridge(marts, "Total OpEx")),
        ("oi_bridge", ed.operating_income_bridge(marts)),
    ):
        categories, series = _waterfall_series(frame)
        add(key, categories, series)

    # --- Budget vs Base, monetary metrics only ---------------------------
    # Gross Margin (bps), Ending Headcount (FTE) and runway (months) are deliberately absent:
    # they do not share an axis with dollars.
    variance = marts["fct_management_variance"].set_index("metric")
    monetary = [
        ("Exit ARR", "exit_arr"), ("Revenue", "total_revenue"),
        ("Gross Profit", "gross_profit"), ("Total OpEx", "total_opex"),
    ]
    add(
        "budget_vs_base",
        [label for label, _ in monetary],
        {
            "Board Budget": [float(variance.loc[k, "budget_amount"]) for _, k in monetary],
            "Base reforecast": [float(variance.loc[k, "base_amount"]) for _, k in monetary],
        },
    )

    # --- Scenario Exit ARR ----------------------------------------------
    scenarios = ed.scenario_summary(marts).set_index("scenario")
    add(
        "scenario_exit_arr", list(ed.SCENARIOS),
        {"Dec-2026 Exit ARR": [
            float(scenarios.loc[name, "dec_2026_exit_arr"]) for name in ed.SCENARIOS
        ]},
    )

    # --- Board-policy runway --------------------------------------------
    policy = ed.runway_policy(marts).set_index("path")
    runway_paths = [
        ("Bear", "Bear"), ("Base", "Base"), ("Bull", "Bull"),
        ("Full Capacity-Close", "Base_FullClose"),
    ]
    add(
        "runway", [label for label, _ in runway_paths],
        {
            "Board-policy runway": [
                float(policy.loc[k, "policy_runway_months"]) for _, k in runway_paths
            ],
            "24-month Board floor": [
                float(policy.loc[k, "board_runway_floor_months"]) for _, k in runway_paths
            ],
        },
        fmt="months",
    )

    # --- Monthly ARR movement -------------------------------------------
    arr = ed.arr_monthly(marts, start=date(2025, 1, 31), end=ed.FY2026_END)
    total = arr[arr["segment"] == "Total"].reset_index(drop=True)
    add(
        "arr_movement", list(total["month_label"]),
        {
            "New Logo": [float(v) for v in total["new_logo_arr"]],
            "Expansion": [float(v) for v in total["expansion_arr"]],
            "Reactivation": [float(v) for v in total["reactivation_arr"]],
            "Contraction": [float(v) for v in total["contraction_arr"]],
            "Churn": [float(v) for v in total["churn_arr"]],
        },
    )

    # --- GTM constraint --------------------------------------------------
    constraint = ed.gtm_constraint_by_segment(marts).set_index("segment")
    segments = list(st.SEGMENT_ORDER)
    add(
        "gtm_constraint", segments,
        {
            "New Logo capacity": [
                float(constraint.loc[x, "h2_capacity_supported_arr"]) for x in segments
            ],
            "Pipeline-supported": [
                float(constraint.loc[x, "h2_pipeline_supported_arr"]) for x in segments
            ],
            "Achievable New Logo ARR": [
                float(constraint.loc[x, "h2_constrained_new_logo_arr"]) for x in segments
            ],
        },
    )

    # --- Scenario trajectory ---------------------------------------------
    trajectory = ed.scenario_trajectory(marts, start=date(2026, 1, 31))
    add(
        "scenario_trajectory", list(trajectory["month_label"]),
        {name: [float(v) for v in trajectory[name]] for name in ed.SCENARIOS},
    )

    # --- Hiring: affordable is not attractive ----------------------------
    hiring = ed.hiring_decision(marts).set_index("case_label")
    cases = [c for c in ed.HIRING_CASE_ORDER if c in hiring.index]
    short = {
        "No Incremental GTM Hiring": "No incremental",
        "Targeted / Runway-Constrained Hiring": "Targeted",
        "Full Capacity-Close Hiring": "Full Capacity-Close",
    }
    add(
        "hiring", [short.get(c, c) for c in cases],
        {
            "Incremental Dec-2027 ARR": [
                float(hiring.loc[c, "dec_2027_incremental_arr"]) for c in cases
            ],
            "Incremental cash consumed": [
                -abs(float(hiring.loc[c, "dec_2027_incremental_cash"])) for c in cases
            ],
        },
    )

    st.data_sheet(ws)
    ws.sheet_state = "hidden"
    for column, width in (("A", 30), ("B", 15), ("C", 15), ("D", 15), ("E", 15), ("F", 15)):
        ws.column_dimensions[column].width = width


# ---------------------------------------------------------------------------
# 1. Executive Summary
# ---------------------------------------------------------------------------
# The Executive Summary page grid: one wide label column then nine uniform value columns.
EXEC_LABEL_W = 27.0
EXEC_COL_W = 13.0

EXEC_SCORECARD_ROWS: list[tuple[str, str, str]] = [
    # (metric_label in tbl_mgmt_variance, display period, unit)
    ("Exit ARR", "Dec-2026", "usd"),
    ("New Logo ARR", "FY2026", "usd"),
    ("Revenue", "FY2026", "usd"),
    ("Gross Profit", "FY2026", "usd"),
    ("Gross Margin", "FY2026", "bps"),
    ("Total OpEx", "FY2026", "usd"),
    ("Operating Income / (Loss)", "FY2026", "usd"),
    ("Ending Headcount", "Dec-2026", "fte"),
]


def build_executive_summary(ctx: Context) -> None:
    ws = ctx.sheet("Executive Summary")
    # Ten uniform content columns. KPI cards span two of them with a gutter between, and the
    # scorecard's metric label sits in the first column and overflows the (empty) second, so
    # cards and table rows land on exactly the same vertical lines.
    st.presentation_sheet(
        ws, content=(EXEC_LABEL_W, EXEC_COL_W, EXEC_COL_W, EXEC_COL_W, EXEC_COL_W,
                     EXEC_COL_W, EXEC_COL_W, EXEC_COL_W, EXEC_COL_W, EXEC_COL_W),
        chart_cols=9)
    chart_col = 13  # column M: the shared chart band on this sheet

    row = st.title_block(ws, COMPANY, SUBTITLE, REPORTING_LINE, ctx.stamp)
    h = ctx.headline

    # --- KPI strip ------------------------------------------------------
    row = st.section(ws, row, COL, "Headline position", width=10)
    cards_top = [
        ("Jun-26 ARR actual", h.jun_2026_arr_actual, "usd_m", "Phase 3 ARR engine", None),
        ("Dec-26 Budget ARR", h.dec_2026_budget_arr, "usd_m", "FY2026 Board-Approved", None),
        ("Dec-26 Base ARR", h.dec_2026_base_arr, "usd_m", "Independent Phase 6 build", None),
        ("ARR vs Budget", h.arr_variance, "usd_m_signed",
         "{:.1%} below Budget".format(abs(h.arr_variance_pct)), P.unfavourable),
        ("FY2026 revenue", h.fy2026_revenue, "usd_m", "H1 actual + H2 reforecast", None),
    ]
    cards_bottom = [
        ("FY2026 gross margin", h.fy2026_gross_margin, "pct", "Base reforecast", P.favourable),
        ("FY2026 operating loss", h.fy2026_operating_income, "usd_m", "Base reforecast",
         P.unfavourable),
        ("Dec-26 headcount", h.dec_2026_ending_headcount, "fte", "Expected-value FTE", None),
        ("Base policy runway", h.base_policy_runway_months, "months",
         "{:+.1f} mo of headroom".format(h.base_headroom_months), P.favourable),
        ("Board runway floor", h.board_runway_floor_months, "months", "Governance floor",
         P.grey),
    ]
    # One column per card with a gutter between, so all five fit inside the ten-column content
    # grid and the strip stops short of the chart band. Card text overflows into its own
    # gutter, which is empty, so nothing is clipped.
    row = st.kpi_strip(ws, row, COL, cards_top, span=1, gutter=1)
    row = st.kpi_strip(ws, row, COL, cards_bottom, span=1, gutter=1)
    ws.row_dimensions[row].height = st.R.spacer
    row += 1

    # --- Budget / Base / Variance ---------------------------------------
    row = st.section(
        ws, row, COL, "FY2026 Board Budget vs Base reforecast", width=10,
        note="Units differ by row. Gross Margin shows Budget and Base as percentages and the "
             "variance in basis points -- a display rendering of the stored bps calculation.")
    row = st.table_header(
        ws, row, COL,
        ["Metric", "", "Budget", "Base reforecast", "Variance", "Var %", "Fav / Unfav"])
    first_data = row
    for label, period, unit in EXEC_SCORECARD_ROWS:
        key = '"' + label + '"'
        st.cell(ws, row, COL, label + "  " + period, style="label")
        if unit == "bps":
            st.cell(
                ws, row, COL + 2,
                "=" + xlookup(key, "tbl_mgmt_variance", "metric_label", "budget_amount") + "/10000",
                fmt="pct", style="value")
            st.cell(
                ws, row, COL + 3,
                "=" + xlookup(key, "tbl_mgmt_variance", "metric_label", "base_amount") + "/10000",
                fmt="pct", style="value")
            st.cell(ws, row, COL + 4, "=(E{r}-D{r})*10000".format(r=row), fmt="bps",
                    style="value")
            st.cell(ws, row, COL + 5, "n/m", style="value_muted")
        else:
            fmt = "usd_m" if unit == "usd" else "fte"
            var_fmt = "usd_m_signed" if unit == "usd" else "fte_signed"
            st.cell(
                ws, row, COL + 2,
                "=" + xlookup(key, "tbl_mgmt_variance", "metric_label", "budget_amount"),
                fmt=fmt, style="value")
            st.cell(
                ws, row, COL + 3,
                "=" + xlookup(key, "tbl_mgmt_variance", "metric_label", "base_amount"),
                fmt=fmt, style="value_bold")
            st.cell(ws, row, COL + 4, "=E{r}-D{r}".format(r=row), fmt=var_fmt, style="value")
            st.cell(
                ws, row, COL + 5, '=IF(D{r}=0,"",E{r}/D{r}-1)'.format(r=row),
                fmt="pct_signed", style="value")
        st.cell(
            ws, row, COL + 6,
            "=" + xlookup(key, "tbl_mgmt_variance", "metric_label", "favorable_unfavorable"),
            style="value")
        ws.row_dimensions[row].height = st.R.body
        row += 1
    st.rule_row(ws, row, COL, 7)
    st.variance_conditional_format(ws, "F{a}:G{b}".format(a=first_data, b=row - 1))
    st.fav_unfav_conditional_format(ws, "H{a}:H{b}".format(a=first_data, b=row - 1))
    ws.row_dimensions[row].height = st.R.spacer
    row += 1
    row = st.source_note(
        ws, row, COL,
        "fct_management_variance (Phase 7). Favourable / unfavourable is the mart's own "
        "centralised polarity value, not re-derived here.")

    # --- Management decision panel --------------------------------------
    row = st.section(
        ws, row, COL, "Management decision panel", width=10,
        note="Every value and every read below is a formula over an approved mart. No "
             "conclusion is typed into this workbook.")
    row = st.table_header(
        ws, row, COL, ["Decision question", "", "Evidence", "", "", "Value", "Read"])
    diag = '"Total"'
    full_close = '"Full Capacity-Close Hiring"'
    targeted = '"Targeted / Runway-Constrained Hiring"'
    panel: list[tuple[str, str, str, str, str]] = [
        (
            "What binds New Logo ARR in H2 2026?",
            "H2 segment-months in which pipeline binds, of 18",
            "=" + xlookup(diag, "tbl_gtm_constraint", "segment", "h2_pipeline_bound_months"),
            "count",
            "=" + xlookup(diag, "tbl_gtm_constraint", "segment", "primary_binding_constraint")),
        (
            "Targeted incremental sales hires",
            "Cumulative incremental hires at Dec-2027",
            "=" + xlookup(targeted, "tbl_hiring", "case_label", "cumulative_hires"),
            "fte",
            '=IF(' + xlookup(targeted, "tbl_hiring", "case_label", "cumulative_hires")
            + '=0,"No incremental hiring computed","Incremental hiring computed")'),
        (
            "Full Capacity-Close: affordable?",
            "Board-policy runway headroom vs the 24-month floor",
            "=" + xlookup(full_close, "tbl_hiring", "case_label", "headroom_months"),
            "months_signed",
            '=IF(' + xlookup(full_close, "tbl_hiring", "case_label", "headroom_months")
            + '>=0,"Clears the Board floor","Breaches the Board floor")'),
        (
            "Full Capacity-Close: attractive?",
            "Incremental Dec-2027 ARR per $1 of incremental cash",
            "=IFERROR(" + xlookup(full_close, "tbl_hiring", "case_label", "dec_2027_incremental_arr")
            + "/-" + xlookup(full_close, "tbl_hiring", "case_label", "dec_2027_incremental_cash")
            + ',"")',
            "ratio",
            '=IF(' + xlookup(full_close, "tbl_hiring", "case_label",
                             "dec_2027_incremental_operating_income")
            + '<0,"Incremental operating income still negative at Dec-2027",'
              '"Incremental operating income positive at Dec-2027")'),
        (
            "Base runway vs the Board floor",
            "Board-policy runway headroom, Base",
            "=" + xlookup('"Base"', "tbl_runway_policy", "path", "headroom_months"),
            "months_signed",
            '=IF(' + xlookup('"Base"', "tbl_runway_policy", "path", "headroom_months")
            + '>=0,"Clears the Board floor","Breaches the Board floor")'),
        (
            "Bear runway vs the Board floor",
            "Board-policy runway headroom, Bear",
            "=" + xlookup('"Bear"', "tbl_runway_policy", "path", "headroom_months"),
            "months_signed",
            '=IF(' + xlookup('"Bear"', "tbl_runway_policy", "path", "headroom_months")
            + '>=0,"Clears the Board floor","Breaches the Board floor")'),
    ]
    for question, evidence, value, fmt, read in panel:
        st.cell(ws, row, COL, question, style="label")
        st.cell(ws, row, COL + 2, evidence, style="label_muted")
        st.cell(ws, row, COL + 5, value, fmt=fmt, style="value_bold")
        st.cell(ws, row, COL + 6, read, style="label_bold")
        ws.row_dimensions[row].height = st.R.body
        row += 1
    st.rule_row(ws, row, COL, 10)
    ws.row_dimensions[row].height = st.R.spacer
    row += 1
    row = st.source_note(
        ws, row, COL,
        "fct_new_logo_diagnosis, fct_hiring_scenario, fct_cash_runway_policy (Phases 6-7).")

    # --- Deterministic commentary ---------------------------------------
    commentary_frame = ctx.frames["tbl_commentary"]
    max_items = int(
        ed.load_commentary_config()["commentary"]["max_executive_summary_items"]
    )
    shown = min(max_items, len(commentary_frame))
    row = st.section(
        ws, row, COL, "Management commentary", width=10,
        note="The top {n} items from Phase 7's deterministic commentary engine, in its own "
             "priority-then-materiality order. Rules-based SQL templates -- no language model "
             "was used anywhere in this pipeline.".format(n=shown))
    row = st.table_header(ws, row, COL, ["Priority", "Section", "Headline"], first_left=True)
    for offset in range(1, 3):
        st.cell(ws, row - 1, COL + offset, style="header_left", fill=P.navy)
    st.cell(ws, row - 1, COL + 1, "Section", style="header_left", fill=P.navy)
    st.cell(ws, row - 1, COL + 2, "Headline", style="header_left", fill=P.navy)
    for rank in range(1, shown + 1):
        key = str(rank)
        st.cell(
            ws, row, COL, "=" + xlookup(key, "tbl_commentary", "exec_rank", "priority"),
            style="label_bold")
        st.cell(
            ws, row, COL + 1, "=" + xlookup(key, "tbl_commentary", "exec_rank", "section"),
            style="label_muted")
        st.cell(
            ws, row, COL + 2, "=" + xlookup(key, "tbl_commentary", "exec_rank", "headline"),
            style="label")
        ws.row_dimensions[row].height = st.R.body
        row += 1
    st.rule_row(ws, row, COL, 10)
    ws.row_dimensions[row].height = st.R.spacer
    row += 1
    row = st.source_note(ws, row, COL, "fct_commentary_output (Phase 7).")

    # --- Charts: four, stacked on the shared vertical rhythm -------------
    # Every one reads a purpose-built block on Chart_Data. Nothing here points at a
    # presentation cell or at a formula Excel would have to calculate before it could render.
    charts = ctx.sheet(CHART_SHEET)
    span = st.CHART_ROW_SPAN
    top = 7

    st.waterfall_chart(
        ws, charts, ctx.blocks["arr_bridge"],
        anchor=st.chart_anchor(chart_col, top),
        title="Dec-2026 Exit ARR: Board Budget to Base reforecast ($M)",
    )
    st.column_chart(
        ws, charts, ctx.blocks["budget_vs_base"],
        anchor=st.chart_anchor(chart_col, top + span),
        title="FY2026 Budget vs Base reforecast ($M)",
        colours=[P.grey_light, P.blue], number_format="usd_m", data_labels=True,
    )
    st.column_chart(
        ws, charts, ctx.blocks["scenario_exit_arr"],
        anchor=st.chart_anchor(chart_col, top + span * 2),
        title="Dec-2026 Exit ARR by operating scenario ($M)",
        colours=[P.blue], legend=False, data_labels=True, number_format="usd_m",
        size=st.CHART_COMPACT,
    )
    st.column_chart(
        ws, charts, ctx.blocks["runway"],
        anchor=st.chart_anchor(chart_col, top + span * 3),
        title="Board-policy runway vs the 24-month floor (months)",
        colours=[P.blue, P.unfavourable], number_format="months", data_labels=True,
        size=st.CHART_COMPACT,
    )

    st.print_area(ws, "B1:H{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 2. ARR & Retention
# ---------------------------------------------------------------------------
def build_arr_retention(ctx: Context) -> None:
    ws = ctx.sheet("ARR & Retention")
    st.presentation_sheet(
        ws, content=(11, 15, 14, 13, 13, 13, 13, 13, 14), chart_cols=9,
    )
    chart_col = 12

    row = st.title_block(ws, COMPANY, SUBTITLE + " -- ARR & Retention", REPORTING_LINE, ctx.stamp)

    arr = ctx.frames["tbl_arr_monthly"]
    total = arr[arr["segment"] == "Total"].reset_index(drop=True)

    row = st.section(
        ws, row, 2, "Monthly ARR waterfall -- actual through Jun-2026, Base reforecast after",
        width=9,
        note="Actual months are shaded blue; Base reforecast months amber. Contraction and "
             "churn are stored signed negative, so ending ARR is a plain sum of the seven "
             "columns.")
    headers = ["Month", "Period", "Beginning ARR", "New Logo", "Expansion", "Reactivation",
               "Contraction", "Churn", "Ending ARR"]
    for offset, text in enumerate(headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset < 2 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    first_arr_row = row
    for _, record in total.iterrows():
        actual = record["period_type"] == "Actual"
        fill = P.actual if actual else P.forecast
        ink = P.actual_ink if actual else P.forecast_ink
        st.cell(ws, row, 2, record["month_label"], style="label", fill=fill, colour=ink)
        st.cell(ws, row, 3, record["period_type"], align="left", fill=fill, colour=ink)
        for offset, column in enumerate(
            ["beginning_arr", "new_logo_arr", "expansion_arr", "reactivation_arr",
             "contraction_arr", "churn_arr", "ending_arr"]
        ):
            bold = column in ("beginning_arr", "ending_arr")
            st.cell(
                ws, row, 4 + offset, float(record[column]), fmt="usd_k", align="right",
                bold=bold, fill=fill)
        row += 1
    last_arr_row = row - 1
    # A visible reconciliation of the ARR identity, computed in Excel over mart values. This is
    # a check on the workbook, not a re-implementation of the Phase 3 engine: each month's
    # ending ARR is beginning plus the five movements, so summing the identity down the whole
    # table has to come to zero.
    st.cell(ws, row, 2, "Waterfall check", bold=True, style="label", border="t")
    st.cell(ws, row, 3, "must be zero", style="label_muted", border="t")
    for col in range(4, 10):
        st.cell(ws, row, col, border="t")
    st.cell(
        ws, row, 10,
        "=ROUND(SUM(D{a}:D{b})+SUM(E{a}:E{b})+SUM(F{a}:F{b})+SUM(G{a}:G{b})"
        "+SUM(H{a}:H{b})+SUM(I{a}:I{b})-SUM(J{a}:J{b}),2)".format(
            a=first_arr_row, b=last_arr_row
        ),
        fmt=USD2, bold=True, align="right", border="t")
    row += 2
    row = st.source_note(
        ws, row, 2,
        "fct_arr_forecast, path Base, segment Total (Phase 6); actual months replicate the "
        "Phase 3 fct_arr_waterfall unchanged.")
    row += 1

    # --- Retention ------------------------------------------------------
    row = st.section(
        ws, row, 2, "TTM retention at 30 June 2026", width=9,
        note="Trailing-twelve-month cohort: every customer with ARR above zero exactly twelve "
             "months earlier. GRR is capped per customer before aggregation, so one large "
             "expansion cannot mask another customer's contraction.")
    for offset, text in enumerate(
        ["Segment", "Cohort customers", "Cohort beginning ARR", "NRR", "GRR", "Logo retention"]
    ):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    for segment in ["Total"] + list(st.SEGMENT_ORDER):
        key = '"' + segment + '"'
        month = '"Jun-26"'
        bold = segment == "Total"
        st.cell(ws, row, 2, segment, style="label_bold" if bold else "label")
        for offset, column in enumerate(
            ["cohort_customers", "cohort_beginning_arr", "nrr", "grr", "logo_retention"]
        ):
            fmt = {"cohort_customers": "count", "cohort_beginning_arr": "usd_m"}.get(column, "pct")
            formula = (
                "=XLOOKUP(1,(tbl_retention[segment]=" + key
                + ")*(tbl_retention[month_label]=" + month
                + "),tbl_retention[" + column + '],"")'
            )
            st.cell(ws, row, 3 + offset, formula, fmt=fmt, style="value_bold" if bold else "value")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_retention_ttm (Phase 4).")
    row += 1

    # --- Segment selector ----------------------------------------------
    row = st.section(
        ws, row, 2, "Segment panel", width=9,
        note="A reader convenience. Changing the segment below re-reads this panel only; every "
             "other figure in the workbook is company Total unless its own row says otherwise.")
    st.cell(ws, row, 2, "Segment:", bold=True, style="label")
    selector = st.cell(
        ws, row, 3, "Total", style="input", fill=P.input_fill, align="center",
        border="tblr")
    validation = DataValidation(
        type="list", formula1='"Total,SMB,Mid-Market,Enterprise"', allow_blank=False,
        errorTitle="Segment", error="Choose Total, SMB, Mid-Market or Enterprise.")
    ws.add_data_validation(validation)
    validation.add(selector)
    segment_cell = "$C$" + str(row)
    row += 2
    panel_rows: list[tuple[str, str, str]] = [
        ("Jun-2026 Exit ARR (actual)", "=" + xlookup(
            segment_cell, "tbl_segment_arr", "segment", "jun_2026_actual_arr"), "usd_m"),
        ("Dec-2026 Exit ARR (Base reforecast)", "=" + xlookup(
            segment_cell, "tbl_segment_arr", "segment", "dec_2026_base_arr"), "usd_m"),
        ("H2 2026 net new ARR", "=" + xlookup(
            segment_cell, "tbl_segment_arr", "segment", "h2_2026_net_new_arr"), "usd_m_signed"),
        ("Share of Dec-2026 ARR", "=" + xlookup(
            segment_cell, "tbl_segment_arr", "segment", "share_of_dec_2026_arr"), "pct"),
        ("TTM NRR at Jun-2026", '=XLOOKUP(1,(tbl_retention[segment]=' + segment_cell
         + ')*(tbl_retention[month_label]="Jun-26"),tbl_retention[nrr],"")', "pct"),
        ("TTM GRR at Jun-2026", '=XLOOKUP(1,(tbl_retention[segment]=' + segment_cell
         + ')*(tbl_retention[month_label]="Jun-26"),tbl_retention[grr],"")', "pct"),
        ("TTM logo retention at Jun-2026", '=XLOOKUP(1,(tbl_retention[segment]=' + segment_cell
         + ')*(tbl_retention[month_label]="Jun-26"),tbl_retention[logo_retention],"")', "pct"),
        ("Forward ATR, next four quarters",
         '=SUMIFS(tbl_atr[atr_arr],tbl_atr[segment],' + segment_cell
         + ',tbl_atr[renewal_quarter],"<=2027Q2")', "usd_m"),
    ]
    for label, formula, fmt in panel_rows:
        st.cell(ws, row, 2, label, style="label")
        st.cell(ws, row, 4, formula, fmt=fmt, bold=True, align="right")
        row += 1
    row += 1
    row = st.source_note(
        ws, row, 2, "fct_arr_forecast, fct_retention_ttm, fct_renewal_base (Phases 4 and 6)."
    )
    row += 1

    # --- Renewal outlook ------------------------------------------------
    row = st.section(
        ws, row, 2, "Forward renewal outlook -- available-to-renew by quarter", width=9,
        note="ATR is each in-force contract's ARR measured today, not its book value at "
             "signing. Renewals cluster in Q4 2026 and Q1 2027: real contract-anniversary "
             "seasonality, not a smoothing assumption.")
    atr = ctx.frames["tbl_atr_pivot"]
    headers = ["Renewal quarter"] + [c for c in atr.columns if c != "renewal_quarter"]
    for offset, text in enumerate(headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    for _, record in atr.iterrows():
        st.cell(ws, row, 2, record["renewal_quarter"], style="label")
        for offset, column in enumerate(headers[1:]):
            st.cell(
                ws, row, 3 + offset, float(record[column]), fmt="usd_m", align="right",
                bold=column == "Total")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_renewal_base (Phase 4).")

    # --- Chart -----------------------------------------------------------
    # One chart on this tab. The TTM retention figures read better as the compact table above
    # (a three-metric, four-segment line chart is unreadable at any size), and the ATR
    # seasonality is already obvious in six rows of quarters.
    st.column_chart(
        ws, ctx.sheet(CHART_SHEET), ctx.blocks["arr_movement"],
        anchor=st.chart_anchor(chart_col, 7),
        title="Monthly ARR movement, company total ($000) -- actual to Jun-26, Base after",
        grouping="stacked", number_format="usd_000",
        colours=[P.blue, P.blue_light, "9DC3A6", P.forecast_ink, P.unfavourable],
        size=st.CHART_WIDE, tick_skip=3,
    )
    st.note(
        ws, 7 + st.CHART_ROW_SPAN, chart_col,
        "Bars are the five movement components. Jan-25 to Jun-26 is actual; Jul-26 onward is "
        "the Base reforecast, where New Logo drops to what the pipeline supports.",
    )

    st.print_area(ws, "B1:J{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 3. GTM
# ---------------------------------------------------------------------------
def build_gtm(ctx: Context) -> None:
    ws = ctx.sheet("GTM")
    st.presentation_sheet(ws, content=(44, 15, 15, 15, 15), chart_cols=9)
    chart_col = 8

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- GTM capacity, pipeline and unit economics",
        REPORTING_LINE, ctx.stamp)

    row = st.section(
        ws, row, 2,
        "Capacity does not equal achievable bookings", width=6,
        note="New Logo ARR is LEAST(New Logo productive capacity, pipeline-supported bookings). "
             "Blended capacity credits Expansion and Renewal Uplift work the New Logo target was "
             "never sized against, so it is shown for context and never compared to that target.")
    headers = ["H2 2026, Base reforecast", "SMB", "Mid-Market", "Enterprise", "Total"]
    for offset, text in enumerate(headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    constraint_rows: list[tuple[str, str, str, bool]] = [
        ("New Logo productive capacity (capacity-supported)", "h2_capacity_supported_arr",
         "usd_k", False),
        ("Pipeline-supported bookings", "h2_pipeline_supported_arr", "usd_k", False),
        ("Constrained New Logo ARR = LEAST of the two", "h2_constrained_new_logo_arr",
         "usd_k", True),
        ("Binding constraint", "primary_binding_constraint", "text", False),
        ("Segment-months pipeline-bound (of 6)", "h2_pipeline_bound_months", "count", False),
        ("Segment-months capacity-bound (of 6)", "h2_capacity_bound_months", "count", False),
        ("FY2026 Budget New Logo ARR", "budget_new_logo_arr", "usd_k", False),
        ("FY2026 Base New Logo ARR", "base_new_logo_arr", "usd_k", False),
        ("New Logo ARR variance vs Budget", "new_logo_arr_variance", "usd_k_signed", False),
    ]
    constraint_first = row
    for label, column, fmt, bold in constraint_rows:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, segment in enumerate(list(st.SEGMENT_ORDER) + ["Total"]):
            key = '"' + segment + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_gtm_constraint", "segment", column),
                fmt=fmt if fmt != "text" else None, style="value_bold" if bold else "value")
        row += 1
    st.variance_conditional_format(ws, "C{r}:F{r}".format(r=constraint_first + 8))
    row += 1
    row = st.source_note(
        ws, row, 2,
        "fct_new_logo_diagnosis and int_gtm_capacity_pipeline_forecast (Phases 6-7). Pipeline "
        "binds in 15 of the 18 H2 2026 segment-months -- a result the model produced, not a "
        "target it was tuned to.")
    row += 1

    row = st.section(
        ws, row, 2, "Sales capacity at 30 June 2026", width=6,
        note="Three different capacity measures, never conflated. Expected attainment is the "
             "trailing realised attainment of fully-ramped reps, by segment.")
    capacity_headers = ["At 30 June 2026, monthly"] + headers[1:]
    for offset, text in enumerate(capacity_headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    capacity_rows = [
        ("Quota-carrying reps", "quota_carrying_reps", "count", False),
        ("Theoretical quota capacity (monthly)", "theoretical_quota_capacity", "usd_k", False),
        ("Blended productive capacity (monthly)", "blended_productive_capacity", "usd_k", False),
        ("New Logo productive capacity (monthly)", "new_logo_productive_capacity", "usd_k", True),
        ("Expected attainment", "expected_attainment", "pct", False),
        ("New Logo share of credited bookings", "new_logo_share_of_bookings", "pct", False),
        ("Actual bookings (month)", "actual_bookings", "usd_k", False),
    ]
    for label, column, fmt, bold in capacity_rows:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, segment in enumerate(list(st.SEGMENT_ORDER) + ["Total"]):
            key = '"' + segment + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_gtm_capacity", "segment", column),
                fmt=fmt, style="value_bold" if bold else "value")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_sales_capacity, int_gtm_new_logo_mix (Phase 5).")
    row += 1

    row = st.section(
        ws, row, 2, "Pipeline, conversion and cycle", width=6,
        note="Win rate is Closed Won / (Closed Won + Closed Lost), New Logo only -- open "
             "pipeline is excluded from the denominator. Median cycle is the headline because "
             "the distribution is right-skewed.")
    pipeline_headers = ["All closed New Logo opportunities to date"] + headers[1:]
    for offset, text in enumerate(pipeline_headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    pipeline_rows = [
        ("New Logo win rate (all-time)", "win_rate", "pct"),
        ("Required pipeline per $1 of target", "required_pipeline_per_dollar", "ratio"),
        ("Median sales cycle", "median_sales_cycle_days", "days"),
        ("Closed-won New Logo opportunities", "closed_won", "count"),
        ("Closed-lost New Logo opportunities", "closed_lost", "count"),
    ]
    for label, column, fmt in pipeline_rows:
        st.cell(ws, row, 2, label, style="label")
        for offset, segment in enumerate(list(st.SEGMENT_ORDER) + ["Total"]):
            key = '"' + segment + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_win_rate", "segment", column),
                fmt=fmt, align="right")
        row += 1
    open_pipeline_row = row
    st.cell(ws, row, 2, "Open CRM pipeline ACV at 30 Jun 2026 (unweighted)", style="label")
    st.cell(
        ws, row, 6, "=SUM(tbl_pipeline[unweighted_acv])", fmt="usd_k", style="value_bold")
    row += 1
    st.cell(ws, row, 2, "Open CRM pipeline ACV (probability-weighted)", style="label")
    st.cell(ws, row, 6, "=SUM(tbl_pipeline[weighted_acv])", fmt="usd_k", align="right")
    row += 1
    st.cell(
        ws, row, 2, "Open pipeline coverage of the FY2026 Budget New Logo ARR target",
        style="label")
    st.cell(
        ws, row, 6,
        "=SUM(tbl_pipeline[unweighted_acv])/"
        + xlookup('"Total"', "tbl_gtm_constraint", "segment", "budget_new_logo_arr"),
        fmt="ratio", style="value_bold")
    row += 2
    row = st.source_note(
        ws, row, 2,
        "fct_pipeline_snapshot, int_crm_opportunity_normalized (Phase 5); Budget New Logo ARR "
        "from fct_new_logo_diagnosis (Phase 7).")
    row += 1

    row = st.section(
        ws, row, 2, "SaaS unit economics -- FY2025", width=6,
        note="Period-summed: bookings and cost sum across the year, then divide once. CAC uses "
             "prior-quarter acquisition S&M; payback uses a company-level blended gross margin, "
             "because the source ledger carries no customer-segment dimension on revenue or COGS.")
    unit_headers = ["FY2025", "SMB", "Mid-Market", "Enterprise", "Blended"]
    for offset, text in enumerate(unit_headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    unit_rows = [
        ("New logos acquired", "new_logos", "count", False),
        ("New Logo ARR", "new_logo_arr", "usd_k", False),
        ("New Logo ARPA", "new_logo_arpa", "usd", False),
        ("New-customer CAC", "cac", "usd", True),
        ("CAC per $1 of New Logo ARR", "cac_per_dollar_new_logo_arr", "ratio", False),
        ("Gross margin used in payback", "gross_margin_pct", "pct", False),
        ("CAC payback", "cac_payback_months", "months", True),
    ]
    for label, column, fmt, bold in unit_rows:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, segment in enumerate(list(st.SEGMENT_ORDER) + ["Blended"]):
            key = '"' + segment + '"'
            st.cell(
                ws, row, 3 + offset, "=" + xlookup(key, "tbl_unit_econ", "segment", column),
                fmt=fmt, style="value_bold" if bold else "value")
        row += 1
    row += 1
    efficiency = ctx.frames["tbl_sales_efficiency"]
    latest = efficiency.iloc[-1]
    fy2025 = efficiency[efficiency["fiscal_quarter"].astype(str).str.startswith("2025")]
    st.cell(ws, row, 2, "Net ARR Sales Efficiency (FY2025 average)", style="label")
    st.cell(
        ws, row, 6, float(fy2025["net_arr_sales_efficiency"].mean()), fmt="rate2",
        style="value_bold")
    row += 1
    st.cell(ws, row, 2, "Magic Number, classic (FY2025 average)", style="label")
    st.cell(ws, row, 6, float(fy2025["magic_number"].mean()), fmt="rate2", align="right")
    row += 1
    st.cell(
        ws, row, 2,
        "Net ARR Sales Efficiency (" + str(latest["fiscal_quarter"]) + ")", style="label")
    st.cell(ws, row, 6, float(latest["net_arr_sales_efficiency"]), fmt="rate2", align="right")
    row += 1
    st.note(
        ws, row, 2,
        "The two are a labelled pair, never averaged into one number: Net ARR Sales Efficiency "
        "is ARR-based and forward-leaning; the Magic Number is recognised-revenue-based and lags.")
    row += 2
    row = st.source_note(ws, row, 2, "fct_unit_economics, fct_sales_efficiency (Phase 5).")

    st.column_chart(
        ws, ctx.sheet(CHART_SHEET), ctx.blocks["gtm_constraint"],
        anchor=st.chart_anchor(chart_col, 7),
        title="H2 2026 New Logo ARR by segment ($000): capacity, pipeline, achievable",
        colours=[P.blue_light, P.blue, P.navy], number_format="usd_000", data_labels=True,
        size=st.CHART_WIDE,
    )
    st.note(
        ws, 7 + st.CHART_ROW_SPAN, chart_col,
        "Read left to right within each segment: what the reps could sell, what the funnel "
        "supports, and what the model books. The third bar is the lesser of the first two -- "
        "pipeline sits below capacity in every segment.",
    )

    st.print_area(ws, "B1:G{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 4. Forecast
# ---------------------------------------------------------------------------
def build_forecast(ctx: Context) -> None:
    ws = ctx.sheet("Forecast")
    st.presentation_sheet(ws, freeze_at="D9", content=(32, 15), gutter=False)
    grid = ed.forecast_grid(ctx.marts)
    months: list[str] = grid.attrs["months"]
    period_type: dict[str, str] = grid.attrs["period_type"]
    first_col = 4
    st.uniform_widths(ws, first_col, first_col + len(months) - 1, 10.5)
    st.column_widths(ws, {get_column_letter(first_col + len(months)): 12.5})

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- FY2026 monthly reforecast", REPORTING_LINE, ctx.stamp
    )

    # Actual / Reforecast banner across the month columns. The band is carried by this one
    # row and by a faint tint on the forecast columns only -- the data cells themselves are
    # left unfilled so the numbers stay the loudest thing on the sheet.
    actual_cols = [i for i, m in enumerate(months) if period_type[m] == "Actual"]
    forecast_cols = [i for i, m in enumerate(months) if period_type[m] != "Actual"]
    banner = row
    for index in actual_cols:
        st.cell(ws, banner, first_col + index, fill=P.actual)
    st.cell(
        ws, banner, first_col + actual_cols[0], "JAN-JUN 2026  ACTUAL", fill=P.actual,
        colour=P.actual_ink, style="subsection",
    )
    for index in forecast_cols:
        st.cell(ws, banner, first_col + index, fill=P.forecast)
    st.cell(
        ws, banner, first_col + forecast_cols[0], "JUL-DEC 2026  BASE REFORECAST",
        fill=P.forecast, colour=P.forecast_ink, style="subsection",
    )
    st.cell(
        ws, banner, first_col + len(months), "FY2026", fill=P.band, colour=P.navy,
        style="subsection", align="center",
    )
    ws.row_dimensions[banner].height = st.R.note
    row += 1

    header = row
    st.cell(ws, header, 2, "FY2026 reforecast", style="header_left", fill=P.navy)
    st.cell(ws, header, 3, "$000 unless stated", style="header_left", fill=P.navy)
    for index, month in enumerate(months):
        st.cell(ws, header, first_col + index, month, style="header", fill=P.navy)
    st.cell(
        ws, header, first_col + len(months), "FY2026", style="header", fill=P.navy,
    )
    ws.row_dimensions[header].height = st.R.body
    row += 1

    letters = [get_column_letter(first_col + i) for i in range(len(months))]
    total_letter = get_column_letter(first_col + len(months))
    line_rows: dict[str, int] = {}

    def write_block(title: str, block: str, start_row: int) -> int:
        current = start_row
        st.cell(ws, current, 2, title, style="label_bold",
                fill=P.blue_pale)
        for col in range(3, first_col + len(months) + 1):
            st.cell(ws, current, col, fill=P.blue_pale)
        current += 1
        subset = grid[grid["block"] == block]
        for _, record in subset.iterrows():
            key = str(record["line_key"])
            kind = str(record["line_kind"])
            unit = str(record["unit"])
            fmt = "pct" if unit == "pct" else ("fte" if unit == "fte" else "usd_000")
            bold = kind in ("subtotal", "total")
            border = "t" if kind in ("subtotal", "total") else None
            indent = 1 if kind == "detail" else 0
            st.cell(
                ws, current, 2, str(record["line_item"]), style="label_bold" if bold else "label",
                indent=indent, border=border)
            st.cell(ws, current, 3, "", border=border)
            for index, month in enumerate(months):
                # Only the reforecast half is tinted, and faintly. Banding every actual cell as
                # well turns the grid into wallpaper and hides the numbers.
                fill = None if period_type[month] == "Actual" else P.forecast
                value: Any = float(record[month])
                cell_ref = letters[index]
                if key == "total_revenue":
                    value = "={c}{a}+{c}{b}".format(
                        c=cell_ref, a=line_rows["subscription_revenue"],
                        b=line_rows["services_revenue"])
                elif key == "total_cogs":
                    value = "={c}{a}+{c}{b}".format(
                        c=cell_ref, a=line_rows["subscription_cogs"],
                        b=line_rows["services_cogs"])
                elif key == "gross_profit":
                    value = "={c}{a}-{c}{b}".format(
                        c=cell_ref, a=line_rows["total_revenue"], b=line_rows["total_cogs"])
                elif key == "total_opex":
                    value = "={c}{a}+{c}{b}+{c}{d}".format(
                        c=cell_ref, a=line_rows["sales_marketing"],
                        b=line_rows["research_development"],
                        d=line_rows["general_administrative"])
                elif key == "operating_income":
                    value = "={c}{a}-{c}{b}".format(
                        c=cell_ref, a=line_rows["gross_profit"], b=line_rows["total_opex"])
                elif key == "gross_margin_pct":
                    value = "={c}{a}/{c}{b}".format(
                        c=cell_ref, a=line_rows["gross_profit"], b=line_rows["total_revenue"])
                elif key == "operating_margin_pct":
                    value = "={c}{a}/{c}{b}".format(
                        c=cell_ref, a=line_rows["operating_income"],
                        b=line_rows["total_revenue"])
                st.cell(
                    ws, current, first_col + index, value, fmt=fmt, style="value_bold" if bold else "value",
                    fill=fill, border=border)
            if kind == "balance" and key == "beginning_arr":
                total_value = "={c}{r}".format(c=letters[0], r=current)
            elif kind == "balance":
                total_value = "={c}{r}".format(c=letters[-1], r=current)
            elif key == "gross_margin_pct":
                total_value = "={t}{a}/{t}{b}".format(
                    t=total_letter, a=line_rows["gross_profit"], b=line_rows["total_revenue"])
            elif key == "operating_margin_pct":
                total_value = "={t}{a}/{t}{b}".format(
                    t=total_letter, a=line_rows["operating_income"],
                    b=line_rows["total_revenue"])
            else:
                total_value = "=SUM({f}{r}:{l}{r})".format(
                    f=letters[0], l=letters[-1], r=current)
            st.cell(
                ws, current, first_col + len(months), total_value, fmt=fmt, align="right",
                bold=True, border=border)
            ws.row_dimensions[current].height = st.R.body
            line_rows[key] = current
            current += 1
        ws.row_dimensions[current].height = st.R.spacer
        return current + 1

    row = write_block("ARR waterfall", "ARR", row)
    # ARR identity check -- an Excel control over mart values, not a re-derivation.
    st.cell(ws, row, 2, "Waterfall check (must be zero)", style="label_muted")
    for index in range(len(months)):
        letter = letters[index]
        formula = "=ROUND({c}{b}+{c}{n}+{c}{e}+{c}{r}+{c}{ct}+{c}{ch}-{c}{end},2)".format(
            c=letter, b=line_rows["beginning_arr"], n=line_rows["new_logo_arr"],
            e=line_rows["expansion_arr"], r=line_rows["reactivation_arr"],
            ct=line_rows["contraction_arr"], ch=line_rows["churn_arr"],
            end=line_rows["ending_arr"])
        st.cell(ws, row, first_col + index, formula, fmt=USD2, style="label_muted",
                align="right")
    row += 2

    row = write_block("Profit and loss", "P&L", row)
    row = write_block("Headcount", "Headcount", row)

    st.note(
        ws, row, 2,
        "FY2027 is a forward runway projection and is deliberately not shown here -- it is not "
        "part of the FY2026 reforecast. See the Scenarios and Runway & Hiring tabs.")
    row += 1
    row = st.source_note(
        ws, row, 2,
        "fct_arr_forecast, fct_pnl_reforecast, fct_headcount_forecast, path Base (Phase 6). "
        "Detail lines are mart values; subtotals, margins and the FY2026 column are Excel "
        "formulas over them.")

    st.print_area(ws, "B1:{c}{r}".format(c=total_letter, r=row + 1))


# ---------------------------------------------------------------------------
# 5. P&L
# ---------------------------------------------------------------------------
def build_pnl(ctx: Context) -> None:
    ws = ctx.sheet("P&L")
    st.presentation_sheet(
        ws, content=(34, 15, 15, 17, 15, 15, 15, 11, 13), gutter=False,
    )

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- management profit and loss", REPORTING_LINE, ctx.stamp
    )
    row = st.section(
        ws, row, 2, "FY2026 management P&L -- Board Budget vs Base reforecast", width=9,
        note="Variance is Base less Budget. Favourable / unfavourable is derived from the "
             "Phase 7 centralised metric polarity, so a cost under-run reads favourable and a "
             "revenue shortfall reads unfavourable without either being asserted here.")
    # Four blocks, in the order a management P&L is read: revenue, gross profit, operating
    # expense, operating income. A block break is a thin rule and a half-height spacer rather
    # than a heading, so the hierarchy is visible without adding six more lines of text.
    BLOCK_OF = {
        "subscription_revenue": "revenue", "services_revenue": "revenue",
        "total_revenue": "revenue",
        "subscription_cogs": "margin", "services_cogs": "margin", "total_cogs": "margin",
        "gross_profit": "margin", "gross_margin_pct": "margin",
        "sales_marketing": "opex", "research_development": "opex",
        "general_administrative": "opex", "total_opex": "opex",
        "operating_income": "result", "operating_margin_pct": "result",
    }

    headers = [
        "$000 unless stated", "FY2025 actual", "H1 2026 actual", "H2 2026 Base",
        "FY2026 Base", "FY2026 Budget", "Variance", "Var %", "F / U",
    ]
    for offset, text in enumerate(headers):
        st.cell(
            ws, row, COL + offset, text,
            style="header_left" if offset == 0 else "header",
            fill=P.navy if offset < 6 else P.blue, wrap=True, valign="bottom",
        )
    ws.row_dimensions[row].height = st.R.header
    row += 1
    first_data = row
    previous_block = None
    for _, record in ctx.frames["tbl_pnl_summary"].iterrows():
        label = str(record["line_item"])
        key_name = str(record["line_key"])
        kind = str(record["line_kind"])
        unit = str(record["unit"])
        key = '"' + label + '"'
        block = BLOCK_OF.get(key_name, "revenue")
        if previous_block is not None and block != previous_block:
            st.rule_row(ws, row, COL, 9)
            ws.row_dimensions[row].height = st.R.spacer
            row += 1
        previous_block = block

        bold = kind in ("subtotal", "total")
        emphasis = key_name == "operating_income"
        border = "t" if bold else None
        indent = 1 if kind == "detail" else 0
        band = P.band if kind == "total" else None
        if band:
            for col in range(COL, COL + 9):
                st.cell(ws, row, col, fill=band)

        st.cell(
            ws, row, COL, label,
            style="label_bold" if bold else "label", indent=indent, border=border, fill=band,
            colour=P.navy if emphasis else None,
        )
        fmt = "pct" if unit == "pct" else "usd_000"
        for offset, column in enumerate(
            ["fy2025_actual", "h1_2026_actual", "h2_2026_base", "fy2026_base", "fy2026_budget"]
        ):
            st.cell(
                ws, row, COL + 1 + offset,
                "=" + xlookup(key, "tbl_pnl_summary", "line_item", column),
                fmt=fmt, style="value_bold" if bold else "value", border=border, fill=band,
            )
        # The variance group is set off by a left rule, so Budget and Base read as one block
        # and the comparison as another.
        var_border = ("t" if bold else "") + "l"
        if unit == "pct":
            st.cell(
                ws, row, COL + 6, "=(F{r}-G{r})*10000".format(r=row), fmt="bps",
                style="value_bold" if bold else "value", border=var_border, fill=band,
            )
            st.cell(ws, row, COL + 7, "n/m", style="value_muted", border=border, fill=band)
        else:
            st.cell(
                ws, row, COL + 6, "=F{r}-G{r}".format(r=row), fmt="usd_000_signed",
                style="value_bold" if bold else "value", border=var_border, fill=band,
            )
            st.cell(
                ws, row, COL + 7, '=IF(G{r}=0,"",F{r}/G{r}-1)'.format(r=row),
                fmt="pct_signed", style="value", border=border, fill=band,
            )
        st.cell(
            ws, row, COL + 8,
            fav_unfav_formula("H" + str(row), "$B$" + str(row), "tbl_pnl_summary"),
            style="value_bold" if bold else "value", border=border, fill=band,
        )
        ws.row_dimensions[row].height = st.R.body
        row += 1
    st.rule_row(ws, row, COL, 9)
    ws.row_dimensions[row].height = st.R.spacer
    row += 1
    st.variance_conditional_format(ws, "H{a}:I{b}".format(a=first_data, b=row - 1))
    st.fav_unfav_conditional_format(ws, "J{a}:J{b}".format(a=first_data, b=row - 1))
    row = st.source_note(
        ws, row, COL,
        "fct_pnl_reforecast, path Base (Phase 6) and int_budget_reforecast_comparison at source "
        "GL grain (Phase 7). Budget Gross Profit and Operating Income are the Phase 7 bridge "
        "anchors.")
    row += 1

    row = st.section(
        ws, row, 2, "Dec-2026 ending headcount by function -- Base reforecast", width=9,
        note="Budget carries Ending Headcount as one company-level statistical figure with no "
             "functional split, so this table has no Budget column. Fabricating one would "
             "invent a departmental plan the source does not contain.")
    for offset, text in enumerate(["Function", "Ending headcount"]):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    headcount = ctx.frames["tbl_headcount_function"]
    first_hc = row
    for _, record in headcount.iterrows():
        st.cell(ws, row, 2, str(record["function"]), style="label")
        st.cell(ws, row, 3, float(record["ending_headcount"]), fmt="fte", align="right")
        row += 1
    st.cell(ws, row, 2, "Total", bold=True, style="label", border="t")
    st.cell(
        ws, row, 3, "=SUM(C{a}:C{b})".format(a=first_hc, b=row - 1), fmt="fte", bold=True,
        align="right", border="t")
    row += 2
    row = st.source_note(ws, row, 2, "fct_headcount_forecast, path Base (Phase 6).")

    st.print_area(ws, "B1:J{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 6. Budget Bridge
# ---------------------------------------------------------------------------
BRIDGE_BLOCKS: list[tuple[str, str, str, str, str]] = [
    ("tbl_arr_bridge", "Dec-2026 Exit ARR", "fct_arr_budget_bridge",
     "Beginning ARR at 31-Dec-2025 is real, shared actual history and identical on both sides, "
     "so the bridge collapses to the five movement variances.", "usd_m"),
    ("tbl_gp_bridge", "FY2026 Gross Profit", "fct_gross_profit_bridge",
     "Every COGS driver is split payroll versus non-payroll and calculated from the same "
     "mechanics Phase 6 used to build Base -- no 'other' catch-all line.", "usd_m"),
    ("tbl_opex_bridge", "FY2026 Total OpEx", "fct_opex_budget_bridge",
     "Payroll, sales commissions and non-payroll run rate, across all three categories. "
     "Category-level detail is in tbl_opex_bridge_all on Data_Bridge.", "usd_m"),
    ("tbl_oi_bridge", "FY2026 Operating Income / (Loss)", "fct_operating_income_bridge",
     "Every revenue, COGS and OpEx variance signed by its actual effect on profit, so the walk "
     "reconciles end to end.", "usd_m"),
]


def build_budget_bridge(ctx: Context) -> None:
    ws = ctx.sheet("Budget Bridge")
    st.presentation_sheet(ws, content=(52, 17, 17), chart_cols=9)
    chart_col = 6

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- Board Budget to Base reforecast bridges", REPORTING_LINE,
        ctx.stamp)
    st.note(
        ws, row, 2,
        "Running balance is an Excel formula over the mart's own component amounts. The "
        "residual line proves the walk closes: it is the mart's stated Base value less the "
        "workbook's own running balance, and it must be zero.")
    row += 2

    for index, (table_name, title, source, description, fmt) in enumerate(BRIDGE_BLOCKS):
        block_key = table_name.replace("tbl_", "")
        frame = ctx.frames[table_name]
        row = st.section(ws, row, 2, title + " -- Budget to Base", width=3, note=description)
        for offset, text in enumerate(["Bridge line", "Amount", "Running balance"]):
            st.cell(
                ws, row, 2 + offset, text,
                style="header_left" if offset == 0 else "header", fill=P.navy,
                valign="bottom")
        row += 1
        first = row
        previous_running: int | None = None
        result_row = None
        for _, record in frame.iterrows():
            label = str(record["line_item"])
            kind = str(record["line_kind"])
            bold = kind in ("anchor", "result")
            fill = P.band if bold else None
            if bold:
                for col in range(2, 5):
                    st.cell(ws, row, col, fill=P.band)
            st.cell(
                ws, row, 2, label, style="label_bold" if bold else "label", fill=fill,
                indent=0 if bold else 1)
            st.cell(
                ws, row, 3, float(record["amount"]),
                fmt="usd_plain" if bold else "usd_signed", style="value_bold" if bold else "value", fill=fill)
            if kind == "anchor":
                st.cell(ws, row, 4, "=C{r}".format(r=row), fmt="usd_plain", align="right",
                        bold=True, fill=fill)
                previous_running = row
            elif kind == "result":
                st.cell(ws, row, 4, "", fill=fill)
                result_row = row
            else:
                st.cell(
                    ws, row, 4, "=D{p}+C{r}".format(p=previous_running, r=row),
                    fmt="usd_plain", align="right")
                previous_running = row
            row += 1
        st.cell(ws, row, 2, "Residual (must be zero)", style="label_muted",
                border="t")
        st.cell(ws, row, 3, "", border="t")
        st.cell(
            ws, row, 4, "=ROUND(C{res}-D{run},2)".format(res=result_row, run=previous_running),
            fmt=USD2, style="value_muted", border="t")
        row += 1
        row = st.source_note(ws, row, 2, source + " (Phase 7).")
        row += 1

        st.waterfall_chart(
            ws, ctx.sheet(CHART_SHEET), ctx.blocks[block_key],
            anchor=st.chart_anchor(chart_col, 7 + index * st.CHART_ROW_SPAN),
            title=title + ": Budget to Base ($M)", number_format=fmt,
        )

    # Gross margin, in basis points.
    row = st.section(
        ws, row, 2, "FY2026 Gross Margin", width=3,
        note="Reported in basis points, never as a bare percentage-point difference.")
    for _, record in ctx.frames["tbl_gm_bridge"].iterrows():
        unit = str(record["unit"])
        st.cell(ws, row, 2, str(record["line_item"]), style="label",
                bold=unit == "bps", indent=0 if unit == "bps" else 1)
        # fct_gross_profit_bridge stores the two margin levels as fractions and the variance
        # in basis points, so each is written in its own unit rather than converted.
        st.cell(
            ws, row, 3, float(record["amount"]),
            fmt="pct" if unit == "pct" else "bps", align="right", bold=unit == "bps")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_gross_profit_bridge (Phase 7).")
    row += 1

    row = st.section(
        ws, row, 2, "FY2026 Revenue -- Subscription, Services and Total", width=3,
        note="A calculated recognition-mechanic decomposition: the same ARR-lag and "
             "attach-ratio mechanics Phase 6 uses to build Base, run over Budget's own ARR "
             "path. Never a fabricated price-volume split.")
    for offset, text in enumerate(["Revenue line", "Bridge line", "Amount"]):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset < 2 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    revenue = ctx.frames["tbl_rev_bridge"]
    for _, record in revenue.iterrows():
        st.cell(ws, row, 2, str(record["revenue_line"]), style="label_muted")
        st.cell(ws, row, 3, str(record["line_item"]), style="label")
        st.cell(ws, row, 4, float(record["amount"]), fmt="usd_plain", align="right")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_revenue_budget_bridge (Phase 7).")
    row += 1

    st.note(
        ws, row, 2,
        "Segment ARR bridges are in tbl_arr_bridge_segment on Data_Bridge. Budget's five ARR "
        "movement components carry no segment grain in the source, so the Budget side of a "
        "segment bridge is an allocation and is labelled budget_grain = 'allocated'. Base's "
        "segment figures are always segment-native.")
    row += 2

    st.print_area(ws, "B1:D{r}".format(r=row))


# ---------------------------------------------------------------------------
# 7. Scenarios
# ---------------------------------------------------------------------------
SCENARIO_PANEL_ROWS: list[tuple[str, str, str]] = [
    ("Dec-2026 Exit ARR", "dec_2026_exit_arr", "usd_m"),
    ("FY2026 revenue", "fy2026_revenue", "usd_m"),
    ("FY2026 operating income / (loss)", "fy2026_operating_income", "usd_m"),
    ("Dec-2027 Exit ARR", "dec_2027_exit_arr", "usd_m"),
    ("Dec-2026 cash", "dec_2026_cash", "usd_m"),
    ("Dec-2027 cash", "dec_2027_cash", "usd_m"),
    ("Board-policy average monthly burn", "policy_avg_monthly_burn", "usd_k"),
    ("Board-policy runway", "policy_runway_months", "months"),
    ("Headroom vs the 24-month floor", "headroom_months", "months_signed"),
    ("Breaches the Board floor", "breaches_floor", "text"),
]

SCENARIO_DRIVER_ROWS: list[tuple[str, str]] = [
    ("Win rate multiplier", "win_rate_multiplier"),
    ("Attainment multiplier", "attainment_multiplier"),
    ("Pipeline creation multiplier", "pipeline_creation_multiplier"),
    ("Retention severity multiplier", "retention_severity_multiplier"),
    ("Expansion multiplier", "expansion_multiplier"),
]


def build_scenarios(ctx: Context) -> None:
    ws = ctx.sheet("Scenarios")
    st.presentation_sheet(ws, content=(40, 16, 16, 16), chart_cols=9)
    chart_col = 7

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- Bear / Base / Bull operating scenarios", REPORTING_LINE,
        ctx.stamp)
    row = st.section(
        ws, row, 2, "Scenario comparison", width=4,
        note="Base is the Board reforecast. Bear and Bull are management operating scenarios, "
             "not alternative forecasts of record. Incremental hiring is deliberately not a "
             "scenario lever -- it is a separate management-action dimension on the next tab.")
    for offset, text in enumerate([""] + list(st.SCENARIO_ORDER)):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header",
            fill=P.navy if text != "Base" else P.blue, wrap=True, valign="bottom")
    row += 1
    first_data = row
    for label, column, fmt in SCENARIO_PANEL_ROWS:
        bold = column in ("dec_2026_exit_arr", "policy_runway_months")
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, scenario in enumerate(st.SCENARIO_ORDER):
            key = '"' + scenario + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_scenario_summary", "scenario", column),
                fmt=fmt if fmt != "text" else None, style="value_bold" if bold else "value",
                fill=P.band if scenario == "Base" else None)
        row += 1
    st.variance_conditional_format(ws, "C{a}:E{b}".format(a=first_data + 8, b=first_data + 8))
    row += 1
    row = st.source_note(
        ws, row, 2, "fct_scenario_monthly and fct_cash_runway_policy (Phase 6)."
    )
    row += 1

    row = st.section(
        ws, row, 2, "Scenario assumptions -- management levers, not history", width=4,
        note="Five multipliers, each tied to one separately modelled mechanism, applied on top "
             "of the data-derived Base rates. Never a blanket revenue or EBITDA multiplier, and "
             "never presented as derived from actuals.")
    for offset, text in enumerate(["Management lever"] + list(st.SCENARIO_ORDER)):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    for label, column in SCENARIO_DRIVER_ROWS:
        st.cell(ws, row, 2, label, style="label")
        for offset, scenario in enumerate(st.SCENARIO_ORDER):
            key = '"' + scenario + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_scenario_summary", "scenario", column),
                fmt="rate2", style="input", fill=P.input_fill)
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "Amber cells are management assumptions (config/assumptions.yml: "
        "forecast.scenario_multipliers). Every other figure on this tab is calculated output. "
        "Reactivation is deliberately not scenario-varied: it is small and historically noisy.")
    row += 2
    row = st.source_note(
        ws, row, 2, "config/assumptions.yml and int_forecast_drivers (Phase 6)."
    )
    row += 1

    # --- Scenario selector ---------------------------------------------
    row = st.section(
        ws, row, 2, "Scenario summary panel", width=4,
        note="This panel alone follows the selector. Base remains the Board reforecast "
             "everywhere else in the workbook.")
    st.cell(ws, row, 2, "Scenario:", bold=True, style="label")
    selector = st.cell(
        ws, row, 3, "Base", style="input", fill=P.input_fill, align="center",
        border="tblr")
    validation = DataValidation(
        type="list", formula1='"Bear,Base,Bull"', allow_blank=False,
        errorTitle="Scenario", error="Choose Bear, Base or Bull.")
    ws.add_data_validation(validation)
    validation.add(selector)
    scenario_cell = "$C$" + str(row)
    row += 2
    for label, column, fmt in SCENARIO_PANEL_ROWS:
        st.cell(ws, row, 2, label, style="label")
        st.cell(
            ws, row, 4, "=" + xlookup(scenario_cell, "tbl_scenario_summary", "scenario", column),
            fmt=fmt if fmt != "text" else None, bold=True, align="right")
        row += 1
    row += 1
    row = st.source_note(ws, row, 2, "fct_scenario_monthly, fct_cash_runway_policy (Phase 6).")

    # One chart. The Board-policy runway comparison lives on the Executive Summary; repeating
    # it here would be the same four bars twice.
    st.line_chart(
        ws, ctx.sheet(CHART_SHEET), ctx.blocks["scenario_trajectory"],
        anchor=st.chart_anchor(chart_col, 7),
        title="Monthly Exit ARR by scenario, Jan-2026 to Dec-2027 ($M)",
        colours=[st.SCENARIO_COLOURS["Bear"], st.SCENARIO_COLOURS["Base"],
                 st.SCENARIO_COLOURS["Bull"]],
        number_format="usd_m", size=st.CHART_WIDE, tick_skip=3,
    )
    st.note(
        ws, 7 + st.CHART_ROW_SPAN, chart_col,
        "Quarter-spaced labels; every month is still plotted. The three paths separate from "
        "Jul-2026, the first forecast month -- everything to the left is shared actual history.",
    )

    st.print_area(ws, "B1:E{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 8. Runway & Hiring
# ---------------------------------------------------------------------------
def build_runway_hiring(ctx: Context) -> None:
    ws = ctx.sheet("Runway & Hiring")
    st.presentation_sheet(ws, content=(46, 18, 18, 18, 15, 13), chart_cols=9)
    chart_col = 9

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- runway and the incremental hiring decision",
        REPORTING_LINE, ctx.stamp)
    st.note(
        ws, row, 2,
        "Two questions, answered separately and never collapsed into one: can the company "
        "afford a hiring case against the Board's runway floor, and is that case a good use of "
        "the spend. Affordable does not mean attractive.")
    row += 2

    cases = list(ctx.frames["tbl_hiring"]["case_label"])
    header = ["Hiring case"] + [c.replace(" Hiring", "") for c in cases]

    # --- A. Affordability ----------------------------------------------
    row = st.section(
        ws, row, COL, "A.  AFFORDABILITY  --  Board-policy runway vs the 24-month floor",
        width=6, accent=P.blue,
        note="The Board-policy view anchors on the approved FY2027 average monthly burn and "
             "moves it only by each case's own modelled burn delta. The model-derived operating "
             "cash proxy supplies deltas, never the level -- it carries no working capital, no "
             "capex and no cash-flow-statement adjustments beyond one D&A add-back.")
    for offset, text in enumerate(header):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    affordability = [
        ("Incremental hires", "cumulative_hires", "fte", False),
        ("Board-policy average monthly burn", "policy_avg_monthly_burn", "usd_k", False),
        ("Board-policy runway", "policy_runway_months", "months", True),
        ("Board runway floor", "board_runway_floor_months", "months", False),
        ("Headroom vs the floor", "headroom_months", "months_signed", True),
        ("Breaches the floor", "breaches_floor_flag", "text", False),
    ]
    headroom_row = None
    for label, column, fmt, bold in affordability:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, case in enumerate(cases):
            key = '"' + case + '"'
            st.cell(
                ws, row, 3 + offset, "=" + xlookup(key, "tbl_hiring", "case_label", column),
                fmt=fmt if fmt != "text" else None, style="value_bold" if bold else "value")
        if column == "headroom_months":
            headroom_row = row
        row += 1
    if headroom_row:
        st.variance_conditional_format(ws, "C{r}:E{r}".format(r=headroom_row))
    row += 1
    row = st.source_note(ws, row, 2, "fct_cash_runway_policy, fct_hiring_scenario (Phase 6).")
    row += 1

    # --- B. Economic attractiveness -------------------------------------
    row = st.section(
        ws, row, COL, "B.  ECONOMIC ATTRACTIVENESS  --  the FY2027 fuller-ramp horizon",
        width=6, accent=P.forecast_ink,
        note="Hires start 31 Oct 2026, so Dec-2026 is only weeks into ramp and understates the "
             "decision-relevant economics. Dec-2027 is the horizon management should judge the "
             "case on; Dec-2026 is carried below as a near-term ramp snapshot only.")
    for offset, text in enumerate(header):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    attractiveness = [
        ("Incremental Dec-2027 Exit ARR", "dec_2027_incremental_arr", "usd_k_signed", True),
        ("Incremental cumulative cash consumed to Dec-2027", "dec_2027_incremental_cash",
         "usd_k_signed", True),
        ("Incremental operating income, Dec-2027", "dec_2027_incremental_operating_income",
         "usd_k_signed", True),
        ("H2 2026 New Logo productive capacity", "h2_2026_new_logo_capacity", "usd_k", False),
        ("Binding constraint on New Logo ARR", "primary_binding_constraint", "text", False),
    ]
    attractiveness_first = row
    for label, column, fmt, bold in attractiveness:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, case in enumerate(cases):
            key = '"' + case + '"'
            st.cell(
                ws, row, 3 + offset, "=" + xlookup(key, "tbl_hiring", "case_label", column),
                fmt=fmt if fmt != "text" else None, style="value_bold" if bold else "value")
        row += 1
    st.cell(
        ws, row, 2, "Incremental Dec-2027 ARR per $1 of incremental cash",
        style="label_bold")
    for offset, case in enumerate(cases):
        key = '"' + case + '"'
        st.cell(
            ws, row, 3 + offset,
            "=IFERROR(" + xlookup(key, "tbl_hiring", "case_label", "dec_2027_incremental_arr")
            + "/-" + xlookup(key, "tbl_hiring", "case_label", "dec_2027_incremental_cash")
            + ',"n/a")',
            fmt="ratio", style="value_bold")
    row += 1
    st.variance_conditional_format(
        ws, "C{a}:E{b}".format(a=attractiveness_first, b=attractiveness_first + 2)
    )
    row += 1

    row = st.section(
        ws, row, 2, "Near-term Dec-2026 ramp impact -- shown separately, not the decision view",
        width=4)
    for offset, text in enumerate(header):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    for label, column in [
        ("Incremental Dec-2026 Exit ARR", "dec_2026_incremental_arr"),
        ("Incremental cumulative cash consumed to Dec-2026", "dec_2026_incremental_cash"),
    ]:
        st.cell(ws, row, 2, label, style="label_muted")
        for offset, case in enumerate(cases):
            key = '"' + case + '"'
            st.cell(
                ws, row, 3 + offset, "=" + xlookup(key, "tbl_hiring", "case_label", column),
                fmt="usd_signed", style="value_muted")
        row += 1
    row += 1
    row = st.source_note(
        ws, row, 2,
        "fct_hiring_scenario, fct_new_logo_diagnosis (Phases 6-7). Hire counts are computed "
        "from the H2 2026 capacity gap by segment, never picked by hand.")
    row += 1

    row = st.section(ws, row, 2, "Board-policy runway, every path", width=4)
    for offset, text in enumerate(
        ["Path", "View", "Policy burn / month", "Policy runway", "Headroom", "Breaches floor"]
    ):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset < 2 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    policy_first = row
    for _, record in ctx.frames["tbl_runway_policy"].iterrows():
        st.cell(ws, row, 2, str(record["label"]), style="label")
        st.cell(ws, row, 3, str(record["view"]), style="label_muted")
        st.cell(ws, row, 4, float(record["policy_avg_monthly_burn"]), fmt="usd_k", align="right")
        st.cell(ws, row, 5, float(record["policy_runway_months"]), fmt="months", align="right",
                bold=True)
        st.cell(ws, row, 6, float(record["headroom_months"]), fmt="months_signed", align="right")
        st.cell(ws, row, 7, str(record["breaches_floor_flag"]), style="value")
        row += 1
    st.variance_conditional_format(ws, "F{a}:F{b}".format(a=policy_first, b=row - 1))
    row += 1
    row = st.source_note(ws, row, 2, "fct_cash_runway_policy (Phase 6).")

    st.column_chart(
        ws, ctx.sheet(CHART_SHEET), ctx.blocks["hiring"],
        anchor=st.chart_anchor(chart_col, 7),
        title="Dec-2027 incremental ARR against incremental cash consumed ($000)",
        colours=[P.blue, P.unfavourable], number_format="usd_000", data_labels=True,
        size=st.CHART_WIDE,
    )
    st.note(
        ws, 7 + st.CHART_ROW_SPAN, chart_col,
        "Full Capacity-Close buys incremental ARR at several times its size in cash, while "
        "pipeline -- not capacity -- is what binds New Logo ARR. Affordable is not the same "
        "question as attractive.",
    )

    st.print_area(ws, "B1:G{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 9. Accounting
# ---------------------------------------------------------------------------
def build_accounting(ctx: Context) -> None:
    ws = ctx.sheet("Accounting")
    # The deferred-revenue rollforward is the widest block here and runs B:I, so the gutter
    # sits at J and the chart band starts at K.
    st.presentation_sheet(
        ws, content=(46, 15, 15, 17, 15, 15, 17, 13), chart_cols=9,
    )
    chart_col = 11

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- SaaS accounting: deferred revenue and ASC 340-40",
        REPORTING_LINE, ctx.stamp)
    st.note(
        ws, row, 2,
        "A reconciliation layer, not a replacement. It reads the frozen Phase 3-7 output and "
        "the source ledger and writes back to neither. The Board reforecast, the runway "
        "calculation and the hiring decision all continue to run on the Phase 6 P&L.")
    row += 2

    # --- Subscription accounting ---------------------------------------
    subscription = ctx.frames["tbl_subscription_accounting"]
    periods = list(subscription["period"])
    row = st.section(
        ws, row, 2, "Bookings, billings, ARR and revenue are four different metrics", width=5,
        note="Collapsing these into one number is the most common way a SaaS model misleads. "
             "Each row is measured on its own basis from its own model; none is derived from "
             "another.")
    for offset, text in enumerate([""] + periods):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    subscription_rows = [
        ("Bookings -- TCV", "bookings_tcv", False),
        ("Bookings -- ACV", "bookings_acv", False),
        ("Subscription billings", "subscription_billings", False),
        ("Exit ARR", "exit_arr", True),
        ("Contract-level analytical revenue", "contract_analytical_revenue", False),
        ("GL subscription revenue (4000 + 4010)", "gl_subscription_revenue", False),
        ("Ending deferred revenue", "ending_deferred_revenue", True),
        ("Ending unbilled receivable", "ending_unbilled_receivable", False),
    ]
    for label, column, bold in subscription_rows:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, period in enumerate(periods):
            key = '"' + period + '"'
            st.cell(
                ws, row, 3 + offset,
                "=" + xlookup(key, "tbl_subscription_accounting", "period", column),
                fmt="usd_m", style="value_bold" if bold else "value")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "The contract schedule runs roughly 2-3% above the source GL: the ledger recognises a "
        "weighted lag of prior month-end ARR, the schedule recognises the current month's "
        "in-force rate. A difference in recognition convention with a stated cause -- neither "
        "series is corrected toward the other, and Phase 6 is not restated.")
    row += 2
    row = st.source_note(
        ws, row, 2,
        "fct_crm_bookings, fct_billings, fct_arr_forecast, fct_revenue_accounting_reconciliation,"
        " fct_deferred_revenue (Phases 5, 6 and 8).")
    row += 1

    # --- Deferred revenue rollforward -----------------------------------
    row = st.section(
        ws, row, 2, "Deferred revenue rollforward", width=8, accent=P.blue,
        note="Beginning + billings - revenue + unbilled receivable movement = ending. There are "
             "no other lines: no true-up, no rounding line, no plug. The unbilled receivable is "
             "carried separately and is never netted into deferred revenue.")
    deferred_headers = [
        "Quarter", "Beginning DR", "Billings", "Revenue recognised (deducted)",
        "Unbilled movement", "Ending DR", "Ending unbilled receivable", "Residual",
    ]
    for offset, text in enumerate(deferred_headers):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    deferred = ctx.frames["tbl_deferred_revenue"]
    for _, record in deferred.iterrows():
        st.cell(ws, row, 2, str(record["fiscal_quarter"]), style="label")
        for offset, column in enumerate([
            "beginning_deferred_revenue", "billings", "revenue_recognised",
            "unbilled_receivable_movement", "ending_deferred_revenue",
            "ending_unbilled_receivable",
        ]):
            st.cell(
                ws, row, 3 + offset, float(record[column]), fmt="usd_plain", align="right",
                bold=column == "ending_deferred_revenue")
        st.cell(
            ws, row, 9, "=ROUND(C{r}+D{r}-E{r}+F{r}-G{r},2)".format(r=row), fmt=USD2,
            style="value_muted")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "Long-term deferred revenue is structurally zero: the longest billing period in this "
        "contract population is twelve months, so no invoice covers service more than eleven "
        "months beyond a month end. A property of the book, not an assumption.")
    row += 2
    row = st.source_note(ws, row, 2, "fct_deferred_revenue (Phase 8).")
    row += 1

    # --- Commission accounting ------------------------------------------
    commission = ctx.frames["tbl_commission"]
    commission_periods = list(commission["period"])
    row = st.section(
        ws, row, 2, "Sales commission -- earned, cash and GAAP expense (ASC 340-40)", width=6, accent=P.blue,
        note="Three numbers, all correct and all different. 41% of earned commission is "
             "expensed as incurred and 59% capitalised, amortised straight-line over 36 months "
             "-- because renewal commission (3% on uplift alone) is not commensurate with the "
             "9% paid to land, so the initial commission relates to the renewal periods too.")
    for offset, text in enumerate([""] + commission_periods):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    commission_rows = [
        ("Beginning commission asset", "beginning_commission_asset", False),
        ("Commission earned", "commission_earned", False),
        ("Expensed as incurred (41%)", "immediate_expense", False),
        ("Capitalised (59%)", "capitalised_commission", False),
        ("Amortisation of prior cohorts", "commission_amortisation", False),
        ("GAAP commission expense", "gaap_commission_expense", True),
        ("Commission paid in cash", "commission_paid_cash", False),
        ("Ending commission asset", "ending_commission_asset", True),
    ]
    for label, column, bold in commission_rows:
        st.cell(ws, row, 2, label, style="label_bold" if bold else "label")
        for offset, period in enumerate(commission_periods):
            key = '"' + period + '"'
            st.cell(
                ws, row, 3 + offset, "=" + xlookup(key, "tbl_commission", "period", column),
                fmt="usd_plain", style="value_bold" if bold else "value")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "P&L expense reconciled, asset analytically derived. Immediate expense ties to account "
        "6030 and amortisation to account 6040 to the cent in every actual month; the asset is "
        "the arithmetic consequence of those two reconciled flows, because fact_gl_actuals is a "
        "P&L extract and the source carries no balance sheet at all.")
    row += 2
    st.note(
        ws, row, 2,
        "Capitalisation is a timing effect, not a saving: it moves when the expense is "
        "recognised, not whether it is paid. Cash commission is carried on the same rows "
        "precisely so the two cannot be conflated.")
    row += 2
    row = st.source_note(ws, row, 2, "fct_commission_asset, path Base (Phase 8).")
    row += 1

    # --- Adjustment and materiality --------------------------------------
    row = st.section(
        ws, row, 2, "The accounting adjustment to the Phase 6 forecast, and its size", width=5, accent=P.blue,
        note="Zero across every actual month by construction -- in actual months the schedule "
             "reproduces the ledger rather than restating it. In the forecast the adjustment is "
             "the amortisation difference alone: a real cohort rollforward against Phase 6's "
             "flat run rate.")
    for offset, text in enumerate([
        "Period", "Commission accounting adjustment", "Phase 6 total revenue",
        "Adjustment as % of revenue", "Largest single month",
    ]):
        st.cell(
            ws, row, 2 + offset, text, style="header_left" if offset == 0 else "header", fill=P.navy, wrap=True,
            valign="bottom")
    row += 1
    for _, record in ctx.frames["tbl_accounting_adjustment"].iterrows():
        st.cell(ws, row, 2, str(record["period"]), style="label")
        st.cell(ws, row, 3, float(record["commission_accounting_adjustment"]), fmt="usd",
                style="value_bold")
        st.cell(ws, row, 4, float(record["phase6_total_revenue"]), fmt="usd_m", align="right")
        st.cell(ws, row, 5, float(record["adjustment_pct_of_revenue"]), fmt="pct2",
                align="right")
        st.cell(ws, row, 6, float(record["largest_monthly_adjustment"]), fmt=USD2, align="right")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "The adjustment is immaterial, and that is the finding rather than a disappointment. At "
        "roughly $0.7M of commission earned a year against $33M of ARR, commission "
        "capitalisation is a real mechanic with a negligible P&L effect. This tab does not "
        "replace the official forecast.")
    row += 2
    row = st.source_note(ws, row, 2, "fct_accounting_enhanced_pnl (Phase 8).")

    # No chart on this tab, deliberately. The rollforward table already shows the build
    # quarter by quarter, and a bar chart of the same ten numbers would make an accounting
    # schedule compete visually with the FP&A tabs it is meant to support.

    st.print_area(ws, "B1:I{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 10. Assumptions
# ---------------------------------------------------------------------------
def build_assumptions(ctx: Context) -> None:
    ws = ctx.sheet("Assumptions")
    st.presentation_sheet(ws, content=(46, 18, 17, 82, 21), gutter=False)

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- assumptions and methodology", REPORTING_LINE, ctx.stamp
    )
    row = st.section(
        ws, row, 2, "Decision-driving assumptions", width=5,
        note="The assumptions a reviewer would actually challenge, not a dump of every "
             "configuration key. Values are read from config/assumptions.yml, "
             "config/commentary_rules.yml and the approved marts on every build, so they cannot "
             "drift out of step with the numbers.")
    for offset, text in enumerate(["Assumption", "Value", "Unit", "Source / rationale", "Type"]):
        st.cell(
            ws, row, 2 + offset, text, style="header_left", fill=P.navy, wrap=True, valign="bottom")
    row += 1
    frame = ed.assumptions_table(ctx.marts)
    type_colour = {
        ed.TYPE_FROZEN: P.navy,
        ed.TYPE_HISTORICAL: P.blue,
        ed.TYPE_MANAGEMENT: P.input_ink,
        ed.TYPE_ACCOUNTING: P.favourable,
    }
    for _, record in frame.iterrows():
        assumption_type = str(record["type"])
        is_input = assumption_type in (ed.TYPE_MANAGEMENT, ed.TYPE_ACCOUNTING)
        st.cell(ws, row, 2, str(record["assumption"]), style="label")
        value = record["value"]
        fmt = None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            unit = str(record["unit"])
            fmt = {"USD": "usd", "USD / month": "usd", "months": "months",
                   "rate / year": "pct", "share of departures": "pct",
                   "share of earned": "pct", "share of Budget": "pct",
                   "count": "count"}.get(unit, "rate2")
        st.cell(
            ws, row, 3, value, fmt=fmt, style="value_bold",
            colour=P.input_ink if is_input else P.ink,
            fill=P.input_fill if is_input else None)
        st.cell(ws, row, 4, str(record["unit"]), style="label_muted")
        st.cell(ws, row, 5, str(record["source"]), style="label_muted")
        st.cell(
            ws, row, 6, assumption_type, bold=True,
            colour=type_colour.get(assumption_type, P.grey), align="left")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "Type legend -- Frozen Phase 1: fixed by the specification and never edited to make a "
        "build pass. Historical derivation: computed from actuals. Management assumption: a "
        "judgement, shown in amber. Accounting policy: the entity's own frozen policy.")
    row += 2
    st.note(
        ws, row, 2,
        "Materiality and priority thresholds are this project's own documented management "
        "reporting convention. PHASE1_SPEC stops at the Phase 6 reforecast and does not define "
        "bridge-commentary materiality, so these are stated plainly rather than implied to be "
        "Board-approved policy.")
    row += 2
    st.source_note(
        ws, row, 2, "config/assumptions.yml, config/commentary_rules.yml, int_forecast_drivers, "
                    "fct_cash_runway_policy.")
    st.print_area(ws, "B1:F{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# 11. Controls
# ---------------------------------------------------------------------------
def build_controls(ctx: Context, *, workbook_checks: list[tuple[str, Any, str]]) -> None:
    ws = ctx.sheet("Controls")
    st.presentation_sheet(ws, content=(34, 11, 74, 13, 13), gutter=False)

    row = st.title_block(
        ws, COMPANY, SUBTITLE + " -- controls and traceability", REPORTING_LINE, ctx.stamp
    )

    controls = ed.controls(ctx.marts)
    st.cell(ws, row, COL, "OVERALL STATUS", style="kpi_label")
    ws.row_dimensions[row].height = st.R.kpi_label
    row += 1
    status_row = row
    for col in range(COL, COL + 5):
        st.cell(ws, row, col)
    ws.row_dimensions[row].height = 30.0
    ws.row_dimensions[row + 1].height = st.R.note
    st.cell(
        ws, row, 2,
        '=IF(SUM(tbl_controls[Violations])=0,"READY / PASS","FAIL")',
        style="status")
    st.cell(
        ws, row + 1, 2,
        '=IF(SUM(tbl_controls[Violations])=0,'
        '"Every upstream control returned zero violation rows.",'
        '"One or more upstream controls returned violation rows -- see the table below.")',
        style="label_muted")
    row += 3

    row = st.section(
        ws, row, 2, "Upstream analytical controls", width=5,
        note="A control passes if and only if its query returns zero rows. Every one is "
             "re-run on every build of the analytical layer, and a violation fails the build "
             "before any mart is exported.")
    display = controls.rename(
        columns={
            "control": "Control", "phase": "Phase", "label": "What it enforces",
            "violation_rows": "Violations", "status": "Status",
        }
    )
    first, last = st.write_table(
        ws, display, name="tbl_controls", top=row, left=2,
        formats={"Violations": "count"})
    ctx.tables["tbl_controls"] = TableRef(
        "Controls", "tbl_controls", 2, first, last, list(display.columns)
    )
    st.column_widths(ws, {"B": 34, "C": 11, "D": 74, "E": 13, "F": 13})
    row = last + 2
    row = st.source_note(
        ws, row, 2,
        "ctl_control_results.csv, written by src/run_sql.py on every build of the analytical "
        "layer.")
    row += 1

    row = st.section(
        ws, row, 2, "Workbook-level checks", width=5,
        note="Checks on this artifact rather than on the analytical layer. Every one is also "
             "asserted independently by src/validate_excel_model.py and tests/test_excel_model.py.")
    for offset, text in enumerate(["Check", "Value", "", "Status"]):
        st.cell(
            ws, row, 2 + offset if offset < 2 else 2 + offset, text, style="header_left", fill=P.navy, wrap=True, valign="bottom")
    row += 1
    for label, value, status in workbook_checks:
        st.cell(ws, row, 2, label, style="label")
        st.cell(ws, row, 3, value, style="label_muted")
        st.cell(
            ws, row, 5, status, bold=True,
            style="label_bold",
            colour=P.favourable if status in ("PASS", "None") else P.navy)
        row += 1
    row += 1

    row = st.section(ws, row, 2, "Full deterministic management commentary", width=5)
    for offset, text in enumerate(["Priority", "Section", "Headline"]):
        st.cell(
            ws, row, 2 + offset, text, style="header_left", fill=P.navy, wrap=True, valign="bottom")
    row += 1
    for rank in range(1, len(ctx.frames["tbl_commentary"]) + 1):
        key = str(rank)
        st.cell(
            ws, row, 2, "=" + xlookup(key, "tbl_commentary", "exec_rank", "priority"),
            style="label_bold")
        st.cell(
            ws, row, 3, "=" + xlookup(key, "tbl_commentary", "exec_rank", "section"),
            style="label_muted")
        st.cell(
            ws, row, 4, "=" + xlookup(key, "tbl_commentary", "exec_rank", "headline"),
            style="label")
        row += 1
    row += 1
    st.note(
        ws, row, 2,
        "Supporting data sheets are hidden, not protected and never veryHidden: right-click any "
        "tab and choose Unhide to inspect every table this workbook reads. The workbook carries "
        "no password, no macros and no external links.")
    row += 2
    st.source_note(ws, row, 2, "fct_commentary_output (Phase 7).")

    # The headline is written green and turns red on its own formula result, so a workbook
    # rebuilt over a failing control cannot show a green PASS.
    st.status_conditional_format(ws, "B{r}:F{r}".format(r=status_row))
    st.print_area(ws, "B1:F{r}".format(r=row + 1))


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def _residual(frames: dict[str, pd.DataFrame], name: str) -> float:
    frame = frames[name]
    result = float(frame[frame["line_kind"] == "result"]["amount"].iloc[0])
    running = float(frame[frame["line_kind"] == "component"]["amount"].sum())
    anchor = float(frame[frame["line_kind"] == "anchor"]["amount"].iloc[0])
    return result - (anchor + running)


def build(
    *, marts_dir: Path = ed.MARTS_DIR, output: Path = OUTPUT_PATH, verbose: bool = True
) -> Path:
    """Read the marts, build every sheet, and save the workbook. Returns the output path."""
    marts = ed.load_marts(marts_dir)
    stamp_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    stamp = (
        "Workbook v" + WORKBOOK_VERSION + "  |  built " + stamp_time
        + "  |  source: data/marts (Phase 3-8 analytical layer, frozen)"
    )

    wb = Workbook()
    wb.remove(wb.active)
    for name in VISIBLE_SHEETS + DATA_SHEETS:
        wb.create_sheet(name)
    wb.calculation.fullCalcOnLoad = True

    ctx = Context(wb, marts, stamp)
    build_data_sheets(ctx)
    build_chart_data(ctx)
    build_executive_summary(ctx)
    build_arr_retention(ctx)
    build_gtm(ctx)
    build_forecast(ctx)
    build_pnl(ctx)
    build_budget_bridge(ctx)
    build_scenarios(ctx)
    build_runway_hiring(ctx)
    build_accounting(ctx)
    build_assumptions(ctx)

    residuals = {
        "Exit ARR": _residual(ctx.frames, "tbl_arr_bridge"),
        "Gross Profit": _residual(ctx.frames, "tbl_gp_bridge"),
        "Total OpEx": _residual(ctx.frames, "tbl_opex_bridge"),
        "Operating Income": _residual(ctx.frames, "tbl_oi_bridge"),
    }
    worst = max(abs(v) for v in residuals.values())
    deferred_residual = float(
        ctx.frames["tbl_deferred_revenue"]["rollforward_residual"].abs().max()
    )
    workbook_checks: list[tuple[str, Any, str]] = [
        ("Workbook version", WORKBOOK_VERSION, "PASS"),
        ("Build timestamp (UTC)", stamp_time, "PASS"),
        ("Source reporting date", "30 June 2026", "PASS"),
        ("Forecast cutover", "Jan-Jun 2026 actual, Jul-Dec 2026 Base reforecast", "PASS"),
        ("Source marts read", str(len(marts)) + " committed CSV extracts", "PASS"),
        ("External workbook links", "None -- every value is embedded", "PASS"),
        ("Macros / VBA", "None -- .xlsx, not .xlsm", "PASS"),
        ("Workbook protection", "None -- formulas and data are inspectable", "PASS"),
        (
            "Largest bridge residual (Exit ARR / GP / OpEx / OI)",
            "${:,.2f}".format(worst) + " against a $1.00 tolerance",
            "PASS" if worst < 1.0 else "FAIL"),
        (
            "Largest deferred-revenue rollforward residual",
            "${:,.2f}".format(deferred_residual),
            "PASS" if deferred_residual < 1.0 else "FAIL"),
        (
            "Formula recalculation",
            "openpyxl does not calculate; the workbook is saved with "
            "full-calculation-on-load and Excel recalculates on open",
            "PASS"),
    ]
    build_controls(ctx, workbook_checks=workbook_checks)

    for name in VISIBLE_SHEETS:
        wb[name].sheet_state = "visible"
        st.finalise_sheet(wb[name])
    wb.active = 0

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if verbose:
        size_kb = output.stat().st_size / 1024
        print("Workbook: " + str(output))
        print("  {:.0f} KB, {} visible sheets, {} supporting data sheets".format(
            size_kb, len(VISIBLE_SHEETS), len(DATA_SHEETS)
        ))
    return output


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="python -m src.build_excel_model",
        description="Build the Phase 9 recruiter-facing Excel FP&A operating model.")
    parser.add_argument(
        "--marts", type=Path, default=ed.MARTS_DIR, help="Directory holding the committed marts."
    )
    parser.add_argument("--out", type=Path, default=OUTPUT_PATH, help="Workbook output path.")
    parser.add_argument(
        "--skip-validation", action="store_true",
        help="Build without running the workbook validation suite afterwards.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        path = build(marts_dir=args.marts, output=args.out)
    except ed.MartError as error:
        print("EXCEL BUILD FAILED - " + str(error))
        return 1

    if args.skip_validation:
        return 0

    from .validate_excel_model import validate

    result = validate(path, marts_dir=args.marts)
    print()
    print(result.summary())
    return 0 if result.passed else 1


if __name__ == "__main__":
    sys.exit(main())
