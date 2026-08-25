"""Presentation vocabulary for the Phase 9 Excel operating model.

One font, one restrained finance palette, one set of number formats, and the small helpers
that write a cell, a section header, a KPI card, an Excel Table or a chart. Everything that
decides what the workbook *looks* like lives here; `src/build_excel_model.py` decides what it
*says*.

Nothing in this module reads a mart or computes a business figure.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass
from typing import Any, Sequence

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

FONT_NAME = "Calibri"


# ---------------------------------------------------------------------------
# Palette -- one navy, one blue, neutral greys, and red/green reserved
# exclusively for unfavourable / favourable variance (PHASE1_SPEC section 12).
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Palette:
    navy: str = "1F3864"          # titles, header bands
    blue: str = "2E5E8C"          # section headers, chart series 1
    blue_light: str = "7FA6C9"    # chart series 2
    blue_pale: str = "DCE6F1"     # header band fill
    ink: str = "262626"           # body text
    grey: str = "595959"          # secondary text, source notes
    grey_light: str = "BFBFBF"    # rules and borders
    band: str = "F2F5F9"          # zebra / subtotal band
    actual: str = "EDF2F8"        # ACTUAL period fill
    actual_ink: str = "1F3864"
    forecast: str = "FDF4E7"      # FORECAST period fill
    forecast_ink: str = "9C6516"
    favourable: str = "1E6B3A"
    unfavourable: str = "9C1F2E"
    input_fill: str = "FFF7E0"    # assumption / input cell
    input_ink: str = "7A5B00"
    white: str = "FFFFFF"


P = Palette()

# Consistent scenario ordering everywhere in the workbook.
SCENARIO_ORDER = ("Bear", "Base", "Bull")
SCENARIO_COLOURS = {"Bear": P.unfavourable, "Base": P.blue, "Bull": P.favourable}

SEGMENT_ORDER = ("SMB", "Mid-Market", "Enterprise")


# ---------------------------------------------------------------------------
# Number formats -- management conventions: parentheses for negatives, "-" for
# zero, no cents on management tables, units always visible.
# ---------------------------------------------------------------------------
NUMFMT: dict[str, str] = {
    "usd_m": '$#,##0.0,,"M";($#,##0.0,,"M");"-"',
    "usd_m_signed": '+$#,##0.0,,"M";-$#,##0.0,,"M";"-"',
    "usd_k": '$#,##0,"K";($#,##0,"K");"-"',
    "usd_k_signed": '+$#,##0,"K";-$#,##0,"K";"-"',
    "usd": '$#,##0;($#,##0);"-"',
    "usd_plain": '#,##0;(#,##0);"-"',
    "usd_signed": '+#,##0;-#,##0;"-"',
    "usd_cents": '#,##0.00;(#,##0.00);"-"',
    # Detailed schedules are stated in $000: the scaling comma divides by a thousand, and the
    # column header carries the unit so the figure is never ambiguous.
    "usd_000": '#,##0,;(#,##0,);"-"',
    "usd_000_signed": '+#,##0,;-#,##0,;"-"',
    "pct": '0.0%;(0.0%);"-"',
    "pct2": '0.00%;(0.00%);"-"',
    "pct_signed": '+0.0%;-0.0%;"-"',
    "bps": '+#,##0" bps";-#,##0" bps";"-"',
    "fte": '#,##0.0;(#,##0.0);"-"',
    "fte_signed": '+#,##0.0;-#,##0.0;"-"',
    "count": '#,##0;(#,##0);"-"',
    "months": '#,##0.0" mo";(#,##0.0)" mo";"-"',
    "months_signed": '+#,##0.0" mo";-#,##0.0" mo";"-"',
    "ratio": '0.00"x";(0.00"x");"-"',
    "rate2": "0.00",
    "rate3": "0.000",
    "days": '#,##0" days"',
    "date_month": "mmm-yy",
    "text": "@",
}



# ---------------------------------------------------------------------------
# TYPOGRAPHY -- the whole scale, defined once
# ---------------------------------------------------------------------------
# Every piece of text in the workbook is one of these tokens. `cell(..., style="body")` resolves
# a token to its size, weight, colour and default alignment, so no size or colour literal needs
# to appear in `build_excel_model.py`. Changing a size here changes it everywhere it is used.
TEXT_STYLES: dict[str, dict[str, Any]] = {
    "title":       {"size": 18,   "bold": True,  "colour": P.navy,  "align": "left"},
    "subtitle":    {"size": 11,   "bold": True,  "colour": P.blue,  "align": "left"},
    "meta":        {"size": 8,    "bold": False, "colour": P.grey,  "align": "left"},
    "section":     {"size": 10.5, "bold": True,  "colour": P.navy,  "align": "left"},
    "subsection":  {"size": 9,    "bold": True,  "colour": P.grey,  "align": "left"},
    "note":        {"size": 8.5,  "bold": False, "colour": P.grey,  "align": "left"},
    "source":      {"size": 8,    "bold": False, "colour": P.grey,  "align": "left"},
    "header":      {"size": 9,    "bold": True,  "colour": P.white, "align": "right"},
    "header_left": {"size": 9,    "bold": True,  "colour": P.white, "align": "left"},
    "label":       {"size": 9.5,  "bold": False, "colour": P.ink,   "align": "left"},
    "label_bold":  {"size": 9.5,  "bold": True,  "colour": P.navy,  "align": "left"},
    "label_muted": {"size": 9,    "bold": False, "colour": P.grey,  "align": "left"},
    "value":       {"size": 9.5,  "bold": False, "colour": P.ink,   "align": "right"},
    "value_bold":  {"size": 9.5,  "bold": True,  "colour": P.navy,  "align": "right"},
    "value_muted": {"size": 9,    "bold": False, "colour": P.grey,  "align": "right"},
    "kpi_label":   {"size": 7.5,  "bold": True,  "colour": P.grey,  "align": "left"},
    "kpi_value":   {"size": 16,   "bold": True,  "colour": P.navy,  "align": "left"},
    "kpi_note":    {"size": 7.5,  "bold": False, "colour": P.grey,  "align": "left"},
    "status":      {"size": 20,   "bold": True,  "colour": P.favourable, "align": "left"},
    "input":       {"size": 9.5,  "bold": True,  "colour": P.input_ink,  "align": "right"},
}


# ---------------------------------------------------------------------------
# SPACING -- row heights and the page grid, defined once
# ---------------------------------------------------------------------------
class Rows:
    """Row heights in points. Compact by default; nothing is left at Excel's 15pt guess."""

    title = 24.0
    subtitle = 15.0
    meta = 11.5
    rule = 5.0
    section = 17.0
    note = 12.5
    spacer = 6.5
    header = 25.0
    body = 14.0
    kpi_label = 12.0
    kpi_value = 21.0
    kpi_note = 11.5
    kpi_rule = 5.0
    panel = 25.0


R = Rows()

MARGIN_WIDTH = 1.8          # column A on every presentation sheet
GUTTER_WIDTH = 2.0          # the column between the content block and the chart band
CONTENT_COL = 2             # every presentation sheet starts writing at column B

# Three named chart sizes, so nothing is sized by hand. All are large enough that a 9.5pt
# axis label and a management-formatted data label stay readable at 100% zoom.
CHART_WIDE = (24.0, 11.0)       # cm -- bridges and long time series
CHART_STANDARD = (19.0, 10.0)   # cm -- the default
CHART_COMPACT = (19.0, 8.0)     # cm -- short category lists

CHART_WIDTH, CHART_HEIGHT = CHART_STANDARD   # backwards-compatible aliases
CHART_ROW_SPAN = 23             # body rows a chart occupies, for stacking without overlap
CHART_TITLE_PT = 1200           # hundredths of a point, the OOXML convention
CHART_AXIS_PT = 950
CHART_LEGEND_PT = 950
CHART_LABEL_PT = 900


# ---------------------------------------------------------------------------
# Modern-function namespacing -- the OOXML representation Excel actually stores
# ---------------------------------------------------------------------------
# Every worksheet function introduced after Excel 2007 is stored in the file format under the
# `_xlfn.` future-function namespace, and a handful of dynamic-array functions under
# `_xlfn._xlws.` as well. Excel writes the prefix itself and hides it in the formula bar, so a
# workbook saved by Excel shows `XLOOKUP(...)` on screen while the XML holds
# `_xlfn.XLOOKUP(...)`.
#
# openpyxl does not do this. It writes the formula string it is given, verbatim, straight into
# the `<f>` element. A bare `XLOOKUP(` therefore reaches Excel as an unrecognised defined name
# and every cell using it resolves to #NAME?. `qualify_formula` closes that gap, and
# `src/validate_excel_model.py` re-opens the saved package and asserts the prefixes are present
# in the worksheet XML rather than trusting what openpyxl reports back in Python.
MODERN_FUNCTIONS: dict[str, str] = {
    # Lookup and logic
    "XLOOKUP": "_xlfn.XLOOKUP",
    "XMATCH": "_xlfn.XMATCH",
    "LET": "_xlfn.LET",
    "LAMBDA": "_xlfn.LAMBDA",
    "IFS": "_xlfn.IFS",
    "SWITCH": "_xlfn.SWITCH",
    "ISOMITTED": "_xlfn.ISOMITTED",
    # Aggregation
    "MAXIFS": "_xlfn.MAXIFS",
    "MINIFS": "_xlfn.MINIFS",
    "PERCENTOF": "_xlfn.PERCENTOF",
    # Text
    "TEXTJOIN": "_xlfn.TEXTJOIN",
    "CONCAT": "_xlfn.CONCAT",
    "TEXTBEFORE": "_xlfn.TEXTBEFORE",
    "TEXTAFTER": "_xlfn.TEXTAFTER",
    "TEXTSPLIT": "_xlfn.TEXTSPLIT",
    "ARRAYTOTEXT": "_xlfn.ARRAYTOTEXT",
    "VALUETOTEXT": "_xlfn.VALUETOTEXT",
    # Dynamic arrays. FILTER and SORT additionally carry the worksheet-function namespace.
    "FILTER": "_xlfn._xlws.FILTER",
    "SORT": "_xlfn._xlws.SORT",
    "SORTBY": "_xlfn.SORTBY",
    "UNIQUE": "_xlfn.UNIQUE",
    "SEQUENCE": "_xlfn.SEQUENCE",
    "RANDARRAY": "_xlfn.RANDARRAY",
    "VSTACK": "_xlfn.VSTACK",
    "HSTACK": "_xlfn.HSTACK",
    "TOCOL": "_xlfn.TOCOL",
    "TOROW": "_xlfn.TOROW",
    "WRAPROWS": "_xlfn.WRAPROWS",
    "WRAPCOLS": "_xlfn.WRAPCOLS",
    "CHOOSECOLS": "_xlfn.CHOOSECOLS",
    "CHOOSEROWS": "_xlfn.CHOOSEROWS",
    "EXPAND": "_xlfn.EXPAND",
    "TAKE": "_xlfn.TAKE",
    "DROP": "_xlfn.DROP",
    "BYROW": "_xlfn.BYROW",
    "BYCOL": "_xlfn.BYCOL",
    "MAP": "_xlfn.MAP",
    "REDUCE": "_xlfn.REDUCE",
    "SCAN": "_xlfn.SCAN",
    "GROUPBY": "_xlfn.GROUPBY",
    "PIVOTBY": "_xlfn.PIVOTBY",
    "TRIMRANGE": "_xlfn.TRIMRANGE",
}

# A double-quoted Excel string literal; `""` is an escaped quote inside one.
_STRING_LITERAL = re.compile(r'"(?:[^"]|"")*"')

# A function call: a name immediately followed by an opening parenthesis. The lookbehind keeps
# the substitution off anything already namespaced (`_xlfn.XLOOKUP`), off a longer identifier
# that merely ends in a function name, and off a structured reference.
_FUNCTION_CALL = re.compile(r"(?<![A-Za-z0-9_.\]])([A-Za-z][A-Za-z0-9_.]*)(\s*\()")


def _qualify_code(fragment: str) -> str:
    def replace(match: "re.Match[str]") -> str:
        return MODERN_FUNCTIONS.get(match.group(1).upper(), match.group(1)) + match.group(2)

    return _FUNCTION_CALL.sub(replace, fragment)


# LET and LAMBDA declare *names*, and a declared name is stored under its own namespace,
# `_xlpm.` (Excel parameter), which is separate from the `_xlfn.` used for function names.
# `_xlfn.LET(v,H10,...)` is still broken: Excel cannot resolve `v`, rejects the formula, and
# drops the whole record -- which surfaces on open as
# "Removed Records: Formula from /xl/worksheets/sheetN.xml".
PARAMETER_NAMESPACE = "_xlpm."
PARAMETER_FUNCTIONS = {"LET", "LAMBDA"}

_NAME = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*$")


def _split_arguments(text: str) -> list[str]:
    """Split a function's argument list on top-level commas, respecting nesting and literals."""
    args: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    index = 0
    while index < len(text):
        char = text[index]
        if in_string:
            current.append(char)
            if char == '"':
                if index + 1 < len(text) and text[index + 1] == '"':
                    current.append('"')
                    index += 1
                else:
                    in_string = False
        elif char == '"':
            in_string = True
            current.append(char)
        elif char in "([{":
            depth += 1
            current.append(char)
        elif char in ")]}":
            depth -= 1
            current.append(char)
        elif char == "," and depth == 0:
            args.append("".join(current))
            current = []
        else:
            current.append(char)
        index += 1
    args.append("".join(current))
    return args


def _namespace_parameter(fragment: str, name: str) -> str:
    """Prefix every free occurrence of a declared name, outside literals and never as a call."""
    pattern = re.compile(
        r"(?<![A-Za-z0-9_.\]])" + re.escape(name) + r"(?![A-Za-z0-9_]*\s*\()(?![A-Za-z0-9_.])"
    )
    out: list[str] = []
    cursor = 0
    for literal in _STRING_LITERAL.finditer(fragment):
        out.append(pattern.sub(PARAMETER_NAMESPACE + name, fragment[cursor:literal.start()]))
        out.append(literal.group(0))
        cursor = literal.end()
    out.append(pattern.sub(PARAMETER_NAMESPACE + name, fragment[cursor:]))
    return "".join(out)


def _find_call(text: str, function: str) -> tuple[int, int, int] | None:
    """Locate `function(` and its matching close paren: (name_start, open_paren, close_paren)."""
    pattern = re.compile(
        r"(?<![A-Za-z0-9_.\]])(?:_xlfn\.)?" + function + r"\s*\(", re.IGNORECASE
    )
    for match in pattern.finditer(text):
        depth = 0
        in_string = False
        index = match.end() - 1
        while index < len(text):
            char = text[index]
            if in_string:
                if char == '"':
                    if index + 1 < len(text) and text[index + 1] == '"':
                        index += 1
                    else:
                        in_string = False
            elif char == '"':
                in_string = True
            elif char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0:
                    return match.start(), match.end() - 1, index
            index += 1
    return None


def _namespace_declared_names(formula: str) -> str:
    """Apply `_xlpm.` to every name declared by a LET or LAMBDA, innermost calls included."""
    for function in PARAMETER_FUNCTIONS:
        while True:
            found = _find_call(formula, function)
            if not found:
                break
            start, open_paren, close_paren = found
            head = formula[:start]
            call = formula[start:open_paren + 1]
            body = formula[open_paren + 1:close_paren]
            tail = formula[close_paren:]

            args = _split_arguments(body)
            if function == "LET":
                # name, value pairs, then a final calculation.
                declared = [i for i in range(0, len(args) - 1, 2)]
            else:
                # LAMBDA: every argument but the last is a parameter name.
                declared = list(range(len(args) - 1))

            names = [
                args[i].strip() for i in declared
                if _NAME.match(args[i].strip()) and not args[i].strip().startswith(
                    PARAMETER_NAMESPACE
                )
            ]
            if not names:
                # Already namespaced. Mark this call so the scan moves past it.
                marker = call.replace("(", "\x00", 1)
                formula = head + marker + body + tail
                continue
            for index in declared:
                args[index] = PARAMETER_NAMESPACE + args[index].strip()
            rebuilt = ",".join(args)
            for name in names:
                rebuilt = _namespace_parameter(rebuilt, name)
                # A declared name is scoped to this call only; re-prefixing an already
                # prefixed occurrence would produce `_xlpm._xlpm.name`.
                rebuilt = rebuilt.replace(
                    PARAMETER_NAMESPACE + PARAMETER_NAMESPACE + name,
                    PARAMETER_NAMESPACE + name,
                )
            formula = head + call.replace("(", "\x00", 1) + rebuilt + tail
    return formula.replace("\x00", "(")


def qualify_formula(value: Any) -> Any:
    """Rewrite a formula into the OOXML representation Excel expects.

    Two distinct namespaces are involved and both are required:

    * `_xlfn.` on the *name of a function* introduced after Excel 2007 (`_xlfn.XLOOKUP`), and
      `_xlfn._xlws.` on a few worksheet-scoped dynamic-array functions;
    * `_xlpm.` on a *name declared by* LET or LAMBDA (`_xlpm.v`), and on every reference to it
      inside that call.

    Getting the first right and the second wrong still produces a formula Excel refuses: it
    cannot resolve the declared name, so it drops the whole record and reports
    "Removed Records: Formula" on open.

    Substitution happens only outside string literals, so a lookup key or a commentary headline
    containing a function or parameter name is never touched. The transform is idempotent.
    """
    if not isinstance(value, str) or not value.startswith("="):
        return value
    out: list[str] = []
    cursor = 0
    for literal in _STRING_LITERAL.finditer(value):
        out.append(_qualify_code(value[cursor:literal.start()]))
        out.append(literal.group(0))
        cursor = literal.end()
    out.append(_qualify_code(value[cursor:]))
    return _namespace_declared_names("".join(out))


def display_formula(value: Any) -> Any:
    """Strip the OOXML namespaces, giving the formula as Excel shows it in the formula bar.

    Used by the validation layer so a structural check can be written against the readable
    form while the stored form keeps its prefixes. Both namespaces are removed: `_xlfn.` on
    function names and `_xlpm.` on names declared by LET or LAMBDA.
    """
    if not isinstance(value, str):
        return value
    return (
        value.replace("_xlfn._xlws.", "").replace("_xlfn.", "").replace(PARAMETER_NAMESPACE, "")
    )


def resolve_format(fmt: str) -> str:
    """Turn a format token into an Excel number-format pattern.

    A token that is not in `NUMFMT` and does not look like an Excel pattern is a typo, and
    raises here rather than being written to the cell verbatim. Silently writing the token
    would produce a format string Excel then has to interpret -- a token containing `d`, `m`,
    `y`, `h` or `s` is read as a *date* pattern, which turns a dollar figure into a datetime.
    """
    if fmt in NUMFMT:
        return NUMFMT[fmt]
    if any(ch in fmt for ch in "#0?@") or fmt == "General":
        return fmt
    raise KeyError(
        "Unknown number-format token " + repr(fmt) + ". Add it to excel_style.NUMFMT or pass a "
        "literal Excel number-format pattern."
    )


# ---------------------------------------------------------------------------
# Cell writing
# ---------------------------------------------------------------------------
_THIN = Side(style="thin", color=P.grey_light)
_MED = Side(style="medium", color=P.navy)


def _border(spec: str | None) -> Border | None:
    if not spec:
        return None
    kw: dict[str, Side] = {}
    if "t" in spec:
        kw["top"] = _THIN
    if "b" in spec:
        kw["bottom"] = _THIN
    if "l" in spec:
        kw["left"] = _THIN
    if "r" in spec:
        kw["right"] = _THIN
    if "T" in spec:
        kw["top"] = _MED
    if "B" in spec:
        kw["bottom"] = _MED
    return Border(**kw)


def cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any = None,
    *,
    style: str | None = None,
    fmt: str | None = None,
    bold: bool | None = None,
    size: float | None = None,
    colour: str | None = None,
    fill: str | None = None,
    align: str | None = None,
    valign: str = "center",
    wrap: bool = False,
    indent: int = 0,
    border: str | None = None,
    underline: str | None = None,
):
    """Write one styled cell and return it. The single entry point for every cell.

    `style` names a token from `TEXT_STYLES`; any explicit argument overrides that token, so a
    one-off emphasis stays possible without inventing a new size. Passing neither falls back to
    the body style rather than to an arbitrary default.
    """
    token = TEXT_STYLES.get(style or "label", TEXT_STYLES["label"])
    size = token["size"] if size is None else size
    bold = token["bold"] if bold is None else bold
    colour = token["colour"] if colour is None else colour
    align = token.get("align") if align is None else align
    target = ws.cell(row=row, column=col)
    if value is not None:
        target.value = qualify_formula(value)
    target.font = Font(name=FONT_NAME, size=size, bold=bold, color=colour, underline=underline)
    if fmt:
        target.number_format = resolve_format(fmt)
    if fill:
        target.fill = PatternFill("solid", fgColor=fill)
    target.alignment = Alignment(
        horizontal=align, vertical=valign, wrap_text=wrap, indent=indent
    )
    bd = _border(border)
    if bd:
        target.border = bd
    return target


def row_values(
    ws: Worksheet, row: int, col: int, values: Sequence[Any], *, fmt: str | None = None, **kw
) -> None:
    """Write a horizontal run of cells sharing one style."""
    for offset, value in enumerate(values):
        cell(ws, row, col + offset, value, fmt=fmt, **kw)


def heights(ws: Worksheet, *rows_and_heights) -> None:
    """Set row heights from (row, height) pairs. Nothing is left at Excel's 15pt guess."""
    for row, height in rows_and_heights:
        ws.row_dimensions[row].height = height


def body_heights(ws: Worksheet, first: int, last: int, height: float | None = None) -> None:
    for row in range(first, last + 1):
        ws.row_dimensions[row].height = R.body if height is None else height


def title_block(ws: Worksheet, company: str, subtitle: str, reporting: str, stamp: str) -> int:
    """The dated header every presentation tab opens with. Returns the next free row.

    Four levels of hierarchy in four rows -- company, what this is, as at when, how it was
    built -- closed by a thin navy rule. Identical on all eleven presentation tabs, so a reader
    moving between them lands on the same shape every time.
    """
    col = CONTENT_COL
    cell(ws, 1, col, company, style="title")
    cell(ws, 2, col, subtitle, style="subtitle")
    cell(ws, 3, col, reporting, style="meta")
    cell(ws, 4, col, stamp, style="meta")
    for c in range(col, col + 24):
        cell(ws, 5, c, border="b")
    heights(ws, (1, R.title), (2, R.subtitle), (3, R.meta), (4, R.meta), (5, R.rule),
            (6, R.spacer))
    return 7


def section(
    ws: Worksheet, row: int, col: int, text: str, width: int = 8, note: str = "",
    accent: str | None = None,
) -> int:
    """A section heading. Returns the next free row.

    A rule under a navy heading, rather than a filled band -- restrained enough to repeat six
    times down a sheet without the page turning into stripes. `accent` recolours the heading
    where two sibling sections need to read as deliberately different things.
    """
    for offset in range(width):
        cell(ws, row, col + offset, border="B")
    cell(ws, row, col, text, style="section", colour=accent, border="B")
    ws.row_dimensions[row].height = R.section
    if note:
        cell(ws, row + 1, col, note, style="note")
        ws.row_dimensions[row + 1].height = R.note
        ws.row_dimensions[row + 2].height = R.spacer
        return row + 3
    ws.row_dimensions[row + 1].height = R.spacer
    return row + 2


def source_note(ws: Worksheet, row: int, col: int, text: str) -> int:
    cell(ws, row, col, "Source: " + text, style="source")
    ws.row_dimensions[row].height = R.note
    ws.row_dimensions[row + 1].height = R.spacer
    return row + 2


def note(ws: Worksheet, row: int, col: int, text: str) -> int:
    cell(ws, row, col, text, style="note")
    ws.row_dimensions[row].height = R.note
    return row + 1


def table_header(
    ws: Worksheet, row: int, col: int, headers: Sequence[str], *, first_left: bool = True
) -> int:
    """A navy header row: first column left-aligned, the rest right-aligned over their numbers."""
    for offset, text in enumerate(headers):
        left = first_left and offset == 0
        cell(
            ws, row, col + offset, text, style="header_left" if left else "header",
            fill=P.navy, wrap=True, valign="bottom",
        )
    ws.row_dimensions[row].height = R.header
    return row + 1


def rule_row(ws: Worksheet, row: int, col: int, width: int, *, weight: str = "t") -> None:
    """A thin separator across a table, instead of boxing every cell."""
    for offset in range(width):
        cell(ws, row, col + offset, border=weight)


def kpi_card(
    ws: Worksheet,
    row: int,
    col: int,
    label: str,
    value: Any,
    fmt: str,
    context: str = "",
    *,
    width: int = 2,
    value_colour: str | None = None,
) -> None:
    """A three-line KPI tile: label, value, context. Every card is the same size.

    Cards sit on a light card fill closed by a thin rule, with a gutter column between them, so
    the strip reads as a row of tiles rather than a block of colour.
    """
    for offset in range(3):
        for c in range(col, col + width):
            cell(ws, row + offset, c, fill=P.band)
    cell(ws, row, col, label.upper(), style="kpi_label", fill=P.band)
    cell(
        ws, row + 1, col, value, fmt=fmt, style="kpi_value", fill=P.band,
        colour=value_colour,
    )
    cell(ws, row + 2, col, context, style="kpi_note", fill=P.band)
    for c in range(col, col + width):
        cell(ws, row + 3, c, border="b")
    heights(ws, (row, R.kpi_label), (row + 1, R.kpi_value), (row + 2, R.kpi_note),
            (row + 3, R.kpi_rule))


def kpi_strip(
    ws: Worksheet, row: int, first_col: int, cards: Sequence[tuple], *, span: int = 2,
    gutter: int = 1,
) -> int:
    """Lay a row of identically sized KPI cards on the sheet grid. Returns the next free row."""
    for index, card in enumerate(cards):
        label, value, fmt, context, colour = card
        kpi_card(
            ws, row, first_col + index * (span + gutter), label, value, fmt, context,
            width=span, value_colour=colour,
        )
    return row + 4


def column_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def uniform_widths(ws: Worksheet, first: int, last: int, width: float) -> None:
    for idx in range(first, last + 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def presentation_sheet(
    ws: Worksheet,
    *,
    freeze_at: str = "A7",
    content: Sequence[float] = (),
    gutter: bool = True,
    chart_cols: int = 0,
    chart_width: float = 10.5,
) -> None:
    """Set up a presentation tab on the shared page grid.

    Every presentation sheet gets the same skeleton: a narrow margin in column A, a content
    block starting at column B whose widths are passed in, an optional gutter, and an optional
    band of uniform columns for charts. Titles, tables and charts therefore start at the same
    left margin on every tab, and charts land on a predictable column.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_at
    ws.sheet_view.zoomScale = 100
    ws.column_dimensions["A"].width = MARGIN_WIDTH
    for offset, width in enumerate(content):
        ws.column_dimensions[get_column_letter(CONTENT_COL + offset)].width = width
    next_col = CONTENT_COL + len(content)
    if gutter:
        ws.column_dimensions[get_column_letter(next_col)].width = GUTTER_WIDTH
        next_col += 1
    for offset in range(chart_cols):
        ws.column_dimensions[get_column_letter(next_col + offset)].width = chart_width
    ws.sheet_properties.tabColor = P.navy


def finalise_sheet(ws: Worksheet) -> None:
    """Give every used row an explicit compact height and set the sheet default.

    Called once per presentation tab after it is written, so a row nobody thought about does
    not silently fall back to Excel's 15pt guess and break the vertical rhythm.
    """
    ws.sheet_format.defaultRowHeight = R.body
    ws.sheet_format.customHeight = True
    for row in range(1, ws.max_row + 1):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = R.body


def chart_anchor(first_col: int, row: int) -> str:
    """The anchor for a chart in the chart band, on the shared vertical rhythm."""
    return get_column_letter(first_col) + str(row)


def print_area(ws: Worksheet, area: str, *, landscape: bool = True) -> None:
    ws.print_area = area
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def variance_conditional_format(ws: Worksheet, cell_range: str) -> None:
    """Red for negative, green for positive -- used only on true variance columns."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan", formula=["0"],
            font=Font(name=FONT_NAME, size=10, color=P.unfavourable),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan", formula=["0"],
            font=Font(name=FONT_NAME, size=10, color=P.favourable),
        ),
    )


STATUS_PASS_FILL = "E3F0E7"
STATUS_FAIL_FILL = "F7E4E6"


def status_conditional_format(ws: Worksheet, cell_range: str) -> None:
    """Colour an overall-status band from its own formula result -- green on PASS, red on FAIL.

    Both the fill and the type are driven by the formula, so the band cannot show a green
    READY / PASS over a failing control roster: the same cell decides the word and the colour.
    """
    size = TEXT_STYLES["status"]["size"]
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"FAIL"'],
            font=Font(name=FONT_NAME, size=size, bold=True, color=P.unfavourable),
            fill=PatternFill("solid", bgColor=STATUS_FAIL_FILL),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"READY / PASS"'],
            font=Font(name=FONT_NAME, size=size, bold=True, color=P.favourable),
            fill=PatternFill("solid", bgColor=STATUS_PASS_FILL),
        ),
    )


def fav_unfav_conditional_format(ws: Worksheet, cell_range: str) -> None:
    """Colour a Fav / Unfav text column by its own word, not by the sign of a number."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Favorable"'],
            font=Font(name=FONT_NAME, size=10, bold=True, color=P.favourable),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Unfavorable"'],
            font=Font(name=FONT_NAME, size=10, bold=True, color=P.unfavourable),
        ),
    )


# ---------------------------------------------------------------------------
# Excel Tables on the supporting data sheets
# ---------------------------------------------------------------------------
def _clean(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "item"):
        try:
            value = value.item()
        except (ValueError, AttributeError):
            value = str(value)
    if isinstance(value, float) and math.isnan(value):
        return None
    if value != value:  # pandas NA
        return None
    return value


def write_table(
    ws: Worksheet,
    frame,
    *,
    name: str,
    top: int = 1,
    left: int = 1,
    formats: dict[str, str] | None = None,
    widths: dict[str, float] | None = None,
) -> tuple[int, int]:
    """Write a DataFrame as a real Excel Table. Returns (first_data_row, last_row).

    Structured references into these tables are how every presentation tab reads its numbers,
    so the table name is the contract: it is asserted by `src/validate_excel_model.py`.
    """
    formats = formats or {}
    columns = [str(c) for c in frame.columns]
    for offset, column in enumerate(columns):
        cell(
            ws, top, left + offset, column, style="header", fill=P.navy, align="center",
            wrap=True, valign="bottom",
        )
    for r, (_, record) in enumerate(frame.iterrows(), start=top + 1):
        for offset, column in enumerate(frame.columns):
            value = _clean(record[column])
            fmt = formats.get(str(column))
            cell(
                ws, r, left + offset, value, fmt=fmt,
                style="label" if fmt in (None, "text") else "value",
            )

    last_row = top + len(frame)
    ref = (
        get_column_letter(left) + str(top) + ":"
        + get_column_letter(left + len(columns) - 1) + str(last_row)
    )
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)

    if widths:
        column_widths(ws, widths)
    else:
        for offset, column in enumerate(columns):
            letter = get_column_letter(left + offset)
            ws.column_dimensions[letter].width = max(12, min(34, len(column) + 4))
    return top + 1, last_row


def data_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Chart data blocks -- a purpose-built presentation layer, not arbitrary cells
# ---------------------------------------------------------------------------
# Every chart reads a block written by `write_chart_block`: one text category column and one or
# more contiguous numeric series columns, stored values only, no blank rows in the middle and no
# formula that Excel has to calculate before the chart can render.
#
# The block records the sheet it was written to and the exact columns it occupies, and the chart
# helpers take the block rather than loose coordinates. A chart can therefore no longer point at
# the right rows on the wrong sheet -- which is precisely the defect that left every waterfall
# in this workbook empty when Excel opened it.


@dataclass
class ChartBlock:
    """Where a chart's data physically lives."""

    sheet: str
    title: str
    first_row: int          # first data row
    last_row: int           # last data row
    label_col: int          # the text category column
    series: dict[str, int]  # series name -> column index

    @property
    def rows(self) -> int:
        return self.last_row - self.first_row + 1

    def category_ref(self) -> str:
        letter = get_column_letter(self.label_col)
        return "'{s}'!${c}${a}:${c}${b}".format(
            s=self.sheet, c=letter, a=self.first_row, b=self.last_row
        )

    def series_ref(self, name: str) -> str:
        letter = get_column_letter(self.series[name])
        return "'{s}'!${c}${a}:${c}${b}".format(
            s=self.sheet, c=letter, a=self.first_row, b=self.last_row
        )

    def series_title_ref(self, name: str) -> str:
        letter = get_column_letter(self.series[name])
        return "'{s}'!${c}${r}".format(s=self.sheet, c=letter, r=self.first_row - 1)


def write_chart_block(
    ws: Worksheet,
    top: int,
    left: int,
    title: str,
    categories: Sequence[Any],
    series: "dict[str, Sequence[Any]]",
    *,
    number_format: str = "usd_plain",
) -> ChartBlock:
    """Write one chart's data as a clean contiguous block and return where it landed.

    A `None` in a series is written as a genuinely empty cell, which Excel plots as a gap: that
    is how a waterfall suppresses the bar and the data label on a movement of zero, without
    dropping the category.
    """
    cell(ws, top, left, title, style="header_left", fill=P.navy)
    for offset, name in enumerate(series, start=1):
        cell(ws, top, left + offset, name, style="header", fill=P.navy)
    for index, label in enumerate(categories):
        row = top + 1 + index
        cell(ws, row, left, label, style="label")
        for offset, values in enumerate(series.values(), start=1):
            value = values[index]
            cell(ws, row, left + offset, value, fmt=number_format, style="value")
    return ChartBlock(
        sheet=ws.title,
        title=title,
        first_row=top + 1,
        last_row=top + len(categories),
        label_col=left,
        series={name: left + 1 + i for i, name in enumerate(series)},
    )


# ---------------------------------------------------------------------------
# Charts -- no 3D, no pies, no Excel default styling, readable at 100% zoom
# ---------------------------------------------------------------------------
def _character(size: int, colour: str, *, bold: bool = False) -> CharacterProperties:
    return CharacterProperties(
        sz=size, b=bold, solidFill=colour, latin=DrawingFont(typeface=FONT_NAME)
    )


def _rich(size: int, colour: str, *, bold: bool = False, rotation: int | None = None) -> RichText:
    """A run of chart text: axis labels, legend entries, data labels."""
    props = _character(size, colour, bold=bold)
    body = RichTextProperties(rot=rotation, vert="horz") if rotation else RichTextProperties()
    return RichText(
        bodyPr=body,
        p=[Paragraph(pPr=ParagraphProperties(defRPr=props), endParaRPr=props)],
    )


def chart_title(text: str) -> Title:
    """A management chart title: one size, one weight, navy. Never Excel's default."""
    props = _character(CHART_TITLE_PT, P.navy, bold=True)
    return Title(
        tx=Text(
            rich=RichText(
                bodyPr=RichTextProperties(),
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=props),
                    r=[RegularTextRun(t=text, rPr=props)],
                )],
            )
        ),
        overlay=False,
    )


def _apply_categories(chart, block: ChartBlock) -> None:
    """Point every series at the block's category column as TEXT.

    openpyxl's `set_categories` always writes a `numRef`, so a text category comes back to
    Excel as a number reference and the axis renders 1, 2, 3 instead of the labels. Categories
    are rewritten here as a `strRef`, which is what a text axis actually needs.
    """
    reference = block.category_ref()
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=reference))


def _standard(
    chart,
    title: str,
    *,
    legend: bool,
    size: tuple[float, float],
    number_format: str | None,
    tick_skip: int | None = None,
) -> None:
    """The house chart standard, applied explicitly so no Excel default leaks through."""
    chart.title = chart_title(title)
    chart.width, chart.height = size
    chart.roundedCorners = False

    # Charts read hidden supporting sheets. Excel's default is to plot visible cells only,
    # which silently renders every one of them empty; this is the setting that allows them.
    chart.visible_cells_only = False
    chart.display_blanks = "gap"

    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill="E8E8E8", w=6350)
    )
    chart.x_axis.majorGridlines = None
    chart.graphical_properties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    for axis in (chart.x_axis, chart.y_axis):
        axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=P.grey_light, w=6350))
        axis.txPr = _rich(CHART_AXIS_PT, P.grey)
        axis.delete = False
        axis.majorTickMark = "out"
        axis.minorTickMark = "none"
    if number_format:
        chart.y_axis.numFmt = resolve_format(number_format)
    if tick_skip and tick_skip > 1:
        # Keep every data point; thin only the tick LABELS so the axis stays readable.
        chart.x_axis.tickLblSkip = tick_skip
        chart.x_axis.tickMarkSkip = tick_skip

    if legend and chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False
        chart.legend.txPr = _rich(CHART_LEGEND_PT, P.grey)
    else:
        chart.legend = None


def _fill(series, colour: str | None) -> None:
    if colour is None:
        series.graphicalProperties = GraphicalProperties(noFill=True)
        series.graphicalProperties.line = LineProperties(noFill=True)
    else:
        series.graphicalProperties = GraphicalProperties(solidFill=colour)
        series.graphicalProperties.line = LineProperties(noFill=True)


def _labels(chart, number_format: str, *, colour: str = P.ink, position: str | None = None):
    labels = DataLabelList()
    labels.showVal = True
    labels.showSerName = False
    labels.showCatName = False
    labels.showLegendKey = False
    labels.numFmt = resolve_format(number_format)
    labels.txPr = _rich(CHART_LABEL_PT, colour, bold=True)
    if position:
        labels.dLblPos = position
    return labels


def column_chart(
    ws: Worksheet,
    data_ws: Worksheet,
    block: ChartBlock,
    *,
    anchor: str,
    title: str,
    series: Sequence[str] | None = None,
    colours: Sequence[str | None] = (),
    number_format: str = "usd_m",
    size: tuple[float, float] = None,
    legend: bool = True,
    data_labels: bool = False,
    horizontal: bool = False,
    grouping: str = "clustered",
    gap_width: int = 60,
    tick_skip: int | None = None,
):
    """A clustered or stacked bar chart over a chart-data block."""
    names = list(series) if series else list(block.series)
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.grouping = grouping
    chart.overlap = 100 if grouping == "stacked" else -10
    chart.gapWidth = gap_width
    for name in names:
        reference = Reference(
            data_ws, min_col=block.series[name],
            min_row=block.first_row - 1, max_row=block.last_row,
        )
        chart.add_data(reference, titles_from_data=True)
    _apply_categories(chart, block)
    _standard(
        chart, title, legend=legend, size=size or CHART_STANDARD,
        number_format=number_format, tick_skip=tick_skip,
    )
    for index, chart_series in enumerate(chart.series):
        _fill(chart_series, colours[index] if index < len(colours) else P.blue)
    if data_labels:
        chart.dataLabels = _labels(chart, number_format)
    ws.add_chart(chart, anchor)
    return chart


def line_chart(
    ws: Worksheet,
    data_ws: Worksheet,
    block: ChartBlock,
    *,
    anchor: str,
    title: str,
    series: Sequence[str] | None = None,
    colours: Sequence[str] = (),
    number_format: str = "usd_m",
    size: tuple[float, float] = None,
    legend: bool = True,
    tick_skip: int | None = None,
):
    names = list(series) if series else list(block.series)
    chart = LineChart()
    for name in names:
        reference = Reference(
            data_ws, min_col=block.series[name],
            min_row=block.first_row - 1, max_row=block.last_row,
        )
        chart.add_data(reference, titles_from_data=True)
    _apply_categories(chart, block)
    _standard(
        chart, title, legend=legend, size=size or CHART_STANDARD,
        number_format=number_format, tick_skip=tick_skip,
    )
    chart.marker = None
    for index, chart_series in enumerate(chart.series):
        chart_series.smooth = False
        line = LineProperties(w=28000)
        if index < len(colours):
            line.solidFill = colours[index]
        chart_series.graphicalProperties = GraphicalProperties(ln=line)
    ws.add_chart(chart, anchor)
    return chart


# The four series a stacked-column waterfall needs, in stacking order.
WATERFALL_SERIES = ("Base", "Anchor", "Increase", "Decrease")


def waterfall_chart(
    ws: Worksheet,
    data_ws: Worksheet,
    block: ChartBlock,
    *,
    anchor: str,
    title: str,
    number_format: str = "usd_m",
    size: tuple[float, float] = None,
):
    """A bridge, built as a stacked column with an invisible floor.

    openpyxl cannot emit Excel's native waterfall chart type, and forcing one is not worth the
    interoperability risk. A stacked column renders identically and is understood by every
    version of Excel: a transparent floor series lifts each movement to its running balance,
    anchors are drawn in navy, increases green and decreases red.

    Movements of zero are written as blanks in the block, so no bar and no data label is drawn
    for them.
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 35
    for name in WATERFALL_SERIES:
        reference = Reference(
            data_ws, min_col=block.series[name],
            min_row=block.first_row - 1, max_row=block.last_row,
        )
        chart.add_data(reference, titles_from_data=True)
    _apply_categories(chart, block)
    _standard(
        chart, title, legend=False, size=size or CHART_WIDE, number_format=number_format,
    )
    for chart_series, colour in zip(
        chart.series, (None, P.navy, P.favourable, P.unfavourable)
    ):
        _fill(chart_series, colour)
    # Label the anchors and the movements, never the invisible floor.
    for chart_series in chart.series[1:]:
        chart_series.dLbls = _labels(chart, number_format, position="ctr")
    chart.x_axis.txPr = _rich(CHART_AXIS_PT, P.grey, rotation=-2700000)
    ws.add_chart(chart, anchor)
    return chart


__all__ = [
    "FONT_NAME", "MODERN_FUNCTIONS", "NUMFMT", "P", "Palette", "Reference",
    "SCENARIO_COLOURS", "SCENARIO_ORDER",
    "SEGMENT_ORDER", "cell", "column_widths", "data_sheet",
    "CHART_COMPACT", "CHART_STANDARD", "CHART_WIDE", "ChartBlock", "body_heights",
    "chart_anchor", "chart_title", "column_chart", "display_formula", "finalise_sheet",
    "write_chart_block",
    "fav_unfav_conditional_format", "heights", "kpi_card", "kpi_strip", "line_chart",
    "note", "rule_row", "table_header", "R", "Rows", "TEXT_STYLES", "CHART_ROW_SPAN",
    "presentation_sheet", "qualify_formula",
    "print_area", "resolve_format", "row_values", "section", "source_note",
    "title_block", "uniform_widths", "variance_conditional_format", "waterfall_chart",
    "write_table",
]
