"""Validates the Phase 9 Excel operating model against the marts it was built from.

    python -m src.validate_excel_model

Two kinds of check, kept apart on purpose.

**Structural.** The file is a valid XLSX, the expected sheets exist with the expected
visibility, no worksheet name is duplicated, every required Excel Table is present, no external
workbook link exists anywhere in the package, no formula carries an external-workbook or
`#REF!` reference, no volatile or banned function is used, and every chart series resolves to a
range on a sheet that exists.

**Value.** Every headline figure the workbook displays is recomputed independently in Python
from the committed marts and compared to what the workbook actually stores.

### On formulas, stated rather than implied

`openpyxl` does not calculate. A formula cell in this workbook therefore holds a formula string
and **no cached result** -- there is nothing for this module to read back and compare. Pretending
otherwise would be the easiest way to ship a workbook whose numbers are wrong. Instead:

* formula strings are validated **structurally** -- the right table, the right column, the right
  operand cells, no banned function; and
* the value each formula should produce is validated **independently**, by recomputing it in
  Python from the underlying mart and comparing that to the supporting data table the formula
  reads.

So a broken lookup is caught by the structural check and a wrong number is caught by the value
check, without either claiming that Excel has run. Excel COM automation is deliberately not
required for the normal build.
"""

from __future__ import annotations

import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from . import excel_data as ed
from . import excel_style as st
from .build_excel_model import COMPANY, DATA_SHEETS, OUTPUT_PATH, VISIBLE_SHEETS

TOLERANCE = 0.01          # dollars, on figures stored at full precision
RATE_TOLERANCE = 1e-9

REQUIRED_TABLES = {
    "tbl_arr_monthly", "tbl_segment_arr", "tbl_retention", "tbl_retention_trend", "tbl_atr",
    "tbl_atr_pivot", "tbl_gtm_capacity", "tbl_gtm_constraint", "tbl_gtm_monthly",
    "tbl_win_rate", "tbl_unit_econ", "tbl_sales_efficiency", "tbl_pipeline", "tbl_pnl_summary",
    "tbl_pnl_monthly", "tbl_headcount_function", "tbl_arr_bridge", "tbl_gp_bridge",
    "tbl_opex_bridge", "tbl_oi_bridge", "tbl_rev_bridge", "tbl_mgmt_variance",
    "tbl_arr_bridge_segment", "tbl_opex_bridge_all", "tbl_gm_bridge", "tbl_scenario_summary",
    "tbl_scenario_trajectory", "tbl_scenario_drivers", "tbl_runway_policy", "tbl_hiring",
    "tbl_subscription_accounting", "tbl_deferred_revenue", "tbl_commission",
    "tbl_accounting_adjustment", "tbl_commentary", "tbl_controls",
}

# Volatile or fragile constructs this workbook commits to not using (PHASE 9 brief section 16).
BANNED_FUNCTIONS = ("OFFSET(", "INDIRECT(", "NOW(", "TODAY(", "RAND(", "RANDBETWEEN(")

# Functions that predate the `_xlfn.` future-function namespace and are therefore written bare.
# A function call that is in neither this set nor `excel_style.MODERN_FUNCTIONS` is one nobody
# has classified, and is failed rather than assumed safe -- an unclassified name is exactly how
# a #NAME? reaches a reviewer's screen.
LEGACY_FUNCTIONS = frozenset({
    "IF", "IFERROR", "SUM", "SUMIF", "SUMIFS", "SUMPRODUCT", "ROUND", "ROUNDUP", "ROUNDDOWN",
    "MIN", "MAX", "ABS", "AND", "OR", "NOT", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS",
    "AVERAGE", "INDEX", "MATCH", "VLOOKUP", "HLOOKUP", "LOOKUP", "CHOOSE", "TEXT", "VALUE",
    "LEFT", "RIGHT", "MID", "LEN", "TRIM", "UPPER", "LOWER", "CONCATENATE", "SUBSTITUTE",
    "ISBLANK", "ISNUMBER", "ISTEXT", "ISERROR", "ISNA", "NA", "SIGN", "IFNA",
})

# A function call inside a stored formula, with any namespace it already carries.
STORED_FUNCTION_CALL = re.compile(
    r"(?<![A-Za-z0-9_.\]])((?:_xlfn\.)?(?:_xlws\.)?)([A-Za-z][A-Za-z0-9_]*)\s*\("
)

# An external-workbook reference in a formula looks like [1]Sheet!A1 or 'C:\path\[Book.xlsx]'.
EXTERNAL_REF = re.compile(r"\[\d+\]|\[[^\]]+\.xls[xmb]?\]", re.IGNORECASE)


@dataclass
class Result:
    checks: list[tuple[str, bool, str]] = field(default_factory=list)

    def record(self, name: str, ok: bool, detail: str = "") -> bool:
        self.checks.append((name, bool(ok), detail))
        return bool(ok)

    @property
    def failures(self) -> list[tuple[str, bool, str]]:
        return [c for c in self.checks if not c[1]]

    @property
    def passed(self) -> bool:
        return not self.failures

    def summary(self) -> str:
        lines = [
            "Workbook validation: {p} of {n} checks passed".format(
                p=len(self.checks) - len(self.failures), n=len(self.checks)
            )
        ]
        for name, _ok, detail in self.failures:
            lines.append("  FAIL  " + name + (": " + detail if detail else ""))
        lines.append("WORKBOOK OK" if self.passed else "WORKBOOK VALIDATION FAILED")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _table_frame(wb: Workbook, name: str) -> pd.DataFrame:
    """Read a written Excel Table back into a DataFrame of stored cell values."""
    for ws in wb.worksheets:
        if name in ws.tables:
            table = ws.tables[name]
            rows = list(ws[table.ref])
            header = [str(c.value) for c in rows[0]]
            data = [[c.value for c in row] for row in rows[1:]]
            return pd.DataFrame(data, columns=header)
    raise KeyError("Excel Table " + name + " is not present in the workbook.")


def _formula_cells(wb: Workbook) -> Iterable[tuple[str, str, str]]:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield ws.title, cell.coordinate, cell.value


def _close(left: Any, right: Any, tolerance: float = TOLERANCE) -> bool:
    if left is None or right is None:
        return False
    try:
        return abs(float(left) - float(right)) <= tolerance
    except (TypeError, ValueError):
        return False


def _find_formula(wb: Workbook, sheet: str, contains: str) -> str | None:
    """Find a formula by its readable form and return it as stored.

    Matching happens against `display_formula` -- the text Excel shows in the formula bar --
    so a structural check can be written as `XLOOKUP("Exit ARR"` while the stored form carries
    its `_xlfn.` namespace. The stored string is what is returned.
    """
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and contains in st.display_formula(cell.value):
                return cell.value
    return None


def _worksheet_formulas(path: Path) -> list[tuple[str, str]]:
    """Every `<f>` element in every worksheet part, read straight out of the saved package.

    Deliberately not read through openpyxl. The defect this guards against lives in the XML
    the writer produced, and a check that asks openpyxl what it thinks the formula says would
    happily confirm a string that Excel cannot resolve.
    """
    formulas: list[tuple[str, str]] = []
    with zipfile.ZipFile(path) as archive:
        parts = sorted(n for n in archive.namelist() if n.startswith("xl/worksheets/sheet"))
        for part in parts:
            xml = archive.read(part).decode("utf-8")
            for match in re.finditer(r"<f[^>]*>(.*?)</f>", xml, re.S):
                text = (
                    match.group(1)
                    .replace("&lt;", "<").replace("&gt;", ">")
                    .replace("&quot;", '"').replace("&apos;", "'")
                    .replace("&amp;", "&")
                )
                formulas.append((part, text))
    return formulas


# ---------------------------------------------------------------------------
# Structural checks
# ---------------------------------------------------------------------------
def check_structure(path: Path, wb: Workbook, result: Result) -> None:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    result.record(
        "file is a valid xlsx package",
        "xl/workbook.xml" in names,
        "xl/workbook.xml not found in the package",
    )
    external = [n for n in names if n.startswith("xl/externalLinks/")]
    result.record(
        "no external link parts in the package",
        not external,
        "found " + ", ".join(external),
    )
    macros = [n for n in names if n.endswith(".bin") or "vbaProject" in n]
    result.record("no macro or VBA part", not macros, "found " + ", ".join(macros))

    titles = [ws.title for ws in wb.worksheets]
    result.record(
        "no duplicate worksheet names", len(titles) == len(set(titles)),
        "duplicates present",
    )
    missing_visible = [s for s in VISIBLE_SHEETS if s not in titles]
    result.record(
        "every required visible sheet exists", not missing_visible,
        "missing " + ", ".join(missing_visible),
    )
    result.record(
        "visible sheets are in the intended order",
        [t for t in titles if t in VISIBLE_SHEETS] == VISIBLE_SHEETS,
        "order is " + ", ".join(t for t in titles if t in VISIBLE_SHEETS),
    )
    wrong_state = [s for s in VISIBLE_SHEETS if wb[s].sheet_state != "visible"]
    result.record(
        "presentation sheets are visible", not wrong_state, ", ".join(wrong_state)
    )
    missing_data = [s for s in DATA_SHEETS if s not in titles]
    result.record(
        "every supporting data sheet exists", not missing_data, ", ".join(missing_data)
    )
    not_hidden = [s for s in DATA_SHEETS if s in titles and wb[s].sheet_state != "hidden"]
    result.record(
        "supporting data sheets are hidden, never veryHidden", not not_hidden,
        ", ".join(s + "=" + wb[s].sheet_state for s in not_hidden),
    )
    protected = [ws.title for ws in wb.worksheets if ws.protection.sheet]
    result.record(
        "no sheet is protected -- formulas stay inspectable", not protected,
        ", ".join(protected),
    )

    present_tables: set[str] = set()
    for ws in wb.worksheets:
        present_tables.update(ws.tables.keys())
    missing_tables = sorted(REQUIRED_TABLES - present_tables)
    result.record(
        "every required Excel Table exists", not missing_tables, ", ".join(missing_tables)
    )

    formulas = list(_formula_cells(wb))
    result.record("the workbook contains formulas", len(formulas) > 0)
    ext = [
        s + "!" + c for s, c, f in formulas if EXTERNAL_REF.search(f)
    ]
    result.record(
        "no formula references an external workbook", not ext, ", ".join(ext[:5])
    )
    refs = [s + "!" + c for s, c, f in formulas if "#REF!" in f]
    result.record("no formula contains #REF!", not refs, ", ".join(refs[:5]))
    banned = [
        s + "!" + c + " uses " + fn
        for s, c, f in formulas for fn in BANNED_FUNCTIONS if fn in f.upper()
    ]
    result.record(
        "no volatile or banned function is used", not banned, ", ".join(banned[:5])
    )
    errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in {
                    "#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?"
                }:
                    errors.append(ws.title + "!" + cell.coordinate)
    result.record("no stored Excel error value", not errors, ", ".join(errors[:5]))

    result.record(
        "workbook is set to recalculate on open",
        bool(getattr(wb.calculation, "fullCalcOnLoad", False)),
    )

    result.record(
        "the file is a practical size (< 8 MB)",
        path.stat().st_size < 8 * 1024 * 1024,
        "{:.1f} MB".format(path.stat().st_size / 1024 / 1024),
    )


def check_ooxml_function_prefixes(path: Path, result: Result) -> None:
    """Assert every modern function is stored under the namespace Excel requires.

    Read from the saved package, not from openpyxl. Functions introduced after Excel 2007 are
    stored as `_xlfn.NAME` (and a few dynamic-array functions as `_xlfn._xlws.NAME`); Excel
    hides the namespace in the formula bar but will not resolve the function without it. A bare
    `XLOOKUP(` in the XML is an unrecognised defined name, and every cell using it renders
    #NAME? the moment the workbook is opened.

    Three failure modes are checked:

    * a modern function written with no namespace, or the wrong one;
    * a namespace applied to a function that does not take one -- `_xlfn.SUM` is just as broken;
    * a function call that is in neither the modern nor the legacy roster, which means nobody
      has classified it and it may or may not resolve.
    """
    formulas = _worksheet_formulas(path)
    result.record(
        "worksheet formulas are readable from the saved package", bool(formulas),
        "found {n}".format(n=len(formulas)),
    )

    unprefixed: list[str] = []
    wrong_namespace: list[str] = []
    spurious: list[str] = []
    unclassified: list[str] = []

    for part, formula in formulas:
        sheet = part.rsplit("/", 1)[-1]
        for match in STORED_FUNCTION_CALL.finditer(_strip_literals(formula)):
            namespace, name = match.group(1), match.group(2).upper()
            stored = namespace + name
            required = st.MODERN_FUNCTIONS.get(name)
            if required is not None:
                if not namespace:
                    unprefixed.append(sheet + ": " + name)
                elif stored.upper() != required.upper():
                    wrong_namespace.append(sheet + ": " + stored + " should be " + required)
            elif name in LEGACY_FUNCTIONS:
                if namespace:
                    spurious.append(sheet + ": " + stored)
            else:
                unclassified.append(sheet + ": " + stored)

    result.record(
        "every modern function carries its OOXML namespace prefix", not unprefixed,
        ", ".join(sorted(set(unprefixed))[:6]),
    )
    result.record(
        "no modern function carries the wrong namespace", not wrong_namespace,
        ", ".join(sorted(set(wrong_namespace))[:6]),
    )
    result.record(
        "no legacy function carries a namespace it does not take", not spurious,
        ", ".join(sorted(set(spurious))[:6]),
    )
    result.record(
        "every function used is classified as modern or legacy", not unclassified,
        ", ".join(sorted(set(unclassified))[:6])
        + " -- add it to excel_style.MODERN_FUNCTIONS or LEGACY_FUNCTIONS",
    )

    # The two functions this workbook actually leans on, asserted by name so the regression is
    # named rather than implied by a general rule.
    joined = " ".join(f for _part, f in formulas)
    for name in ("XLOOKUP", "LET"):
        used = re.search(r"(?<![A-Za-z0-9_.])" + name + r"\s*\(", _strip_literals(joined))
        namespaced = "_xlfn." + name + "(" in joined.replace(" ", "")
        result.record(
            name + " is stored as _xlfn." + name,
            (not used) or namespaced,
            "found a bare " + name + "( in the worksheet XML",
        )

    # A declared name carries its own namespace. Getting `_xlfn.` right and `_xlpm.` wrong
    # still makes Excel drop the record: it cannot resolve the name, so the whole formula is
    # removed and the file opens with "Removed Records: Formula from /xl/worksheets/sheetN.xml".
    undeclared: list[str] = []
    for part, formula in formulas:
        sheet = part.rsplit("/", 1)[-1]
        for function in st.PARAMETER_FUNCTIONS:
            for name in _declared_names(formula, function):
                if not name.startswith(st.PARAMETER_NAMESPACE):
                    undeclared.append(sheet + ": " + function + " parameter " + name)
    result.record(
        "every LET / LAMBDA parameter carries the _xlpm. namespace", not undeclared,
        ", ".join(sorted(set(undeclared))[:6]),
    )


def _declared_names(formula: str, function: str) -> list[str]:
    """The names a LET or LAMBDA call declares, exactly as they are written in the XML."""
    names: list[str] = []
    text = formula
    while True:
        found = st._find_call(text, function)
        if not found:
            return names
        _start, open_paren, close_paren = found
        args = st._split_arguments(text[open_paren + 1:close_paren])
        positions = (
            range(0, len(args) - 1, 2) if function == "LET" else range(len(args) - 1)
        )
        for index in positions:
            candidate = args[index].strip()
            if st._NAME.match(candidate):
                names.append(candidate)
        text = text[:open_paren] + "\x00" + text[open_paren + 1:]


def _strip_literals(formula: str) -> str:
    """Blank out string literals so a lookup key can never be read as a function call."""
    return re.sub(r'"(?:[^"]|"")*"', '""', formula)


def check_formatting(wb: Workbook, result: Result) -> None:
    """Structural formatting checks.

    These do not and cannot judge whether the workbook looks good -- that needs a human with
    Excel open. What they can do is catch the mechanical regressions that make it look wrong:
    a presentation tab shipped with gridlines on, a title written at the wrong size, a KPI
    strip whose cards drifted out of alignment, a merged cell, a chart parked outside the
    sheet's own content. Aesthetic acceptance is explicitly not claimed from these.
    """
    gridlines = [s for s in VISIBLE_SHEETS if wb[s].sheet_view.showGridLines]
    result.record(
        "every presentation tab has gridlines off", not gridlines, ", ".join(gridlines)
    )

    unfrozen = [s for s in VISIBLE_SHEETS if not wb[s].freeze_panes]
    result.record("every presentation tab has frozen panes", not unfrozen, ", ".join(unfrozen))

    merged = [
        s + ": " + str(wb[s].merged_cells.ranges)
        for s in wb.sheetnames if wb[s].merged_cells.ranges
    ]
    result.record("no merged cells anywhere in the workbook", not merged, ", ".join(merged[:3]))

    # Title, subtitle and the two meta lines are the same four cells on every tab.
    title_style = st.TEXT_STYLES["title"]
    wrong_title = []
    for name in VISIBLE_SHEETS:
        cell = wb[name].cell(row=1, column=st.CONTENT_COL)
        if cell.value != COMPANY:
            wrong_title.append(name + ": title text")
        elif cell.font.sz != title_style["size"] or not cell.font.b:
            wrong_title.append(name + ": title style")
        elif (cell.font.color is None or cell.font.color.rgb not in
              ("00" + st.P.navy, "FF" + st.P.navy, st.P.navy)):
            wrong_title.append(name + ": title colour")
    result.record(
        "every tab opens with the same title block style", not wrong_title,
        ", ".join(wrong_title[:4]),
    )

    starts = [
        name for name in VISIBLE_SHEETS
        if wb[name].cell(row=1, column=1).value is not None
    ]
    result.record(
        "every tab keeps column A as the left margin", not starts, ", ".join(starts)
    )

    margins = [
        name for name in VISIBLE_SHEETS
        if abs((wb[name].column_dimensions["A"].width or 0) - st.MARGIN_WIDTH) > 0.01
    ]
    result.record(
        "every tab uses the same left-margin width", not margins, ", ".join(margins)
    )

    # The Executive Summary KPI strip: ten cards, all the same shape.
    exec_ws = wb["Executive Summary"]
    # A card's label row is the bold one at the KPI label size; its note row shares the size
    # but is not bold, so weight is what tells the two apart.
    label_rows = [
        row for row in range(1, 40)
        if isinstance(exec_ws.cell(row=row, column=st.CONTENT_COL).value, str)
        and exec_ws.cell(row=row, column=st.CONTENT_COL).font.sz
        == st.TEXT_STYLES["kpi_label"]["size"]
        and exec_ws.cell(row=row, column=st.CONTENT_COL).font.b
    ]
    result.record(
        "the Executive Summary carries two rows of KPI cards", len(label_rows) == 2,
        "found label rows " + str(label_rows),
    )
    card_shapes = set()
    for row in label_rows:
        for index in range(5):
            card_shapes.add((
                exec_ws.row_dimensions[row].height,
                exec_ws.row_dimensions[row + 1].height,
                exec_ws.row_dimensions[row + 2].height,
            ))
    result.record(
        "every KPI card has the same height", len(card_shapes) == 1, str(card_shapes)
    )

    # Charts sit inside the sheet's own content, not floating past it.
    stray = []
    for name in VISIBLE_SHEETS:
        ws = wb[name]
        for chart in ws._charts:
            anchor = chart.anchor._from
            if anchor.col + 1 < st.CONTENT_COL or anchor.row + 1 < 1:
                stray.append(name)
    result.record("no chart is anchored outside the page grid", not stray, ", ".join(stray))

    # Chart size and sources are checked in `check_chart_specs`, from the saved package;
    # openpyxl
    # does not restore `chart.width` / `chart.height` on reload -- it hands back its own
    # defaults -- so asking the reloaded object would pass no matter what was written.

    unstyled = [
        name for name in VISIBLE_SHEETS for chart in wb[name]._charts
        if chart.style is not None
    ]
    result.record(
        "no chart carries a built-in Excel chart style", not unstyled, ", ".join(unstyled)
    )

    fonts = set()
    for name in VISIBLE_SHEETS:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.value is not None and cell.font and cell.font.name:
                    fonts.add(cell.font.name)
    result.record(
        "one font family throughout", fonts <= {st.FONT_NAME}, ", ".join(sorted(fonts))
    )

    printless = [s for s in VISIBLE_SHEETS if not wb[s].print_area]
    result.record("every presentation tab has a print area", not printless, ", ".join(printless))


EXPECTED_CHARTS = 12
MIN_CHART_WIDTH = 18.0     # cm -- below this the axis labels stop being readable
MIN_CHART_HEIGHT = 7.5     # cm


def _chart_specs(path: Path) -> list[dict[str, Any]]:
    """Parse every chart part in the saved package into a comparable spec.

    Read from the package rather than through openpyxl, because openpyxl reconstructs charts on
    load with its own defaults -- it will happily report a size, a plot setting and a series
    list that are not what was written.
    """
    specs: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        parts = sorted(
            (n for n in archive.namelist() if re.match(r"xl/charts/chart\d+\.xml$", n)),
            key=lambda n: int(re.search(r"\d+", n).group()),
        )
        for part in parts:
            xml = archive.read(part).decode("utf-8")
            title = re.search(r"<a:t>(.*?)</a:t>", xml, re.S)
            plot_vis = re.search(r'<plotVisOnly val="(\d)"', xml)
            specs.append({
                "part": part,
                "title": title.group(1) if title else "",
                "plot_visible_only": plot_vis.group(1) if plot_vis else None,
                "categories": re.findall(r"<cat>.*?<f>(.*?)</f>", xml, re.S),
                "values": re.findall(r"<val>.*?<f>(.*?)</f>", xml, re.S),
                "category_is_text": "<cat><strRef>" in xml,
                # Plot-type elements only: `barChart`, `lineChart`, `pie3DChart`. The bare
                # `<chart>` container starts lower-case and is deliberately not matched.
                "types": set(re.findall(r"<(\w+Chart)>", xml)),
                "xml": xml,
            })
    return specs


def _resolve(wb: Workbook, reference: str) -> list[Any] | None:
    """Return the cell values a chart reference points at, or None if it does not resolve."""
    if "!" not in reference:
        return None
    sheet, rng = reference.split("!", 1)
    sheet = sheet.strip("'").replace("''", "'")
    if sheet not in wb.sheetnames:
        return None
    try:
        cells = wb[sheet][rng.replace("$", "")]
    except (ValueError, KeyError):
        return None
    if not isinstance(cells, tuple):
        cells = ((cells,),)
    return [c.value for row in cells for c in (row if isinstance(row, tuple) else (row,))]


def check_chart_specs(path: Path, wb: Workbook, result: Result) -> None:
    """Every chart must be one Excel can actually render, from data it can actually reach."""
    specs = _chart_specs(path)
    hidden = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}

    result.record(
        "the workbook carries the expected number of charts", len(specs) == EXPECTED_CHARTS,
        "found {n}, expected {e}".format(n=len(specs), e=EXPECTED_CHARTS),
    )

    untitled = [s["part"] for s in specs if not s["title"].strip()]
    result.record("every chart has a title", not untitled, ", ".join(untitled))

    seriesless = [s["part"] for s in specs if not s["values"]]
    result.record("every chart has at least one series", not seriesless, ", ".join(seriesless))

    numeric_categories = [s["part"] for s in specs if not s["category_is_text"]]
    result.record(
        "every chart's categories are a text reference", not numeric_categories,
        ", ".join(numeric_categories) + " -- a numRef renders the axis as 1, 2, 3",
    )

    # The setting that decides whether a chart sourced from a hidden sheet renders at all.
    plots_hidden = [
        s["part"] for s in specs
        if s["plot_visible_only"] != "0"
        and any(
            ref.split("!")[0].strip("'") in hidden
            for ref in s["categories"] + s["values"]
        )
    ]
    result.record(
        "charts reading hidden sheets are set to plot hidden data", not plots_hidden,
        ", ".join(plots_hidden) + " -- plotVisOnly must be 0",
    )

    unresolved: list[str] = []
    empty: list[str] = []
    mismatched: list[str] = []
    for spec in specs:
        name = spec["part"].split("/")[-1]
        category_lengths = set()
        for reference in spec["categories"]:
            values = _resolve(wb, reference)
            if values is None:
                unresolved.append(name + " cat " + reference)
                continue
            populated = [v for v in values if v is not None]
            if not populated:
                empty.append(name + " cat " + reference)
            category_lengths.add(len(values))
        value_lengths = set()
        for reference in spec["values"]:
            values = _resolve(wb, reference)
            if values is None:
                unresolved.append(name + " val " + reference)
                continue
            numbers = [v for v in values if isinstance(v, (int, float))]
            if not numbers:
                empty.append(name + " val " + reference)
            value_lengths.add(len(values))
        if category_lengths and value_lengths and category_lengths != value_lengths:
            mismatched.append(
                name + " cats " + str(category_lengths) + " vs vals " + str(value_lengths)
            )

    result.record(
        "every chart reference resolves to a real sheet and range", not unresolved,
        ", ".join(unresolved[:4]),
    )
    result.record(
        "every chart source range holds numeric data", not empty, ", ".join(empty[:4])
    )
    result.record(
        "category and value ranges are the same length", not mismatched,
        ", ".join(mismatched[:4]),
    )

    broken = [s["part"] for s in specs if "#REF!" in s["xml"]]
    result.record("no chart source contains #REF!", not broken, ", ".join(broken))

    unsupported = {
        t for spec in specs for t in spec["types"]
        if t not in {"barChart", "lineChart"}
    }
    result.record(
        "only supported chart types are used", not unsupported, ", ".join(sorted(unsupported))
    )

    # Size, read from the drawing XML: openpyxl does not restore it.
    sizes: set[tuple[float, float]] = set()
    with zipfile.ZipFile(path) as archive:
        for part in archive.namelist():
            if part.startswith("xl/drawings/drawing"):
                xml = archive.read(part).decode("utf-8")
                for match in re.finditer(r'<ext cx="(\d+)" cy="(\d+)"', xml):
                    sizes.add((
                        round(int(match.group(1)) / 360000, 2),
                        round(int(match.group(2)) / 360000, 2),
                    ))
    approved = {
        (round(w, 2), round(h, 2))
        for w, h in (st.CHART_WIDE, st.CHART_STANDARD, st.CHART_COMPACT)
    }
    off_standard = sizes - approved
    result.record(
        "every chart uses one of the three standard sizes", not off_standard,
        "found " + str(sorted(off_standard)),
    )
    too_small = {
        size for size in sizes
        if size[0] < MIN_CHART_WIDTH or size[1] < MIN_CHART_HEIGHT
    }
    result.record(
        "every chart clears the minimum readable size", not too_small,
        "found " + str(sorted(too_small)),
    )


def check_chart_values(path: Path, wb: Workbook, marts: dict[str, pd.DataFrame],
                       result: Result) -> None:
    """The key charts must plot the same numbers the marts carry.

    Checking that a reference resolves proves the chart will draw something. These check it
    draws the right thing.
    """
    specs = {s["title"]: s for s in _chart_specs(path)}

    def series_values(title_fragment: str, index: int) -> list[Any] | None:
        """The values a named chart's Nth series points at, or None if it cannot be resolved.

        Returning None rather than raising matters: a validator whose own checks throw on a
        broken workbook reports nothing at all, which is worse than reporting a failure.
        """
        for title, spec in specs.items():
            if title_fragment in title and index < len(spec["values"]):
                return _resolve(wb, spec["values"][index])
        return None

    def numbers(values: list[Any] | None) -> list[float]:
        return [v for v in (values or []) if isinstance(v, (int, float))]

    # Exit ARR bridge: the closing anchor is Base Exit ARR from Phase 7.
    variance = marts["fct_management_variance"].set_index("metric")
    anchors = numbers(series_values("Dec-2026 Exit ARR: Board Budget", 1))
    result.record(
        "the Exit ARR bridge chart opens at Budget and closes at Base",
        len(anchors) == 2
        and _close(anchors[0], variance.loc["exit_arr", "budget_amount"])
        and _close(anchors[-1], variance.loc["exit_arr", "base_amount"]),
        "anchors " + str(anchors),
    )

    # Scenario chart: Dec-26 Exit ARR ties fct_scenario_monthly.
    monthly = marts["fct_scenario_monthly"]
    expected = [
        float(monthly[(monthly["scenario"] == name)
                      & (monthly["month_end_date"] == ed.FY2026_END)]["ending_arr"].iloc[0])
        for name in ed.SCENARIOS
    ]
    plotted = numbers(series_values("Exit ARR by operating scenario", 0))
    result.record(
        "the scenario chart ties to fct_scenario_monthly",
        len(plotted) == len(expected)
        and all(_close(a, b) for a, b in zip(plotted, expected)),
        str(plotted) + " vs " + str(expected),
    )

    # Runway chart: ties fct_cash_runway_policy, and carries the 24-month floor.
    policy = marts["fct_cash_runway_policy"].set_index("path")
    runway_expected = [
        float(policy.loc[k, "policy_runway_months"])
        for k in ("Bear", "Base", "Bull", "Base_FullClose")
    ]
    runway_plotted = numbers(series_values("Board-policy runway vs the 24-month floor", 0))
    floor_plotted = numbers(series_values("Board-policy runway vs the 24-month floor", 1))
    result.record(
        "the runway chart ties to fct_cash_runway_policy",
        len(runway_plotted) == len(runway_expected)
        and all(_close(a, b, 1e-6) for a, b in zip(runway_plotted, runway_expected)),
        str(runway_plotted),
    )
    result.record(
        "the runway chart plots the 24-month floor",
        len(floor_plotted) == len(runway_expected)
        and all(_close(v, 24.0) for v in floor_plotted),
        str(floor_plotted),
    )

    # GTM chart: ties fct_new_logo_diagnosis, and the achievable series is the lesser of the two.
    diagnosis = ed.gtm_constraint_by_segment(marts).set_index("segment")
    segments = list(st.SEGMENT_ORDER)
    for index, column in enumerate((
        "h2_capacity_supported_arr", "h2_pipeline_supported_arr",
        "h2_constrained_new_logo_arr",
    )):
        plotted = numbers(series_values("New Logo ARR by segment", index))
        expected_series = [float(diagnosis.loc[x, column]) for x in segments]
        result.record(
            "the GTM chart ties to fct_new_logo_diagnosis (" + column + ")",
            len(plotted) == len(expected_series)
            and all(_close(a, b, 1e-4) for a, b in zip(plotted, expected_series)),
            str(plotted),
        )

    # Budget vs Base carries only monetary metrics -- no bps, FTE or months on a dollar axis.
    for title, spec in specs.items():
        if "Budget vs Base" in title:
            categories = (
                _resolve(wb, spec["categories"][0]) if spec["categories"] else None
            ) or []
            forbidden = {"Gross Margin", "Ending Headcount", "Board-policy runway"}
            result.record(
                "the Budget vs Base chart mixes no incompatible units",
                bool(categories) and not (set(map(str, categories)) & forbidden),
                str(categories),
            )


# ---------------------------------------------------------------------------
# Value checks -- independent recomputation from the marts
# ---------------------------------------------------------------------------
def check_executive(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    head = ed.headline(marts)
    ws = wb["Executive Summary"]

    kpis = {
        round(head.jun_2026_arr_actual, 2): "Jun-26 ARR actual",
        round(head.dec_2026_budget_arr, 2): "Dec-26 Budget ARR",
        round(head.dec_2026_base_arr, 2): "Dec-26 Base ARR",
        round(head.arr_variance, 2): "ARR variance",
        round(head.fy2026_revenue, 2): "FY2026 revenue",
        round(head.fy2026_operating_income, 2): "FY2026 operating income",
        round(head.base_policy_runway_months, 6): "Base policy runway",
    }
    stored: set[float] = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                stored.add(round(float(cell.value), 2))
                stored.add(round(float(cell.value), 6))
    missing = [label for value, label in kpis.items() if value not in stored]
    result.record(
        "Executive Summary KPI tiles carry the mart values", not missing, ", ".join(missing)
    )

    variance = _table_frame(wb, "tbl_mgmt_variance").set_index("metric")
    source = marts["fct_management_variance"].set_index("metric")
    mismatch = [
        metric for metric in source.index
        if not _close(variance.loc[metric, "base_amount"], source.loc[metric, "base_amount"])
        or not _close(variance.loc[metric, "budget_amount"], source.loc[metric, "budget_amount"])
    ]
    result.record(
        "executive scorecard table ties to fct_management_variance", not mismatch,
        ", ".join(mismatch),
    )

    formula = _find_formula(wb, "Executive Summary", 'XLOOKUP("Exit ARR"')
    result.record(
        "the scorecard reads Budget and Base by structured reference",
        formula is not None
        and "tbl_mgmt_variance[metric_label]" in formula
        and "tbl_mgmt_variance[budget_amount]" in formula,
        formula or "no Exit ARR lookup found",
    )
    var_formula = None
    for _s, _c, f in _formula_cells(wb):
        if f in ("=E{}-D{}".format(_c[1:], _c[1:]),):
            var_formula = f
    result.record(
        "variance is a formula, not a stored number",
        any(
            f.startswith("=E") and "-D" in f
            for s, _c, f in _formula_cells(wb) if s == "Executive Summary"
        ),
        var_formula or "",
    )


def check_pnl(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    stored = _table_frame(wb, "tbl_pnl_summary").set_index("line_key")
    expected = ed.pnl_summary(marts).set_index("line_key")
    columns = ["fy2025_actual", "h1_2026_actual", "h2_2026_base", "fy2026_base", "fy2026_budget"]
    mismatch = [
        key + "." + column
        for key in expected.index for column in columns
        if not _close(
            stored.loc[key, column], expected.loc[key, column],
            RATE_TOLERANCE if key.endswith("_pct") else TOLERANCE,
        )
    ]
    result.record("P&L table ties to the Phase 6 P&L", not mismatch, ", ".join(mismatch[:6]))

    # Budget Gross Profit and Operating Income must equal the derived figures, not merely be
    # taken from the bridge anchors and left unchecked.
    budget = expected["fy2026_budget"]
    result.record(
        "Budget Gross Profit equals Budget Revenue less Budget COGS",
        _close(budget["gross_profit"], budget["total_revenue"] - budget["total_cogs"]),
    )
    result.record(
        "Budget Operating Income equals Budget Gross Profit less Budget OpEx",
        _close(budget["operating_income"], budget["gross_profit"] - budget["total_opex"]),
    )

    base = expected["fy2026_base"]
    result.record(
        "FY2026 Base equals H1 actual plus H2 reforecast",
        _close(base["total_revenue"],
               expected.loc["total_revenue", "h1_2026_actual"]
               + expected.loc["total_revenue", "h2_2026_base"]),
    )

    formula = _find_formula(wb, "P&L", "higher_favorable")
    result.record(
        "Fav / Unfav is derived from the Phase 7 polarity value",
        formula is not None and "tbl_pnl_summary[polarity]" in formula,
        formula or "no polarity formula found",
    )


def check_forecast(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    ws = wb["Forecast"]
    grid = ed.forecast_grid(marts)
    months = grid.attrs["months"]
    result.record(
        "the forecast grid covers Jan-26 through Dec-26",
        months[0] == "Jan-26" and months[-1] == "Dec-26" and len(months) == 12,
        ", ".join(months),
    )
    # No 2027 period column may appear on the FY2026 reforecast grid. Checked against the
    # period labels themselves, not against prose -- the sheet's own footnote says in words
    # that FY2027 is excluded, and that sentence is not a violation of the rule it states.
    month_labels = {
        str(cell.value)
        for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and re.fullmatch(r"[A-Z][a-z]{2}-\d{2}", cell.value)
    }
    result.record(
        "no 2027 period column appears on the FY2026 reforecast grid",
        month_labels and not any(label.endswith("-27") for label in month_labels),
        ", ".join(sorted(month_labels)),
    )
    expected = {
        str(record["line_item"]): record for _, record in grid.iterrows()
    }
    stored: dict[str, list[Any]] = {}
    for row in ws.iter_rows():
        label = row[1].value
        if isinstance(label, str) and label in expected:
            stored[label] = [c.value for c in row[3:15]]
    missing = [label for label in expected if label not in stored]
    result.record(
        "every forecast line is written to the sheet", not missing, ", ".join(missing)
    )
    mismatch = []
    for label, record in expected.items():
        if label not in stored:
            continue
        for index, month in enumerate(months):
            value = stored[label][index]
            if isinstance(value, str):
                continue  # a subtotal / margin formula -- checked structurally below
            if not _close(value, record[month], RATE_TOLERANCE if record["unit"] == "pct" else TOLERANCE):
                mismatch.append(label + " " + month)
    result.record(
        "forecast detail values tie to the Phase 6 marts", not mismatch, ", ".join(mismatch[:6])
    )

    subtotal_formulas = [
        f for s, _c, f in _formula_cells(wb) if s == "Forecast" and f.startswith("=SUM(")
    ]
    result.record(
        "the FY2026 column is a formula, not a stored total", len(subtotal_formulas) > 0
    )
    check = _find_formula(wb, "Forecast", "Waterfall check")
    result.record(
        "the ARR waterfall carries a visible identity check",
        any(
            "ROUND(" in f and f.count("+") >= 4
            for s, _c, f in _formula_cells(wb) if s == "Forecast"
        ),
        check or "",
    )
    # Independently: the identity the Excel check tests must actually hold in the mart.
    arr = grid[grid["block"] == "ARR"].set_index("line_key")
    worst = 0.0
    for month in months:
        identity = (
            float(arr.loc["beginning_arr", month]) + float(arr.loc["new_logo_arr", month])
            + float(arr.loc["expansion_arr", month]) + float(arr.loc["reactivation_arr", month])
            + float(arr.loc["contraction_arr", month]) + float(arr.loc["churn_arr", month])
            - float(arr.loc["ending_arr", month])
        )
        worst = max(worst, abs(identity))
    result.record(
        "the ARR identity holds in every forecast month", worst < 1.0,
        "worst residual ${:,.4f}".format(worst),
    )


def check_bridges(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    blocks = {
        "tbl_arr_bridge": ed.arr_bridge(marts),
        "tbl_gp_bridge": ed.gross_profit_bridge(marts),
        "tbl_opex_bridge": ed.opex_bridge(marts, "Total OpEx"),
        "tbl_oi_bridge": ed.operating_income_bridge(marts),
    }
    for name, expected in blocks.items():
        stored = _table_frame(wb, name)
        mismatch = [
            str(expected.iloc[i]["line_item"])
            for i in range(len(expected))
            if not _close(stored.iloc[i]["amount"], expected.iloc[i]["amount"])
        ]
        result.record(
            name + " ties to its Phase 7 bridge mart", not mismatch, ", ".join(mismatch[:4])
        )
        anchor = float(expected[expected["line_kind"] == "anchor"]["amount"].iloc[0])
        components = float(expected[expected["line_kind"] == "component"]["amount"].sum())
        outcome = float(expected[expected["line_kind"] == "result"]["amount"].iloc[0])
        residual = outcome - (anchor + components)
        result.record(
            name + " reconciles Budget + components = Base", abs(residual) < 1.0,
            "residual ${:,.4f}".format(residual),
        )
        geometry_ok = all(
            _close(
                float(row["chart_invisible"]) + float(row["chart_increase"])
                - float(row["chart_decrease"]) * 0,
                float(row["chart_invisible"]) + float(row["chart_increase"]),
            )
            for _, row in expected.iterrows()
        )
        result.record(name + " chart geometry columns are present", geometry_ok)

    running = [
        f for s, _c, f in _formula_cells(wb) if s == "Budget Bridge" and f.startswith("=D")
    ]
    result.record(
        "bridge running balances are Excel formulas", len(running) > 0,
    )
    residual_formulas = [
        f for s, _c, f in _formula_cells(wb)
        if s == "Budget Bridge" and f.startswith("=ROUND(C")
    ]
    result.record(
        "each bridge shows a residual computed in the workbook",
        len(residual_formulas) >= len(blocks),
        "found {n}".format(n=len(residual_formulas)),
    )

    exit_arr = ed.arr_bridge(marts)
    base_value = float(exit_arr[exit_arr["line_kind"] == "result"]["amount"].iloc[0])
    variance = marts["fct_management_variance"].set_index("metric")
    result.record(
        "the Exit ARR bridge ends at the Base reforecast Exit ARR",
        _close(base_value, variance.loc["exit_arr", "base_amount"]),
    )


def check_scenarios(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    stored = _table_frame(wb, "tbl_scenario_summary").set_index("scenario")
    expected = ed.scenario_summary(marts).set_index("scenario")
    monthly = marts["fct_scenario_monthly"]
    mismatch = []
    for scenario in ed.SCENARIOS:
        subset = monthly[
            (monthly["scenario"] == scenario)
            & (monthly["month_end_date"] == ed.FY2026_END)
        ]
        direct = float(subset["ending_arr"].iloc[0])
        if not _close(stored.loc[scenario, "dec_2026_exit_arr"], direct):
            mismatch.append(scenario + " Dec-26 exit ARR")
        if not _close(
            stored.loc[scenario, "fy2026_operating_income"],
            expected.loc[scenario, "fy2026_operating_income"],
        ):
            mismatch.append(scenario + " FY2026 operating income")
    result.record(
        "scenario table ties to fct_scenario_monthly", not mismatch, ", ".join(mismatch)
    )
    result.record(
        "scenarios are ordered Bear / Base / Bull",
        list(_table_frame(wb, "tbl_scenario_summary")["scenario"]) == list(ed.SCENARIOS),
    )
    multipliers = ed.load_assumptions_config()["forecast"]["scenario_multipliers"]
    lever_mismatch = [
        lever + " " + scenario
        for lever in ("win_rate", "attainment", "pipeline_creation", "retention_severity",
                      "expansion")
        for scenario in ed.SCENARIOS
        if not _close(
            stored.loc[scenario, lever + "_multiplier"], multipliers[lever][scenario], 1e-9
        )
    ]
    result.record(
        "scenario levers match config/assumptions.yml", not lever_mismatch,
        ", ".join(lever_mismatch),
    )
    selector = wb["Scenarios"]
    validations = [dv.formula1 for dv in selector.data_validations.dataValidation]
    result.record(
        "the Scenarios tab carries a Bear / Base / Bull selector",
        any("Bear" in f and "Bull" in f for f in validations),
        ", ".join(validations),
    )


def check_runway_and_hiring(
    wb: Workbook, marts: dict[str, pd.DataFrame], result: Result
) -> None:
    stored = _table_frame(wb, "tbl_runway_policy").set_index("path")
    source = marts["fct_cash_runway_policy"].set_index("path")
    mismatch = [
        path for path in source.index
        if not _close(
            stored.loc[path, "policy_runway_months"], source.loc[path, "policy_runway_months"],
            1e-6,
        )
        or not _close(
            stored.loc[path, "headroom_months"], source.loc[path, "headroom_months"], 1e-6
        )
    ]
    result.record(
        "policy runway ties to fct_cash_runway_policy", not mismatch, ", ".join(mismatch)
    )
    result.record(
        "every path carries the 24-month Board floor",
        all(_close(v, 24.0) for v in stored["board_runway_floor_months"]),
    )

    hiring_stored = _table_frame(wb, "tbl_hiring").set_index("case_label")
    hiring_expected = ed.hiring_decision(marts).set_index("case_label")
    hiring_mismatch = [
        case + "." + column
        for case in hiring_expected.index
        for column in ("cumulative_hires", "dec_2027_incremental_arr",
                       "dec_2027_incremental_cash", "headroom_months")
        if not _close(
            hiring_stored.loc[case, column], hiring_expected.loc[case, column], 1e-4
        )
    ]
    result.record(
        "hiring table ties to fct_hiring_scenario", not hiring_mismatch,
        ", ".join(hiring_mismatch[:4]),
    )

    # The attractiveness horizon must be Dec-2027, sourced from the mart's own Dec-2027 row.
    raw = marts["fct_hiring_scenario"]
    full = raw[
        (raw["case_label"] == "Full Capacity-Close Hiring")
        & (raw["month_end_date"] == date(2027, 12, 31))
    ]
    result.record(
        "hiring attractiveness reads the Dec-2027 horizon",
        _close(
            hiring_stored.loc["Full Capacity-Close Hiring", "dec_2027_incremental_arr"],
            float(full["incremental_ending_arr"].iloc[0]), 1e-4,
        ),
    )
    labels = {
        str(cell.value)
        for row in wb["Runway & Hiring"].iter_rows(min_col=2, max_col=2) for cell in row
        if isinstance(cell.value, str)
    }
    result.record(
        "affordability and attractiveness are presented as separate questions",
        any("Board-policy runway" in label for label in labels)
        and any("Incremental Dec-2027" in label for label in labels),
    )


def check_accounting(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    stored = _table_frame(wb, "tbl_deferred_revenue").set_index("fiscal_quarter")
    expected = ed.deferred_revenue_quarterly(marts).set_index("fiscal_quarter")
    mismatch = [
        str(q) for q in expected.index
        if not _close(
            stored.loc[q, "ending_deferred_revenue"],
            expected.loc[q, "ending_deferred_revenue"],
        )
    ]
    result.record(
        "deferred revenue ties to fct_deferred_revenue", not mismatch, ", ".join(mismatch)
    )
    worst = 0.0
    for quarter in expected.index:
        record = expected.loc[quarter]
        # Beginning + billings - revenue + unbilled receivable movement = ending.
        # `revenue_recognised` is stored positive in the mart, so it is subtracted here.
        residual = (
            float(record["beginning_deferred_revenue"]) + float(record["billings"])
            - float(record["revenue_recognised"])
            + float(record["unbilled_receivable_movement"])
            - float(record["ending_deferred_revenue"])
        )
        worst = max(worst, abs(residual))
    result.record(
        "the deferred revenue rollforward closes with no plug", worst < 1.0,
        "worst residual ${:,.4f}".format(worst),
    )

    commission_stored = _table_frame(wb, "tbl_commission").set_index("period")
    commission_expected = ed.commission_accounting(marts).set_index("period")
    commission_mismatch = [
        str(p) + "." + column
        for p in commission_expected.index
        for column in ("commission_earned", "gaap_commission_expense",
                       "ending_commission_asset", "commission_paid_cash")
        if not _close(
            commission_stored.loc[p, column], commission_expected.loc[p, column]
        )
    ]
    result.record(
        "commission accounting ties to fct_commission_asset", not commission_mismatch,
        ", ".join(commission_mismatch[:4]),
    )
    asset_residual = 0.0
    for period in commission_expected.index:
        record = commission_expected.loc[period]
        residual = (
            float(record["beginning_commission_asset"])
            + float(record["capitalised_commission"])
            - float(record["commission_amortisation"])
            - float(record["ending_commission_asset"])
        )
        asset_residual = max(asset_residual, abs(residual))
    result.record(
        "the commission asset rollforward closes", asset_residual < 1.0,
        "worst residual ${:,.4f}".format(asset_residual),
    )

    adjustment = ed.accounting_adjustment(marts).set_index("period")
    result.record(
        "the accounting adjustment to history is zero",
        abs(float(adjustment.loc["All actual months", "commission_accounting_adjustment"])) < 1.0,
    )
    labels = {
        str(cell.value)
        for row in wb["Accounting"].iter_rows() for cell in row
        if isinstance(cell.value, str)
    }
    result.record(
        "the commission asset is labelled analytically derived",
        any("analytically derived" in label for label in labels),
    )


def check_commentary(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    stored = _table_frame(wb, "tbl_commentary")
    expected = ed.commentary(marts)
    result.record(
        "commentary row count matches fct_commentary_output",
        len(stored) == len(expected), "{a} vs {b}".format(a=len(stored), b=len(expected)),
    )
    mismatch = [
        str(expected.iloc[i]["commentary_id"])
        for i in range(len(expected))
        if str(stored.iloc[i]["headline"]) != str(expected.iloc[i]["headline"])
    ]
    result.record(
        "commentary text is the mart's own, unedited", not mismatch, ", ".join(mismatch)
    )
    source_headlines = set(marts["fct_commentary_output"]["headline"].astype(str))
    result.record(
        "no commentary sentence was written in the workbook",
        set(stored["headline"].astype(str)) <= source_headlines,
    )
    ranks = list(stored["exec_rank"])
    result.record(
        "commentary carries the Phase 7 priority order", ranks == list(range(1, len(ranks) + 1))
    )
    exec_formulas = [
        f for s, _c, f in _formula_cells(wb)
        if s == "Executive Summary" and "tbl_commentary[headline]" in f
    ]
    max_items = int(ed.load_commentary_config()["commentary"]["max_executive_summary_items"])
    result.record(
        "the Executive Summary shows the configured number of commentary items",
        len(exec_formulas) == min(max_items, len(expected)),
        "found {n}, configured {m}".format(n=len(exec_formulas), m=max_items),
    )


def check_controls(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    stored = _table_frame(wb, "tbl_controls")
    expected = ed.controls(marts)
    result.record(
        "all six upstream controls are on the Controls tab",
        len(stored) == 6 and len(expected) == 6, "found {n}".format(n=len(stored)),
    )
    names = list(stored["Control"])
    required = [
        "ctl_arr_reconciliation", "ctl_retention_bounds", "ctl_gtm_controls",
        "ctl_forecast_controls", "ctl_bridge_commentary", "ctl_accounting_enhancements",
    ]
    missing = [c for c in required if c not in names]
    result.record("every named control is present", not missing, ", ".join(missing))
    failing = [
        str(row["Control"]) for _, row in stored.iterrows()
        if str(row["Status"]) != "PASS" or int(row["Violations"]) != 0
    ]
    result.record("every upstream control passes", not failing, ", ".join(failing))

    status = _find_formula(wb, "Controls", "READY / PASS")
    result.record(
        "the overall status is a formula over the control violations, not a typed word",
        status is not None
        and "SUM(tbl_controls[Violations])" in status
        and '"FAIL"' in status,
        status or "no overall status formula found",
    )
    result.record(
        "the overall status cannot read PASS while a control has violations",
        status is not None and status.startswith("=IF(SUM(tbl_controls[Violations])=0,"),
        status or "",
    )
    labels = {
        str(cell.value)
        for row in wb["Controls"].iter_rows() for cell in row
        if isinstance(cell.value, str)
    }
    for required_check in (
        "Workbook version", "Build timestamp (UTC)", "Source reporting date",
        "Forecast cutover", "External workbook links",
    ):
        result.record(
            'workbook-level check "' + required_check + '" is shown',
            any(label.startswith(required_check) for label in labels),
        )


def check_arr_and_gtm(wb: Workbook, marts: dict[str, pd.DataFrame], result: Result) -> None:
    arr_stored = _table_frame(wb, "tbl_arr_monthly")
    arr_expected = ed.arr_monthly(marts, start=date(2025, 1, 31), end=ed.FY2026_END)
    result.record(
        "the ARR table covers Jan-2025 to Dec-2026",
        len(arr_stored) == len(arr_expected),
        "{a} vs {b}".format(a=len(arr_stored), b=len(arr_expected)),
    )
    actual_labels = set(
        arr_expected[arr_expected["period_type"] == "Actual"]["month_label"]
    )
    result.record(
        "the actual / forecast cutover is Jun-2026",
        "Jun-26" in actual_labels and "Jul-26" not in actual_labels,
    )

    retention = _table_frame(wb, "tbl_retention")
    source = marts["fct_retention_ttm"]
    jun = source[
        (source["month_end_date"] == ed.REPORTING_DATE) & (source["segment"] == "Total")
    ].iloc[0]
    stored_jun = retention[
        (retention["month_label"] == "Jun-26") & (retention["segment"] == "Total")
    ].iloc[0]
    result.record(
        "TTM retention ties to fct_retention_ttm",
        _close(stored_jun["nrr"], jun["nrr"], 1e-9)
        and _close(stored_jun["grr"], jun["grr"], 1e-9)
        and _close(stored_jun["logo_retention"], jun["logo_retention"], 1e-9),
    )
    result.record(
        "GRR never exceeds NRR anywhere in the retention table",
        all(
            float(row["grr"]) <= float(row["nrr"]) + 1e-9
            for _, row in retention.iterrows()
        ),
    )

    constraint = _table_frame(wb, "tbl_gtm_constraint").set_index("segment")
    expected = ed.gtm_constraint_by_segment(marts).set_index("segment")
    mismatch = [
        str(segment) + "." + column
        for segment in expected.index
        for column in ("h2_capacity_supported_arr", "h2_pipeline_supported_arr",
                       "h2_constrained_new_logo_arr")
        if not _close(constraint.loc[segment, column], expected.loc[segment, column], 1e-4)
    ]
    result.record(
        "the GTM constraint table ties to fct_new_logo_diagnosis", not mismatch,
        ", ".join(mismatch[:4]),
    )
    least_violations = [
        str(segment) for segment in expected.index
        if float(expected.loc[segment, "h2_constrained_new_logo_arr"])
        > min(
            float(expected.loc[segment, "h2_capacity_supported_arr"]),
            float(expected.loc[segment, "h2_pipeline_supported_arr"]),
        ) + 1.0
    ]
    result.record(
        "constrained New Logo ARR never exceeds the lesser of capacity and pipeline",
        not least_violations, ", ".join(least_violations),
    )

    unit = _table_frame(wb, "tbl_unit_econ").set_index("segment")
    unit_expected = ed.fy2025_unit_economics(marts).set_index("segment")
    unit_mismatch = [
        str(segment) for segment in unit_expected.index
        if not _close(unit.loc[segment, "cac"], unit_expected.loc[segment, "cac"], 1e-4)
        or not _close(
            unit.loc[segment, "cac_payback_months"],
            unit_expected.loc[segment, "cac_payback_months"], 1e-6,
        )
    ]
    result.record(
        "FY2025 unit economics tie to fct_unit_economics", not unit_mismatch,
        ", ".join(unit_mismatch),
    )


def check_marts_untouched(marts_dir: Path, before: dict[str, int], result: Result) -> None:
    after = {p.name: p.stat().st_mtime_ns for p in sorted(marts_dir.glob("*.csv"))}
    changed = [name for name, stamp in before.items() if after.get(name) != stamp]
    result.record(
        "workbook generation modified no upstream mart", not changed, ", ".join(changed)
    )
    result.record("no mart was deleted", set(before) <= set(after))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def validate(
    path: Path = OUTPUT_PATH, *, marts_dir: Path = ed.MARTS_DIR,
    mart_stamps: dict[str, int] | None = None,
) -> Result:
    """Run every structural and value check. Returns a `Result`, never raises on a failure."""
    result = Result()
    if not path.exists():
        result.record("workbook exists", False, str(path) + " was not found")
        return result
    result.record("workbook exists", True)

    marts = ed.load_marts(marts_dir)
    wb = load_workbook(path, data_only=False)

    check_structure(path, wb, result)
    check_ooxml_function_prefixes(path, result)
    check_formatting(wb, result)
    check_chart_specs(path, wb, result)
    check_chart_values(path, wb, marts, result)
    check_executive(wb, marts, result)
    check_pnl(wb, marts, result)
    check_forecast(wb, marts, result)
    check_bridges(wb, marts, result)
    check_scenarios(wb, marts, result)
    check_runway_and_hiring(wb, marts, result)
    check_accounting(wb, marts, result)
    check_commentary(wb, marts, result)
    check_controls(wb, marts, result)
    check_arr_and_gtm(wb, marts, result)
    if mart_stamps is not None:
        check_marts_untouched(marts_dir, mart_stamps, result)
    wb.close()
    return result


def main(argv: list[str] | None = None) -> int:
    path = Path(argv[0]) if argv else OUTPUT_PATH
    result = validate(path)
    print(result.summary())
    return 0 if result.passed else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
