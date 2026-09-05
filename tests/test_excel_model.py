"""Phase 9 -- the recruiter-facing Excel FP&A operating model.

Two things are tested here, and they are different questions.

1. **The build is reproducible and correct.** The workbook is rebuilt from the committed marts
   into a temporary directory and put through `src/validate_excel_model.py` in full: structure,
   external links, formulas, charts, and an independent Python recomputation of every headline
   figure from the marts.

2. **The two artifacts are the ones they claim to be.** The builder writes
   `build/generated/...xlsx`; `excel/Helio_SaaS_FP&A_Operating_Model.xlsx` holds the REVIEWED
   workbook, restructured in native Excel after the build. The generated one is put through
   the full structural validation when it is present. The reviewed one deliberately does not
   match that structure - its Executive Summary was rebuilt, tables were demoted and charts
   re-placed - so it is checked for the things that must still hold: it exists, it opens, and
   every control in it reads PASS.

`openpyxl` does not calculate formulas, so no test in this file claims a formula's *result* was
read back from the file. Formula strings are asserted structurally; the values they should
produce are recomputed in Python from the marts. See `src/validate_excel_model.py`.
"""

from __future__ import annotations

import re
import zipfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import excel_data as ed
from src import excel_style as st
from src import validate_excel_model as vx
from src.build_excel_model import (
    DATA_SHEETS,
    GENERATED_PATH,
    PUBLISHED_PATH,
    VISIBLE_SHEETS,
    build,
)



@pytest.fixture(scope="module")
def marts() -> dict:
    return ed.load_marts()


@pytest.fixture(scope="module")
def mart_stamps() -> dict[str, int]:
    return {p.name: p.stat().st_mtime_ns for p in sorted(ed.MARTS_DIR.glob("*.csv"))}


@pytest.fixture(scope="module")
def rebuilt(tmp_path_factory, mart_stamps) -> Path:
    """Rebuild the workbook from the committed marts into a temporary directory."""
    out = tmp_path_factory.mktemp("excel") / "Helio_SaaS_FP&A_Operating_Model.xlsx"
    return build(output=out, verbose=False)


@pytest.fixture(scope="module")
def rebuilt_result(rebuilt, mart_stamps) -> vx.Result:
    return vx.validate(rebuilt, mart_stamps=mart_stamps)


# ---------------------------------------------------------------------------
# The build itself
# ---------------------------------------------------------------------------
def test_workbook_is_produced_and_is_a_valid_xlsx(rebuilt: Path) -> None:
    assert rebuilt.exists() and rebuilt.stat().st_size > 0
    with zipfile.ZipFile(rebuilt) as archive:
        assert archive.testzip() is None
        assert "xl/workbook.xml" in archive.namelist()


def test_every_validation_check_passes_on_a_fresh_build(rebuilt_result: vx.Result) -> None:
    assert rebuilt_result.passed, rebuilt_result.summary()


def test_generated_workbook_is_current_with_the_marts() -> None:
    """The builder's own output, when it is on disk, must survive the same validation."""
    if not GENERATED_PATH.exists():
        pytest.skip(
            f"{GENERATED_PATH} not built. The reproducibility test above already builds and "
            "validates a fresh workbook in a temporary directory; this one only re-checks a "
            "build artefact left in the tree. Run `python -m src.build_excel_model`."
        )
    result = vx.validate(GENERATED_PATH)
    assert result.passed, result.summary()


def test_the_published_workbook_is_the_reviewed_one_and_its_controls_pass() -> None:
    """`excel/` holds the reviewed workbook, not the builder's output.

    It is not put through `validate_excel_model`: that validator encodes the generated
    layout, and the reviewed workbook deliberately departs from it. What must still be true
    is that it is present, that it opens, and that nothing in it reports a failed control.
    """
    assert PUBLISHED_PATH.exists(), (
        f"{PUBLISHED_PATH} is missing. It is the reviewed portfolio workbook and is copied "
        "in by hand - `python -m src.build_excel_model` does not write it."
    )
    wb = load_workbook(PUBLISHED_PATH, data_only=True)
    try:
        # Anchored on the header label "Status", never on an address: a restructured sheet
        # silently stops matching an address, and a bare PASS/FAIL scan picks up things that
        # are not controls at all - the Model Guide's legend swatch, and the Executive
        # Summary's decision band, where a Bear FAIL is the correct answer.
        statuses: list[str] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            heads = [
                c
                for row in ws.iter_rows()
                for c in row
                if isinstance(c.value, str) and c.value.strip() == "Status"
            ]
            for head in heads:
                for r in range(head.row + 1, min(head.row + 30, ws.max_row) + 1):
                    value = ws.cell(row=r, column=head.column).value
                    text = value.strip() if isinstance(value, str) else ""
                    if text in ("PASS", "FAIL"):
                        statuses.append(text)
                    elif not str(ws.cell(row=r, column=2).value or "").strip():
                        break
        assert len(statuses) == 26, (
            f"expected 26 controls in the reviewed workbook, found {len(statuses)}"
        )
        assert "FAIL" not in statuses, (
            f"the reviewed workbook reports a failed control "
            f"({statuses.count('FAIL')} of {len(statuses)})"
        )
    finally:
        wb.close()


def test_the_build_fails_loudly_when_a_required_mart_is_missing(tmp_path: Path) -> None:
    empty = tmp_path / "marts"
    empty.mkdir()
    with pytest.raises(ed.MartError) as excinfo:
        build(marts_dir=empty, output=tmp_path / "out.xlsx", verbose=False)
    assert "fct_arr_waterfall" in str(excinfo.value) or "missing" in str(excinfo.value)
    assert not (tmp_path / "out.xlsx").exists(), "a blank workbook was written anyway"


def test_workbook_generation_modifies_no_upstream_mart(rebuilt_result: vx.Result) -> None:
    names = {name for name, _ok, _detail in rebuilt_result.checks}
    assert "workbook generation modified no upstream mart" in names
    failed = [
        name for name, ok, _detail in rebuilt_result.checks
        if name.startswith("workbook generation modified") and not ok
    ]
    assert not failed


# ---------------------------------------------------------------------------
# Structure
# ---------------------------------------------------------------------------
def test_expected_visible_sheets_exist_in_order_with_no_duplicates(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    titles = wb.sheetnames
    assert len(titles) == len(set(titles))
    assert [t for t in titles if t in VISIBLE_SHEETS] == VISIBLE_SHEETS
    assert all(wb[name].sheet_state == "visible" for name in VISIBLE_SHEETS)
    wb.close()


def test_supporting_data_sheets_are_hidden_but_remain_accessible(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    for name in DATA_SHEETS:
        assert wb[name].sheet_state == "hidden", name + " is not hidden"
        assert wb[name].sheet_state != "veryHidden"
        assert wb[name].max_row > 1, name + " is empty"
        assert not wb[name].protection.sheet
    wb.close()


def test_no_external_links_and_no_macros(rebuilt: Path) -> None:
    with zipfile.ZipFile(rebuilt) as archive:
        names = archive.namelist()
    assert not [n for n in names if n.startswith("xl/externalLinks/")]
    assert not [n for n in names if "vbaProject" in n or n.endswith(".bin")]

    wb = load_workbook(rebuilt)
    offenders = [
        sheet + "!" + coord
        for sheet, coord, formula in vx._formula_cells(wb)
        if vx.EXTERNAL_REF.search(formula) or "#REF!" in formula
    ]
    wb.close()
    assert not offenders, offenders[:5]


def test_no_volatile_or_banned_function_is_used(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    offenders = [
        sheet + "!" + coord + " " + fn
        for sheet, coord, formula in vx._formula_cells(wb)
        for fn in vx.BANNED_FUNCTIONS
        if fn in formula.upper()
    ]
    wb.close()
    assert not offenders, offenders[:5]


# ---------------------------------------------------------------------------
# Values -- every one recomputed from the marts, not read back from a formula
# ---------------------------------------------------------------------------
def test_executive_scorecard_ties_to_the_management_variance_mart(
    rebuilt: Path, marts: dict
) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_mgmt_variance").set_index("metric")
    wb.close()
    source = marts["fct_management_variance"].set_index("metric")
    for metric in source.index:
        assert vx._close(stored.loc[metric, "budget_amount"], source.loc[metric, "budget_amount"])
        assert vx._close(stored.loc[metric, "base_amount"], source.loc[metric, "base_amount"])
        assert (
            stored.loc[metric, "favorable_unfavorable"]
            == source.loc[metric, "favorable_unfavorable"]
        )


def test_exit_arr_waterfall_ties_to_the_phase_7_bridge(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_arr_bridge")
    wb.close()
    expected = ed.arr_bridge(marts)
    assert len(stored) == len(expected)
    for i in range(len(expected)):
        assert vx._close(stored.iloc[i]["amount"], expected.iloc[i]["amount"])
    anchor = float(expected[expected["line_kind"] == "anchor"]["amount"].iloc[0])
    components = float(expected[expected["line_kind"] == "component"]["amount"].sum())
    outcome = float(expected[expected["line_kind"] == "result"]["amount"].iloc[0])
    assert abs(outcome - (anchor + components)) < 1.0

    variance = marts["fct_management_variance"].set_index("metric")
    assert vx._close(outcome, variance.loc["exit_arr", "base_amount"])


def test_pnl_ties_to_the_phase_6_reforecast(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_pnl_summary").set_index("line_key")
    wb.close()
    expected = ed.pnl_summary(marts).set_index("line_key")
    for key in expected.index:
        for column in ("fy2025_actual", "h1_2026_actual", "h2_2026_base", "fy2026_base",
                       "fy2026_budget"):
            tolerance = 1e-9 if key.endswith("_pct") else vx.TOLERANCE
            assert vx._close(stored.loc[key, column], expected.loc[key, column], tolerance), (
                key + "." + column
            )
    # FY2026 Base is H1 realised actual plus the H2 reforecast, not a separately stored figure.
    assert vx._close(
        expected.loc["total_revenue", "fy2026_base"],
        expected.loc["total_revenue", "h1_2026_actual"]
        + expected.loc["total_revenue", "h2_2026_base"],
    )


def test_scenario_table_ties_to_fct_scenario_monthly(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_scenario_summary").set_index("scenario")
    wb.close()
    monthly = marts["fct_scenario_monthly"]
    for scenario in ed.SCENARIOS:
        dec26 = monthly[
            (monthly["scenario"] == scenario) & (monthly["month_end_date"] == ed.FY2026_END)
        ]
        dec27 = monthly[
            (monthly["scenario"] == scenario)
            & (monthly["month_end_date"] == date(2027, 12, 31))
        ]
        assert vx._close(stored.loc[scenario, "dec_2026_exit_arr"], dec26["ending_arr"].iloc[0])
        assert vx._close(stored.loc[scenario, "dec_2027_exit_arr"], dec27["ending_arr"].iloc[0])


def test_policy_runway_ties_to_fct_cash_runway_policy(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_runway_policy").set_index("path")
    wb.close()
    source = marts["fct_cash_runway_policy"].set_index("path")
    for path in source.index:
        assert vx._close(
            stored.loc[path, "policy_runway_months"],
            source.loc[path, "policy_runway_months"], 1e-6,
        )
        assert vx._close(
            stored.loc[path, "headroom_months"], source.loc[path, "headroom_months"], 1e-6
        )
        assert vx._close(stored.loc[path, "board_runway_floor_months"], 24.0)
    # Affordability is answered on the Board-policy view, never on the operating cash proxy.
    assert "fct_cash_runway" not in set(stored.columns)


def test_hiring_table_ties_to_fct_hiring_scenario_on_the_fy2027_horizon(
    rebuilt: Path, marts: dict
) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_hiring").set_index("case_label")
    wb.close()
    raw = marts["fct_hiring_scenario"]
    for case in ed.HIRING_CASE_ORDER:
        dec27 = raw[
            (raw["case_label"] == case) & (raw["month_end_date"] == date(2027, 12, 31))
        ]
        assert vx._close(
            stored.loc[case, "dec_2027_incremental_arr"],
            dec27["incremental_ending_arr"].iloc[0], 1e-4,
        )
        assert vx._close(
            stored.loc[case, "cumulative_hires"], dec27["cumulative_hires"].iloc[0], 1e-9
        )
    # The near-term ramp figures are carried separately and are not the decision view.
    assert "dec_2026_incremental_arr" in stored.columns


def test_accounting_summary_ties_to_phase_8(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    deferred = vx._table_frame(wb, "tbl_deferred_revenue").set_index("fiscal_quarter")
    commission = vx._table_frame(wb, "tbl_commission").set_index("period")
    wb.close()

    expected_dr = ed.deferred_revenue_quarterly(marts).set_index("fiscal_quarter")
    for quarter in expected_dr.index:
        record = expected_dr.loc[quarter]
        assert vx._close(
            deferred.loc[quarter, "ending_deferred_revenue"],
            record["ending_deferred_revenue"],
        )
        residual = (
            float(record["beginning_deferred_revenue"]) + float(record["billings"])
            - float(record["revenue_recognised"])
            + float(record["unbilled_receivable_movement"])
            - float(record["ending_deferred_revenue"])
        )
        assert abs(residual) < 1.0, str(quarter) + " rollforward residual " + str(residual)

    expected_commission = ed.commission_accounting(marts).set_index("period")
    for period in expected_commission.index:
        record = expected_commission.loc[period]
        assert vx._close(
            commission.loc[period, "gaap_commission_expense"], record["gaap_commission_expense"]
        )
        rollforward = (
            float(record["beginning_commission_asset"])
            + float(record["capitalised_commission"])
            - float(record["commission_amortisation"])
            - float(record["ending_commission_asset"])
        )
        assert abs(rollforward) < 1.0, str(period) + " asset rollforward " + str(rollforward)


def test_commentary_comes_from_the_mart_and_is_never_written_in_the_workbook(
    rebuilt: Path, marts: dict
) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_commentary")
    exec_lookups = [
        formula for sheet, _coord, formula in vx._formula_cells(wb)
        if sheet == "Executive Summary" and "tbl_commentary[headline]" in formula
    ]
    wb.close()

    source = marts["fct_commentary_output"]
    assert len(stored) == len(source)
    assert set(stored["headline"].astype(str)) == set(source["headline"].astype(str))
    # The Executive Summary retrieves commentary; it never restates it.
    configured = int(ed.load_commentary_config()["commentary"]["max_executive_summary_items"])
    assert len(exec_lookups) == min(configured, len(source))
    assert list(stored["exec_rank"]) == list(range(1, len(stored) + 1))


def test_overall_control_status_reflects_the_upstream_controls(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    controls = vx._table_frame(wb, "tbl_controls")
    status = vx._find_formula(wb, "Controls", "READY / PASS")
    wb.close()

    assert len(controls) == 6
    assert set(controls["Control"]) == {
        "ctl_arr_reconciliation", "ctl_retention_bounds", "ctl_gtm_controls",
        "ctl_forecast_controls", "ctl_bridge_commentary", "ctl_accounting_enhancements",
    }
    assert all(int(v) == 0 for v in controls["Violations"])
    assert all(str(s) == "PASS" for s in controls["Status"])
    # Structural, not cosmetic: the headline cannot read PASS unless the violations sum to zero.
    assert status is not None
    assert status.startswith('=IF(SUM(tbl_controls[Violations])=0,')
    assert '"FAIL"' in status


def test_formula_driven_variance_fields_are_structurally_correct(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    ws = wb["P&L"]
    # Find the Total Revenue row, then check its variance, variance % and Fav/Unfav formulas.
    target = None
    for row in ws.iter_rows(min_col=2, max_col=2):
        if row[0].value == "Total Revenue":
            target = row[0].row
            break
    assert target is not None
    variance = ws.cell(row=target, column=8).value
    variance_pct = ws.cell(row=target, column=9).value
    fav = ws.cell(row=target, column=10).value
    wb.close()

    assert variance == "=F{r}-G{r}".format(r=target), variance          # Base less Budget
    assert variance_pct == '=IF(G{r}=0,"",F{r}/G{r}-1)'.format(r=target), variance_pct
    # Plain nested IF over XLOOKUP -- no LET, so no declared name to namespace.
    assert fav.startswith("=IF(_xlfn.XLOOKUP("), fav
    assert "LET(" not in fav, fav
    assert "tbl_pnl_summary[polarity]" in fav                            # centralised polarity
    assert '"contextual"' in fav                                         # never labelled F/U
    assert '"Favorable"' in fav and '"Unfavorable"' in fav
    assert "H{r}".format(r=target) in fav                                # reads the variance cell


def test_gtm_shows_capacity_pipeline_and_the_binding_constraint(
    rebuilt: Path, marts: dict
) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_gtm_constraint").set_index("segment")
    wb.close()
    expected = ed.gtm_constraint_by_segment(marts).set_index("segment")
    for segment in expected.index:
        capacity = float(expected.loc[segment, "h2_capacity_supported_arr"])
        pipeline = float(expected.loc[segment, "h2_pipeline_supported_arr"])
        constrained = float(expected.loc[segment, "h2_constrained_new_logo_arr"])
        assert vx._close(
            stored.loc[segment, "h2_constrained_new_logo_arr"], constrained, 1e-4
        )
        # The central point of the tab: capacity alone does not equal achievable bookings.
        assert constrained <= min(capacity, pipeline) + 1.0
    assert str(expected.loc["Total", "primary_binding_constraint"]) in {"Pipeline", "Capacity"}


def test_the_actual_to_forecast_cutover_is_30_june_2026(rebuilt: Path, marts: dict) -> None:
    wb = load_workbook(rebuilt)
    stored = vx._table_frame(wb, "tbl_arr_monthly")
    wb.close()
    actual = {
        str(row["month_label"]) for _, row in stored.iterrows()
        if str(row["period_type"]) == "Actual"
    }
    forecast = {
        str(row["month_label"]) for _, row in stored.iterrows()
        if str(row["period_type"]) == "Base Reforecast"
    }
    assert "Jun-26" in actual and "Jul-26" not in actual
    assert "Jul-26" in forecast and "Dec-26" in forecast
    assert not actual & forecast


# ---------------------------------------------------------------------------
# OOXML function namespacing -- read from the saved package, not from openpyxl
# ---------------------------------------------------------------------------
# Functions introduced after Excel 2007 are stored under the `_xlfn.` future-function
# namespace. Excel writes the prefix itself and hides it in the formula bar; openpyxl writes
# whatever string it is handed, verbatim. A bare `XLOOKUP(` in the worksheet XML reaches Excel
# as an unrecognised defined name and every cell using it renders #NAME? on open.
#
# These tests read the XML out of the ZIP rather than asking openpyxl what the formula says,
# because the defect lives in what was written, not in what Python reports back.
def _worksheet_xml(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        return {
            name: archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet")
        }


def _stored_formulas(path: Path) -> list[str]:
    formulas = []
    for xml in _worksheet_xml(path).values():
        for match in re.finditer(r"<f[^>]*>(.*?)</f>", xml, re.S):
            formulas.append(match.group(1))
    return formulas


def _code_only(formula: str) -> str:
    """Blank string literals, so a lookup key is never mistaken for a function call."""
    return re.sub(r'"(?:[^"]|"")*"', '""', formula)


def test_xlookup_and_let_are_namespaced_in_the_worksheet_xml(rebuilt: Path) -> None:
    """The reported defect, asserted directly: bare XLOOKUP or LET in the XML must fail."""
    formulas = _stored_formulas(rebuilt)
    assert formulas, "no formulas found in the worksheet XML"

    for name in ("XLOOKUP", "LET"):
        bare = [
            f for f in formulas
            if re.search(r"(?<![A-Za-z0-9_.])" + name + r"\s*\(", _code_only(f))
        ]
        assert not bare, name + " emitted without its _xlfn. prefix: " + str(bare[:2])

    namespaced = [f for f in formulas if "_xlfn.XLOOKUP(" in f]
    assert len(namespaced) > 400, "expected _xlfn.XLOOKUP throughout, found " + str(len(namespaced))


def test_every_modern_function_carries_its_required_namespace(rebuilt: Path) -> None:
    """Generalised: every function in `MODERN_FUNCTIONS` must be stored as that mapping says."""
    offenders = []
    for formula in _stored_formulas(rebuilt):
        for match in vx.STORED_FUNCTION_CALL.finditer(_code_only(formula)):
            namespace, name = match.group(1), match.group(2).upper()
            required = st.MODERN_FUNCTIONS.get(name)
            if required is not None and (namespace + name).upper() != required.upper():
                offenders.append(namespace + name + " should be " + required)
    assert not offenders, sorted(set(offenders))[:5]


def test_no_formula_uses_a_function_excel_will_not_recognise(rebuilt: Path) -> None:
    """All 668 formulas, every sheet: no unclassified name, no namespace on a legacy function."""
    formulas = _stored_formulas(rebuilt)
    assert len(formulas) > 600, "expected the full formula population, found " + str(len(formulas))

    unclassified, spurious = [], []
    for formula in formulas:
        for match in vx.STORED_FUNCTION_CALL.finditer(_code_only(formula)):
            namespace, name = match.group(1), match.group(2).upper()
            if name in st.MODERN_FUNCTIONS:
                continue
            if name in vx.LEGACY_FUNCTIONS:
                if namespace:
                    spurious.append(namespace + name)
            else:
                unclassified.append(namespace + name)
    assert not unclassified, (
        "unclassified function(s) " + str(sorted(set(unclassified))[:5])
        + " -- add to excel_style.MODERN_FUNCTIONS or validate_excel_model.LEGACY_FUNCTIONS"
    )
    assert not spurious, sorted(set(spurious))[:5]


def test_the_namespace_check_fails_on_a_workbook_with_the_prefix_stripped(
    rebuilt: Path, tmp_path: Path
) -> None:
    """Mutation test: the guard must actually fire, not merely be present.

    A check that never fails proves nothing. This rewrites the saved package with every
    `_xlfn.` removed -- exactly the defect that was reported -- and asserts that validation
    rejects it.
    """
    mutant = tmp_path / "mutant.xlsx"
    with zipfile.ZipFile(rebuilt) as source, zipfile.ZipFile(
        mutant, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                data = (
                    data.decode("utf-8")
                    .replace("_xlfn._xlws.", "")
                    .replace("_xlfn.", "")
                    .encode("utf-8")
                )
            target.writestr(item, data)

    result = vx.validate(mutant)
    assert not result.passed, "stripping every _xlfn. prefix was not detected"
    failed = {name for name, ok, _detail in result.checks if not ok}
    assert "every modern function carries its OOXML namespace prefix" in failed
    assert "XLOOKUP is stored as _xlfn.XLOOKUP" in failed


def test_excel_recalculates_the_workbook_on_open(rebuilt: Path) -> None:
    """Namespacing only matters if Excel actually evaluates the formulas when the file opens."""
    wb = load_workbook(rebuilt)
    assert wb.calculation.fullCalcOnLoad is True
    wb.close()
    with zipfile.ZipFile(rebuilt) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
    assert "fullCalcOnLoad=\"1\"" in workbook_xml, workbook_xml[-400:]


# ---------------------------------------------------------------------------
# Declared names -- the defect Excel reported as "Removed Records: Formula"
# ---------------------------------------------------------------------------
# `_xlfn.` namespaces a FUNCTION name. A name DECLARED by LET or LAMBDA needs a second,
# different namespace, `_xlpm.`. Getting the first right and the second wrong still produces a
# formula Excel refuses: it cannot resolve the declared name, drops the whole record, and opens
# the file with "Removed Records: Formula from /xl/worksheets/sheetN.xml".
#
# That is exactly what happened to the fourteen P&L Fav / Unfav cells, which were written as
# `_xlfn.LET(v,H10,p,_xlfn.XLOOKUP(...),...)` -- correct function names, bare parameter names.
def _sheet_part_for(path: Path, sheet_name: str) -> str:
    """Resolve a worksheet's XML part from the workbook's own sheet order."""
    with zipfile.ZipFile(path) as archive:
        names = re.findall(r'<sheet name="([^"]+)"', archive.read("xl/workbook.xml").decode("utf-8"))
    escaped = sheet_name.replace("&", "&amp;")
    assert escaped in names, sheet_name + " not found in " + str(names)
    return "xl/worksheets/sheet{n}.xml".format(n=names.index(escaped) + 1)


def test_sheet5_is_the_pnl_worksheet(rebuilt: Path) -> None:
    """The part Excel named in the recovery message maps to the P&L tab."""
    assert _sheet_part_for(rebuilt, "P&L") == "xl/worksheets/sheet5.xml"


def test_pnl_fav_unfav_formulas_survive_in_the_saved_xml(rebuilt: Path) -> None:
    """The fourteen cells Excel removed must be present, namespaced, and free of declared names."""
    part = _sheet_part_for(rebuilt, "P&L")
    with zipfile.ZipFile(rebuilt) as archive:
        xml = archive.read(part).decode("utf-8")

    fav = re.findall(r'<c r="(J\d+)"[^>]*>(?:(?!</c>).)*?<f[^>]*>(.*?)</f>', xml, re.S)
    assert len(fav) == 14, "expected 14 Fav / Unfav formulas on P&L, found " + str(len(fav))

    for coord, formula in fav:
        assert "_xlfn.XLOOKUP(" in formula, coord + ": " + formula
        assert "tbl_pnl_summary[polarity]" in formula, coord
        assert "contextual" in formula, coord
        # No declared name means no `_xlpm.` obligation and nothing for Excel to reject.
        assert not re.search(r"(?<![A-Za-z0-9_.])LET\s*\(", formula), coord + " still uses LET"


def test_every_declared_name_is_namespaced_across_all_worksheets(rebuilt: Path) -> None:
    """Generalised guard: any LET / LAMBDA anywhere must declare `_xlpm.`-namespaced names."""
    offenders = []
    for name, xml in _worksheet_xml(rebuilt).items():
        for match in re.finditer(r"<f[^>]*>(.*?)</f>", xml, re.S):
            formula = match.group(1)
            for function in st.PARAMETER_FUNCTIONS:
                for declared in vx._declared_names(formula, function):
                    if not declared.startswith(st.PARAMETER_NAMESPACE):
                        offenders.append(name + ": " + function + "(" + declared)
    assert not offenders, sorted(set(offenders))[:5]


def test_the_declared_name_guard_fails_on_a_bare_let_parameter(tmp_path: Path) -> None:
    """Mutation test for the reported defect, reproduced exactly.

    The workbook no longer ships a LET, so this builds the broken formula the P&L used to
    carry, writes it into a worksheet, and asserts the validator rejects it. Without this the
    guard would be untested code: nothing in the current workbook can trigger it.
    """
    from openpyxl import Workbook as _Workbook

    broken = (
        '_xlfn.LET(v,H10,p,_xlfn.XLOOKUP($B$10,t[line_item],t[polarity]),'
        'IF(p="contextual","n/a",IF(v=0,"-",'
        'IF((p="higher_favorable")=(v>0),"Favorable","Unfavorable"))))'
    )
    # The serializer would fix this, so it is written past it to reproduce the shipped defect.
    wb = _Workbook()
    wb.active.title = "P&L"
    wb.active["J10"] = "=" + broken
    path = tmp_path / "broken.xlsx"
    wb.save(path)
    wb.close()

    with zipfile.ZipFile(path) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    formula = re.search(r"<f[^>]*>(.*?)</f>", xml, re.S).group(1)
    declared = vx._declared_names(formula, "LET")
    assert declared == ["v", "p"], declared
    assert all(not d.startswith(st.PARAMETER_NAMESPACE) for d in declared)

    # And the same formula, put through the serializer, comes out correct.
    fixed = st.qualify_formula("=" + broken)
    assert "_xlfn.LET(_xlpm.v," in fixed, fixed
    assert "_xlpm.p" in fixed, fixed
    assert vx._declared_names(fixed, "LET") == ["_xlpm.v", "_xlpm.p"]


def test_declared_name_serialization_round_trips(  ) -> None:
    """Unit coverage for the `_xlpm.` serializer, including nesting and literals."""
    cases = {
        "=LET(x,1,y,2,x+y)": "=_xlfn.LET(_xlpm.x,1,_xlpm.y,2,_xlpm.x+_xlpm.y)",
        "=LET(rate,C5,LET(base,D5,base*rate))":
            "=_xlfn.LET(_xlpm.rate,C5,_xlfn.LET(_xlpm.base,D5,_xlpm.base*_xlpm.rate))",
        "=LAMBDA(a,b,a+b)": "=_xlfn.LAMBDA(_xlpm.a,_xlpm.b,_xlpm.a+_xlpm.b)",
        # A parameter name appearing inside a string literal must not be touched.
        '=LET(v,H10,IF(v>0,"v is positive","v is not"))':
            '=_xlfn.LET(_xlpm.v,H10,IF(_xlpm.v>0,"v is positive","v is not"))',
        # A function name inside a literal must not be touched either.
        '=XLOOKUP("LET",t[a],t[b])': '=_xlfn.XLOOKUP("LET",t[a],t[b])',
    }
    for source, expected in cases.items():
        produced = st.qualify_formula(source)
        assert produced == expected, source + " -> " + produced
        assert st.qualify_formula(produced) == produced, "not idempotent: " + produced
        assert st.display_formula(produced) == source, st.display_formula(produced)


# ---------------------------------------------------------------------------
# Formatting -- structural only
# ---------------------------------------------------------------------------
# These check the mechanics of the design system, not whether the result looks good. A
# gridline left on, a title at the wrong size, KPI cards out of alignment, a merged cell or a
# second chart size are all regressions a test can catch. Whether the workbook reads as a
# polished management model is a judgement that needs Excel open in front of a person, and
# nothing here claims otherwise.
def test_presentation_tabs_share_the_page_grid(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    for name in VISIBLE_SHEETS:
        ws = wb[name]
        assert not ws.sheet_view.showGridLines, name + " has gridlines on"
        assert ws.freeze_panes, name + " has no frozen panes"
        assert ws.print_area, name + " has no print area"
        assert ws.cell(row=1, column=1).value is None, name + " writes into the margin column"
        assert abs((ws.column_dimensions["A"].width or 0) - st.MARGIN_WIDTH) < 0.01, name
        assert ws.cell(row=1, column=st.CONTENT_COL).value == "Helio Systems, Inc.", name
    wb.close()


def test_no_merged_cells_anywhere(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    offenders = [ws.title for ws in wb.worksheets if ws.merged_cells.ranges]
    wb.close()
    assert not offenders, offenders


def test_one_font_family_and_the_approved_title_style(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    fonts, sizes = set(), set()
    for name in VISIBLE_SHEETS:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.value is not None and cell.font and cell.font.name:
                    fonts.add(cell.font.name)
                    sizes.add(cell.font.sz)
        title = wb[name].cell(row=1, column=st.CONTENT_COL)
        assert title.font.sz == st.TEXT_STYLES["title"]["size"], name
        assert title.font.b, name
    wb.close()
    assert fonts == {st.FONT_NAME}, fonts
    # Every size in use must come from the scale; nothing ad hoc.
    approved = {float(style["size"]) for style in st.TEXT_STYLES.values()}
    assert {float(s) for s in sizes} <= approved, sorted({float(s) for s in sizes} - approved)


def test_kpi_cards_share_one_shape(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    ws = wb["Executive Summary"]
    label_rows = [
        row for row in range(1, 40)
        if isinstance(ws.cell(row=row, column=st.CONTENT_COL).value, str)
        and ws.cell(row=row, column=st.CONTENT_COL).font.sz == st.TEXT_STYLES["kpi_label"]["size"]
        and ws.cell(row=row, column=st.CONTENT_COL).font.b
    ]
    assert len(label_rows) == 2, label_rows
    shapes, columns = set(), set()
    for row in label_rows:
        shapes.add((
            ws.row_dimensions[row].height,
            ws.row_dimensions[row + 1].height,
            ws.row_dimensions[row + 2].height,
        ))
        found = [
            col for col in range(st.CONTENT_COL, st.CONTENT_COL + 12)
            if ws.cell(row=row, column=col).value is not None
        ]
        columns.add(tuple(found))
    wb.close()
    assert len(shapes) == 1, shapes                 # identical heights
    assert len(columns) == 1, columns               # identical column positions
    assert len(next(iter(columns))) == 5, columns   # five cards per row


def test_every_chart_uses_a_standard_size_and_no_excel_style(rebuilt: Path) -> None:
    """Sizes come from the three named standards, read from the drawing XML.

    openpyxl hands back its own default width and height on reload, so the saved package is
    the only place the real size can be read.
    """
    approved = {
        (round(w, 2), round(h, 2))
        for w, h in (st.CHART_WIDE, st.CHART_STANDARD, st.CHART_COMPACT)
    }
    found = set()
    with zipfile.ZipFile(rebuilt) as archive:
        for name in archive.namelist():
            if name.startswith("xl/drawings/drawing"):
                xml = archive.read(name).decode("utf-8")
                for match in re.finditer(r'<ext cx="(\d+)" cy="(\d+)"', xml):
                    found.add((
                        round(int(match.group(1)) / 360000, 2),
                        round(int(match.group(2)) / 360000, 2),
                    ))
    assert found, "no chart drawings in the package"
    assert found <= approved, sorted(found - approved)
    for width, height in found:
        assert width >= vx.MIN_CHART_WIDTH and height >= vx.MIN_CHART_HEIGHT, (width, height)

    wb = load_workbook(rebuilt)
    styled = [
        name for name in VISIBLE_SHEETS for chart in wb[name]._charts
        if chart.style is not None
    ]
    wb.close()
    assert not styled, styled


# ---------------------------------------------------------------------------
# Charts -- the defect that left five of them empty in Excel
# ---------------------------------------------------------------------------
# Two independent causes, both invisible to "openpyxl created a chart object":
#   1. every waterfall referenced the presentation sheet while its coordinates belonged to
#      Data_Bridge, so the ranges pointed at empty cells on the wrong sheet;
#   2. every chart carried plotVisOnly=1 while its source sat on a hidden Data_* sheet, which
#      tells Excel to plot visible cells only.
def test_every_chart_reference_resolves_to_populated_numeric_data(rebuilt: Path) -> None:
    wb = load_workbook(rebuilt)
    specs = vx._chart_specs(rebuilt)
    assert len(specs) == vx.EXPECTED_CHARTS, len(specs)
    for spec in specs:
        name = spec["part"].split("/")[-1] + " " + spec["title"][:40]
        assert spec["title"].strip(), name
        assert spec["values"], name + " has no series"
        assert spec["category_is_text"], name + " uses a numRef for its categories"
        lengths = set()
        for reference in spec["categories"]:
            values = vx._resolve(wb, reference)
            assert values is not None, name + " category " + reference
            assert any(v is not None for v in values), name + " category range is empty"
            lengths.add(len(values))
        for reference in spec["values"]:
            values = vx._resolve(wb, reference)
            assert values is not None, name + " series " + reference
            assert any(isinstance(v, (int, float)) for v in values), name + " series is empty"
            lengths.add(len(values))
        assert len(lengths) == 1, name + " category/value lengths differ: " + str(lengths)
    wb.close()


def test_charts_reading_hidden_sheets_plot_hidden_data(rebuilt: Path) -> None:
    """plotVisOnly must be 0, or Excel renders every chart on a hidden source empty."""
    wb = load_workbook(rebuilt)
    hidden = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
    wb.close()
    for spec in vx._chart_specs(rebuilt):
        sources = {
            ref.split("!")[0].strip("'") for ref in spec["categories"] + spec["values"]
        }
        if sources & hidden:
            assert spec["plot_visible_only"] == "0", (
                spec["part"] + " sources " + str(sources & hidden)
                + " but plotVisOnly=" + str(spec["plot_visible_only"])
            )


def test_no_chart_sources_a_presentation_sheet(rebuilt: Path) -> None:
    """Charts read the dedicated chart-data layer, never a presentation cell or a formula.

    This is the check that would have caught the empty waterfalls: they referenced
    'Executive Summary' and 'Budget Bridge' rather than the sheet holding their data.
    """
    for spec in vx._chart_specs(rebuilt):
        for reference in spec["categories"] + spec["values"]:
            sheet = reference.split("!")[0].strip("'")
            assert sheet not in VISIBLE_SHEETS, (
                spec["part"] + " reads presentation sheet " + sheet
            )
            assert sheet == "Chart_Data", spec["part"] + " reads " + sheet


def test_key_charts_tie_to_the_marts(rebuilt: Path, marts: dict) -> None:
    """A chart that renders is not the same as a chart that is right."""
    wb = load_workbook(rebuilt)
    specs = {s["title"]: s for s in vx._chart_specs(rebuilt)}

    def series(fragment: str, index: int):
        for title, spec in specs.items():
            if fragment in title:
                return vx._resolve(wb, spec["values"][index])
        raise AssertionError("no chart titled ~" + fragment)

    variance = marts["fct_management_variance"].set_index("metric")
    anchors = [v for v in series("Dec-2026 Exit ARR: Board Budget", 1)
               if isinstance(v, (int, float))]
    assert len(anchors) == 2, anchors
    assert vx._close(anchors[0], variance.loc["exit_arr", "budget_amount"])
    assert vx._close(anchors[-1], variance.loc["exit_arr", "base_amount"])

    monthly = marts["fct_scenario_monthly"]
    expected = [
        float(monthly[(monthly["scenario"] == name)
                      & (monthly["month_end_date"] == ed.FY2026_END)]["ending_arr"].iloc[0])
        for name in ed.SCENARIOS
    ]
    assert all(vx._close(a, b) for a, b in zip(series("by operating scenario", 0), expected))

    policy = marts["fct_cash_runway_policy"].set_index("path")
    runway = series("Board-policy runway vs the 24-month floor", 0)
    for value, key in zip(runway, ("Bear", "Base", "Bull", "Base_FullClose")):
        assert vx._close(value, policy.loc[key, "policy_runway_months"], 1e-6), key
    assert all(vx._close(v, 24.0)
               for v in series("Board-policy runway vs the 24-month floor", 1))

    diagnosis = ed.gtm_constraint_by_segment(marts).set_index("segment")
    for index, column in enumerate((
        "h2_capacity_supported_arr", "h2_pipeline_supported_arr",
        "h2_constrained_new_logo_arr",
    )):
        plotted = series("New Logo ARR by segment", index)
        for value, segment in zip(plotted, st.SEGMENT_ORDER):
            assert vx._close(value, diagnosis.loc[segment, column], 1e-4), segment + column
    wb.close()


def test_row_heights_are_explicit_and_compact(rebuilt: Path) -> None:
    """No row is left at Excel's default, and the body height is the compact one."""
    wb = load_workbook(rebuilt)
    for name in VISIBLE_SHEETS:
        ws = wb[name]
        assert ws.sheet_format.defaultRowHeight == st.R.body, name
        unset = [r for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].height is None]
        assert not unset, name + " rows " + str(unset[:5])
    wb.close()
